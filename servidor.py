# -*- coding: utf-8 -*-
import os, re, shutil, urllib.parse, sqlite3
from flask import Flask, request, render_template, jsonify, session, redirect, url_for, send_file
import pandas as pd
import secrets
import json
from datetime import datetime, timedelta
from pathlib import Path
import socket
from io import BytesIO
from collections import Counter

from procesador import procesar_fichadas, aplanar_registros_por_tramo
from departamentos import clave_canonica, nombre_visible
from pdf_generator import LEGAJOS_EXCLUIR_INFORMES

app = Flask(__name__)
app.secret_key   = os.environ.get("SECRET_KEY",      "cm_horas_secret_2026")
SUPERVISOR_PASS  = os.environ.get("SUPERVISOR_PASS",  "cm2026")
FIRMA_SUPERVISOR = os.environ.get("FIRMA_SUPERVISOR", "CM - Carola Martin")
# URL base para links de WhatsApp — si está vacío usa el host del request
WA_BASE_URL = os.environ.get("WA_BASE_URL", "")
# Clave para disparar el reporte semanal de francos desde un cron externo
# (ver /tareas/reporte-francos-semanal) — Scheduled Tasks de PythonAnywhere
# requiere plan pago; esto es la alternativa gratuita. Sin configurar,
# la ruta queda deshabilitada (falla cerrado).
REPORTE_FRANCOS_TOKEN = os.environ.get("REPORTE_FRANCOS_TOKEN", "")

SESION_FILE    = Path("sesion.json")
CONFIRM_DIR    = Path("confirmaciones")
SEMANAS_DIR    = Path("semanas")
PERIODOS_DIR   = Path("periodos")
DATOS_DIR      = Path("datos")
DB_FILE        = DATOS_DIR / "cierres.db"
TELEFONOS_FILE = Path("recursos/telefonos.json")

# Prefijos de área por legajo (excepciones a la regla general)
_AREA_CODES = {100: "343", 141: "3435", 151: "353"}
_AREA_DEFAULT = "3437"

# Apellidos compuestos (formato fichadas: "APELLIDO NOMBRE"). Sin esta lista,
# el primer nombre de pila se confunde con la segunda palabra del apellido.
_APELLIDOS_COMPUESTOS = ["RUIZ DIAZ", "SAINT PAUL"]

_DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

def _dia_semana(fecha_str):
    try:
        dt = datetime.strptime(str(fecha_str)[:10], "%Y-%m-%d")
        return _DIAS_ES[dt.weekday()]
    except Exception:
        return ""

@app.template_filter("dia_semana")
def filter_dia_semana(fecha_str):
    return _dia_semana(fecha_str)


# ═══════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════
def _autenticado():
    return session.get("auth") is True

def _requiere_auth():
    return redirect(url_for("login", next=request.path))


# ═══════════════════════════════════════════════
# PERSISTENCIA
# ═══════════════════════════════════════════════
def _cargar_sesion():
    """Lee sesion.json (período activo en memoria). Retorna {} si no existe o está corrupto."""
    if SESION_FILE.exists():
        try:
            return json.loads(SESION_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _guardar_sesion(s):
    """Persiste el período activo en sesion.json. No commitear — contiene datos personales."""
    SESION_FILE.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding="utf-8")

def _cargar_metadata():
    f = SEMANAS_DIR / "metadata.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"semana_actual": 0, "semanas": []}

def _guardar_metadata(m):
    SEMANAS_DIR.mkdir(exist_ok=True)
    (SEMANAS_DIR / "metadata.json").write_text(
        json.dumps(m, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def _get_db():
    DATOS_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row
    return conn

def _init_db():
    """Crea las tablas si no existen y aplica migraciones de columnas nuevas via ALTER TABLE.
    Se llama al arrancar el servidor. Las migraciones usan try/except para ser idempotentes."""
    with _get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS periodos (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                cerrado_en  TEXT,
                semana_desde INTEGER,
                semana_hasta INTEGER,
                archivo     TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS periodo_empleados (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                periodo_id  INTEGER REFERENCES periodos(id),
                legajo      TEXT,
                nombre      TEXT,
                departamento TEXT,
                ot50        TEXT,
                ot100       TEXT,
                comidas     INTEGER DEFAULT 0,
                francos     INTEGER DEFAULT 0,
                tardanzas   INTEGER DEFAULT 0,
                semanas     TEXT,
                confirmado  INTEGER DEFAULT 0
            )
        """)
        # Migración: agregar columnas de fechas si no existen
        for col in (
            "fecha_desde TEXT DEFAULT ''",
            "fecha_hasta TEXT DEFAULT ''",
            "estado TEXT DEFAULT 'ACTIVO'",
            "fecha_anulacion TEXT DEFAULT ''",
            "motivo_anulacion TEXT DEFAULT ''",
            "usuario_anulacion TEXT DEFAULT ''",
        ):
            try:
                conn.execute(f"ALTER TABLE periodos ADD COLUMN {col}")
            except Exception:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_tomados (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                legajo         TEXT NOT NULL,
                nombre         TEXT NOT NULL,
                tipo           TEXT NOT NULL DEFAULT 'RANGO',
                fecha_desde    TEXT DEFAULT '',
                fecha_hasta    TEXT DEFAULT '',
                fechas_sueltas TEXT DEFAULT '[]',
                dias           INTEGER NOT NULL DEFAULT 1,
                estado         TEXT DEFAULT 'Aprobado',
                observaciones  TEXT DEFAULT '',
                cargado_en     TEXT DEFAULT ''
            )
        """)
        for col in (
            "fecha_emision TEXT DEFAULT ''",
            "autorizado_por TEXT DEFAULT ''",
            "fecha_anulacion TEXT DEFAULT ''",
            "motivo_anulacion TEXT DEFAULT ''",
            "usuario_anulacion TEXT DEFAULT ''",
        ):
            try:
                conn.execute(f"ALTER TABLE francos_tomados ADD COLUMN {col}")
            except Exception:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_saldo_inicial (
                legajo     TEXT PRIMARY KEY,
                nombre     TEXT NOT NULL,
                saldo      INTEGER NOT NULL DEFAULT 0,
                nota       TEXT DEFAULT '',
                cargado_en TEXT DEFAULT ''
            )
        """)
        # Columnas para el cierre mensual (no alteran datos existentes)
        for col in (
            "tomados_al_corte INTEGER DEFAULT 0",
            "gen_extra_al_corte INTEGER DEFAULT 0",
            "fecha_corte TEXT DEFAULT '2026-05-21'",
        ):
            try:
                conn.execute(f"ALTER TABLE francos_saldo_inicial ADD COLUMN {col}")
            except Exception:
                pass
        # Columna para guardar snapshot de saldo anterior (reversión al anular)
        try:
            conn.execute("ALTER TABLE periodos ADD COLUMN saldo_anterior TEXT DEFAULT '{}'")
        except Exception:
            pass
        # Marca/log de corrección retroactiva del snapshot de francos del cierre
        try:
            conn.execute("ALTER TABLE periodos ADD COLUMN francos_corregido_en TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE cierres_francos ADD COLUMN saldo_anterior TEXT DEFAULT '{}'")
        except Exception:
            pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_generados (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                legajo       TEXT NOT NULL,
                nombre       TEXT NOT NULL,
                departamento TEXT NOT NULL DEFAULT '',
                descripcion  TEXT NOT NULL DEFAULT '',
                dias         INTEGER NOT NULL DEFAULT 1,
                cargado_en   TEXT DEFAULT ''
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_semana_manual (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                legajo       TEXT NOT NULL,
                nombre       TEXT NOT NULL,
                departamento TEXT NOT NULL DEFAULT '',
                semana_num   INTEGER NOT NULL,
                mes          TEXT NOT NULL DEFAULT '',
                dias         INTEGER NOT NULL DEFAULT 0,
                guardado_en  TEXT DEFAULT '',
                UNIQUE(legajo, semana_num, mes)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_cierre_detalle (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                periodo_id    INTEGER NOT NULL,
                legajo        TEXT NOT NULL,
                nombre        TEXT NOT NULL,
                departamento  TEXT NOT NULL DEFAULT '',
                tipo          TEXT NOT NULL DEFAULT '',
                fecha_desde   TEXT DEFAULT '',
                fecha_hasta   TEXT DEFAULT '',
                fechas_sueltas TEXT DEFAULT '[]',
                dias          INTEGER NOT NULL DEFAULT 0,
                estado        TEXT DEFAULT '',
                fecha_emision TEXT DEFAULT '',
                autorizado_por TEXT DEFAULT '',
                observaciones TEXT DEFAULT '',
                francos_tomados_id INTEGER
            )
        """)
        # Referencia al francos_tomados.id de origen, para poder revertir el
        # estado 'Cerrado' sin ambigüedad al anular un cierre (ver
        # _revertir_estado_francos_cierre). NULL en filas históricas previas
        # a esta columna -- se usa un fallback por tupla de campos para esas.
        try:
            conn.execute("ALTER TABLE francos_cierre_detalle ADD COLUMN francos_tomados_id INTEGER")
        except Exception:
            pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_semana_parcial (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                legajo       TEXT NOT NULL,
                nombre       TEXT NOT NULL,
                departamento TEXT NOT NULL DEFAULT '',
                semana_num   INTEGER NOT NULL,
                dias         INTEGER NOT NULL DEFAULT 0,
                guardado_en  TEXT DEFAULT ''
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS supervisores (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre        TEXT NOT NULL,
                email         TEXT NOT NULL,
                departamentos TEXT NOT NULL DEFAULT '[]',
                activo        INTEGER NOT NULL DEFAULT 1,
                cargado_en    TEXT DEFAULT ''
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS empleados_extra (
                legajo        TEXT PRIMARY KEY,
                nombre        TEXT NOT NULL,
                departamento  TEXT NOT NULL DEFAULT '',
                activo        INTEGER NOT NULL DEFAULT 1,
                cargado_en    TEXT DEFAULT ''
            )
        """)
        # Gatti y Mancioni se administran como empleados sin fichadas del
        # departamento Ingenieros. Si ya existen (por ejemplo en producción),
        # se corrigen sus nombres y su asignación actual según la nómina.
        # Los registros históricos de periodo_empleados no se modifican.
        for legajo, nombre in (("100", "MANCIONI, Martin"), ("101", "GATTI, Marcelo")):
            conn.execute("""
                INSERT INTO empleados_extra
                    (legajo, nombre, departamento, activo, cargado_en)
                VALUES (?, ?, 'Ingenieros', 1, CURRENT_TIMESTAMP)
                ON CONFLICT(legajo) DO UPDATE SET
                    nombre = excluded.nombre,
                    departamento = 'Ingenieros',
                    activo = 1
            """, (legajo, nombre))
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cierres_francos (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                cerrado_en    TEXT NOT NULL,
                departamento  TEXT NOT NULL,
                fecha_hasta   TEXT NOT NULL,
                total_dias    INTEGER DEFAULT 0,
                estado        TEXT NOT NULL DEFAULT 'ACTIVO'
            )
        """)
        # Relación auditable entre cada movimiento manual y el cierre que lo
        # incorporó. Las migraciones son idempotentes para bases existentes.
        for tabla in ("francos_tomados", "francos_generados", "francos_semana_manual"):
            try:
                conn.execute(f"ALTER TABLE {tabla} ADD COLUMN cierre_francos_id INTEGER")
            except sqlite3.OperationalError:
                pass
            conn.execute(
                f"CREATE INDEX IF NOT EXISTS idx_{tabla}_cierre_francos "
                f"ON {tabla}(cierre_francos_id)"
            )
        for tabla, columna in (
            ("francos_tomados", "estado_antes_cierre TEXT DEFAULT ''"),
            ("cierres_francos", "base_anterior TEXT DEFAULT '{}'"),
            ("cierres_francos", "fecha_anulacion TEXT DEFAULT ''"),
            ("cierres_francos", "motivo_anulacion TEXT DEFAULT ''"),
            ("cierres_francos", "usuario_anulacion TEXT DEFAULT ''"),
        ):
            try:
                conn.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna}")
            except sqlite3.OperationalError:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS francos_anulaciones_cerrados (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                francos_tomados_id  INTEGER NOT NULL,
                legajo              TEXT NOT NULL,
                nombre              TEXT NOT NULL,
                departamento        TEXT NOT NULL DEFAULT '',
                tipo                TEXT NOT NULL,
                fecha_desde         TEXT DEFAULT '',
                fecha_hasta         TEXT DEFAULT '',
                fechas_sueltas      TEXT DEFAULT '[]',
                dias                INTEGER NOT NULL,
                motivo              TEXT NOT NULL,
                usuario             TEXT DEFAULT '',
                anulado_en          TEXT NOT NULL,
                periodo_origen_id   INTEGER,
                periodo_aplicado_id INTEGER,
                aplicado_en         TEXT DEFAULT ''
            )
        """)
        conn.commit()
    # Importar JSON viejos si los hay
    if PERIODOS_DIR.exists():
        with _get_db() as conn:
            ya = {r[0] for r in conn.execute("SELECT archivo FROM periodos").fetchall()}
            for f in sorted(PERIODOS_DIR.glob("periodo_*.json")):
                if f.name in ya:
                    continue
                try:
                    d = json.loads(f.read_text(encoding="utf-8"))
                    emps = d.get("empleados", d.get("confirmaciones", []))
                    sems = d.get("semanas", [])
                    cur = conn.execute(
                        "INSERT INTO periodos (cerrado_en, semana_desde, semana_hasta, archivo) VALUES (?,?,?,?)",
                        (d.get("cerrado_en",""), min(sems,default=0), max(sems,default=0), f.name)
                    )
                    pid = cur.lastrowid
                    for e in emps:
                        conn.execute(
                            "INSERT INTO periodo_empleados (periodo_id,legajo,nombre,departamento,ot50,ot100,comidas,francos,tardanzas,semanas,confirmado) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                            (pid, str(e.get("legajo","")), e.get("nombre",""), e.get("departamento",""),
                             e.get("ot50","0h"), e.get("ot100","0h"),
                             e.get("comidas",0), e.get("francos",0), e.get("tardanzas",0),
                             json.dumps(e.get("semanas",[])),
                             1 if e.get("confirmado") else 0)
                        )
                    conn.commit()
                except Exception:
                    pass

def _guardar_semana_csv(n, df):
    SEMANAS_DIR.mkdir(exist_ok=True)
    df.to_csv(SEMANAS_DIR / f"semana_{n}.csv", index=False, encoding="utf-8")

def _cargar_semana_csv(n):
    f = SEMANAS_DIR / f"semana_{n}.csv"
    return pd.read_csv(f, encoding="utf-8") if f.exists() else None

def _archivos_confirmacion():
    """Confirmaciones activas + archivadas, excluyendo las de cierres ANULADOS.

    Sin este filtro, una confirmación archivada de un período anulado puede
    ser "readoptada" por _resolver_semana_confirmacion (que empareja por
    fecha cuando no encuentra el número de semana exacto) en el próximo
    cierre que caiga en el mismo rango de fechas, resucitando totales viejos
    en vez de usar los datos frescos del reproceso."""
    CONFIRM_DIR.mkdir(exist_ok=True)
    archivos = sorted(CONFIRM_DIR.rglob("*.json"), reverse=True)
    try:
        with _get_db() as conn:
            carpetas_anuladas = {
                Path(row["archivo"]).stem
                for row in conn.execute(
                    "SELECT archivo FROM periodos WHERE COALESCE(estado,'ACTIVO')='ANULADO' AND archivo IS NOT NULL"
                )
            }
    except Exception:
        carpetas_anuladas = set()
    if not carpetas_anuladas:
        return archivos
    return [f for f in archivos if f.parent.name not in carpetas_anuladas]

def _parse_fecha(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _clave_confirmacion(data):
    return (
        _normalizar_departamento_web(data.get("departamento", "") or "Todos"),
        str(data.get("legajo", "")),
        data.get("semana", 0),
    )

def _fechas_confirmacion(data):
    fechas = []
    for d in data.get("dias", []) or []:
        f = _parse_fecha(d.get("fecha", ""))
        if f:
            fechas.append(f)
    return fechas

def _resolver_semana_confirmacion(data, meta):
    sem_conf = data.get("semana", 0)
    depto_conf = _normalizar_departamento_web(data.get("departamento", "") or "Todos")
    for s in meta.get("semanas", []):
        if s.get("numero") == sem_conf and _normalizar_departamento_web(s.get("departamento", "") or "Todos") == depto_conf:
            return sem_conf, s.get("num_depto", sem_conf)

    fechas = _fechas_confirmacion(data)
    if not fechas:
        return sem_conf, data.get("semana_depto", sem_conf)
    desde_conf, hasta_conf = min(fechas), max(fechas)
    for s in meta.get("semanas", []):
        if _normalizar_departamento_web(s.get("departamento", "") or "Todos") != depto_conf:
            continue
        desde_sem = _parse_fecha(s.get("fecha_desde", ""))
        hasta_sem = _parse_fecha(s.get("fecha_hasta", ""))
        if desde_sem and hasta_sem and desde_conf <= hasta_sem and hasta_conf >= desde_sem:
            return s.get("numero", sem_conf), s.get("num_depto", s.get("numero", sem_conf))
    return sem_conf, data.get("semana_depto", sem_conf)

def _legajo_key(x):
    leg = str(x.get("legajo", ""))
    try:
        return int(leg)
    except ValueError:
        return 0

def _score_confirmacion(data):
    dias = data.get("dias", []) or []
    descs = sum(1 for d in dias if str(d.get("descripcion", "")).strip())
    confirmado_en = data.get("confirmado_en", "") or ""
    return (descs, confirmado_en)

def _semana_meta_por_numero(meta, semana, departamento=None):
    departamento = _normalizar_departamento_web(departamento)
    return next((
        s for s in meta.get("semanas", [])
        if s.get("numero") == semana
        and (not departamento or _normalizar_departamento_web(s.get("departamento", "") or "Todos") == departamento)
    ), None)

def _ajustar_confirmacion_a_semana(data, meta, semana, departamento=None):
    sem = _semana_meta_por_numero(meta, semana, departamento)
    if not sem:
        return data
    desde = _parse_fecha(sem.get("fecha_desde", ""))
    hasta = _parse_fecha(sem.get("fecha_hasta", ""))
    if not desde or not hasta:
        return data

    dias = []
    for dia in data.get("dias", []) or []:
        fecha = _parse_fecha(dia.get("fecha", ""))
        if fecha and desde <= fecha <= hasta:
            dias.append(dia)

    if len(dias) == len(data.get("dias", []) or []):
        return data

    data = dict(data)
    data["dias"] = dias
    ot50 = ot100 = timedelta(0)
    comidas = francos = tardanzas = 0
    for dia in dias:
        ot50 += _parse_td(dia.get("ot50", "00:00:00"))
        ot100 += _parse_td(dia.get("ot100", "00:00:00"))
        comidas += int(dia.get("comida", 0))
        francos += int(dia.get("franco", 0))
        tardanzas += int(dia.get("tardanzas", dia.get("tarde", 0)))
    data["totales"] = {
        "ot50": _fmt_hm(ot50),
        "ot100": _fmt_hm(ot100),
        "comidas": comidas,
        "francos": francos,
        "tardanzas": tardanzas,
    }
    return data

def _leer_historial(semana=None, departamento=None):
    CONFIRM_DIR.mkdir(exist_ok=True)
    # Solo leer confirmaciones del período activo (semanas en metadata)
    departamento = _normalizar_departamento_web(departamento)
    meta = _cargar_metadata()
    semanas_activas = {s.get("numero") for s in meta.get("semanas", [])}
    items = []
    vistos = set()  # departamento + legajo + semana ya cubiertos por archivos
    mejores = {}
    for f in _archivos_confirmacion():
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            depto_conf = _normalizar_departamento_web(data.get("departamento", "") or "Todos")
            sem_conf, sem_depto = _resolver_semana_confirmacion(data, meta)
            data["semana_original"] = data.get("semana", sem_conf)
            data["semana"] = sem_conf
            data["semana_depto"] = sem_depto
            data = _ajustar_confirmacion_a_semana(data, meta, sem_conf, depto_conf)
            # Ignorar confirmaciones de períodos anteriores
            if semanas_activas and sem_conf not in semanas_activas:
                continue
            if departamento and depto_conf != departamento:
                continue
            if semana is None or sem_depto == semana:
                key = _clave_confirmacion(data)
                actual = mejores.get(key)
                if actual is None or _score_confirmacion(data) > _score_confirmacion(actual):
                    mejores[key] = data
        except Exception:
            continue
    items.extend(sorted(mejores.values(), key=_legajo_key))
    vistos.update(mejores.keys())
    # Fallback: empleados confirmados en _sesion sin archivo en confirmaciones/
    for d in _sesion.values():
        if not d.get("confirmado"):
            continue
        depto_conf = _normalizar_departamento_web(d.get("departamento", "") or "Todos")
        if departamento and depto_conf != departamento:
            continue
        key = _clave_confirmacion(d)
        if key in vistos:
            continue
        if semana is not None and d.get("semana_depto", d.get("semana")) != semana:
            continue
        items.append({
            "legajo":       d["legajo"],
            "nombre":       d["nombre"],
            "departamento": d.get("departamento", ""),
            "confirmado_en": d.get("confirmado_en"),
            "semana":       d.get("semana", 0),
            "semana_depto": d.get("semana_depto", d.get("semana", 0)),
            "totales":      d.get("totales", {}),
            "excluido_ot":  d.get("excluido_ot", False),
            "liquida_ot":   d.get("liquida_ot", True),
            "observacion_liquidacion": d.get("observacion_liquidacion", ""),
            "dias": [
                {"fecha": x["fecha"], "ot50": x["ot50"], "ot100": x["ot100"],
                 "franco": x.get("franco", 0), "comida": x.get("comida", 0),
                 "tipo_dia": x.get("tipo_dia", "normal"),
                 "dia_semana": x.get("dia_semana") or _dia_semana(x["fecha"]),
                 "descripcion": x.get("descripcion", ""),
                 "tramos": x.get("tramos", [])}
                for x in d.get("dias", []) if x.get("tiene_ot")
            ],
        })
        vistos.add(key)
    return sorted(items, key=_legajo_key)

def _cargar_telefonos():
    if TELEFONOS_FILE.exists():
        try:
            return json.loads(TELEFONOS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _guardar_telefonos(t):
    Path("recursos").mkdir(exist_ok=True)
    TELEFONOS_FILE.write_text(json.dumps(t, ensure_ascii=False, indent=2), encoding="utf-8")

def _normalizar_valor_excel(v):
    texto = str(v).strip()
    if texto.lower() in ("nan", "none", ""):
        return ""
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto

def _primer_nombre(nombre):
    """Extrae el primer nombre de pila de 'APELLIDO NOMBRE' (formato fichadas)."""
    nombre = nombre.strip()
    for apellido in _APELLIDOS_COMPUESTOS:
        if nombre.upper().startswith(apellido + " "):
            resto = nombre[len(apellido):].split()
            return resto[0] if resto else nombre.split()[0]
    partes = nombre.split()
    return partes[1] if len(partes) > 1 else partes[0]

def _wa_url(legajo, nombre, url, totales=None, dias=None):
    tel = _cargar_telefonos()
    phone = re.sub(r'\D', '', str(tel.get(str(legajo), "")))
    if not phone:
        return ""
    area = _AREA_CODES.get(int(legajo), _AREA_DEFAULT)
    if WA_BASE_URL:
        from urllib.parse import urlparse
        path = urlparse(url).path
        url  = WA_BASE_URL.rstrip("/") + path
    nombre_corto = _primer_nombre(nombre)
    if dias or totales:
        lineas = [f"Hola {nombre_corto}, tus horas extras registradas son:"]
        if dias:
            for d in dias:
                if not d.get("tiene_ot") and not d.get("ot50") and not d.get("ot100") and not d.get("franco") and not d.get("comida"):
                    continue
                if d.get("ot50") == "00:00:00" and d.get("ot100") == "00:00:00" and not d.get("franco") and not d.get("comida"):
                    continue
                dia_nombre = d.get("dia_semana") or _dia_semana(d.get("fecha", ""))
                fecha_fmt  = d.get("fecha_fmt") or d.get("fecha", "")[5:]
                partes = []
                if d.get("ot50")  and d["ot50"]  != "00:00:00": partes.append(f"{d['ot50'][:5]} (50%)")
                if d.get("ot100") and d["ot100"] != "00:00:00": partes.append(f"{d['ot100'][:5]} (100%)")
                if d.get("franco"):  partes.append("Franco")
                if d.get("comida"):  partes.append("Comida")
                if partes:
                    prefijo = f"{dia_nombre} {fecha_fmt}" if dia_nombre else fecha_fmt
                    lineas.append(f"• {prefijo}: {', '.join(partes)}")
        if len(lineas) == 1 and totales:
            ot50  = totales.get("ot50",  "0h")
            ot100 = totales.get("ot100", "0h")
            fra   = totales.get("francos",  0)
            com   = totales.get("comidas",  0)
            tar   = totales.get("tardanzas", 0)
            if ot50  not in ("0h", "0:00", ""): lineas.append(f"• OT 50%: {ot50}")
            if ot100 not in ("0h", "0:00", ""): lineas.append(f"• OT 100%: {ot100}")
            if fra: lineas.append(f"• Francos: {fra}")
            if com: lineas.append(f"• Comidas: {com}")
            if tar: lineas.append(f"• Tardanzas: {tar}")
        if len(lineas) == 1:
            lineas.append("• Sin horas extras este período")
        lineas.append(f"Podés verlas y confirmarlas en: {url}")
        texto = "\n".join(lineas)
    else:
        texto = f"Hola {nombre_corto}, confirmá tus horas extras en este link: {url}"
    msg = urllib.parse.quote(texto)
    return f"https://wa.me/549{area}{phone}?text={msg}"

def _normalizar_cargado_en(valor):
    """Normaliza una fecha/hora de corte (p.ej. periodos.cerrado_en, en
    formato ISO con 'T' y posibles microsegundos) reemplazando únicamente el
    separador 'T' por un espacio, SIN truncar la precisión — para poder
    compararla como string con francos_tomados.cargado_en
    ('YYYY-MM-DD HH:MM:SS') preservando los microsegundos cuando existan.
    Dos cortes dentro del mismo segundo con microsegundos distintos no deben
    perder esa diferencia (de lo contrario la ventana incremental podría
    calcularse mal)."""
    return (valor or "").replace("T", " ")


def _fmt_fecha_corte_pdf(fecha_corte):
    """Formatea una fecha de corte ('YYYY-MM-DD' o 'YYYY-MM-DD HH:MM:SS[.ffffff]')
    a 'DD/MM/YYYY' para mostrar en PDFs. Devuelve "" si no se puede parsear
    (p.ej. cierre sin fecha de corte registrada)."""
    try:
        return datetime.strptime((fecha_corte or "")[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return ""


# Expresión SQL que normaliza francos_tomados.cargado_en de la misma forma
# que _normalizar_cargado_en (separador 'T' -> espacio, sin truncar
# microsegundos), recortando además espacios sobrantes. Se usa de forma
# IDÉNTICA para detectar valores vacíos/legacy y para los límites inferior y
# superior de la ventana incremental, así "YYYY-MM-DDTHH:MM:SS[.ffffff]" y
# "YYYY-MM-DD HH:MM:SS[.ffffff]" se comparan de forma equivalente.
_SQL_CARGADO_EN_NORM = "REPLACE(TRIM(COALESCE(cargado_en,'')), 'T', ' ')"
# Igual que _SQL_CARGADO_EN_NORM pero recortado al día (los cierres
# manuales de francos -- cierres_francos -- usan un corte de solo fecha,
# sin hora; sin este recorte, cualquier cargado_en con hora >00:00:00 del
# día del corte quedaría excluido por comparación de string).
_SQL_CARGADO_EN_DIA = "SUBSTR(REPLACE(TRIM(COALESCE(cargado_en,'')), 'T', ' '), 1, 10)"


def _fecha_corte_anterior(conn, pid, departamento, fecha_corte_actual):
    """Busca, entre los cierres activos (estado distinto de 'ANULADO') del
    mismo departamento y distintos de `pid`, el de fecha/hora de corte
    (periodos.cerrado_en) "inmediatamente anterior" a `fecha_corte_actual`:
    el de mayor corte estrictamente menor que `fecha_corte_actual`.

    Si dos cierres del mismo departamento tienen exactamente el mismo
    cerrado_en (caso límite, p.ej. dos cierres en el mismo microsegundo), se
    usa el id como desempate secundario: el de menor id se considera
    "anterior" al de mayor id. El id NUNCA es el criterio principal — solo
    desempata cortes idénticos.

    Devuelve esa fecha de corte normalizada ('YYYY-MM-DD HH:MM:SS[.ffffff]'),
    o "" si no existe un cierre anterior válido — en cuyo caso la ventana del
    cierre actual no tiene límite inferior."""
    depto_canon = _normalizar_departamento_web(departamento)
    rows = conn.execute("""
        SELECT p.id AS pid, p.cerrado_en AS cerrado_en,
               (SELECT pe.departamento FROM periodo_empleados pe
                WHERE pe.periodo_id = p.id LIMIT 1) AS departamento
        FROM periodos p
        WHERE p.id != ?
          AND COALESCE(p.estado, 'ACTIVO') != 'ANULADO'
    """, (pid,)).fetchall()
    candidatos = []
    for r in rows:
        if _normalizar_departamento_web(r["departamento"] or "") != depto_canon:
            continue
        corte = _normalizar_cargado_en(r["cerrado_en"])
        if not corte:
            continue
        if corte < fecha_corte_actual or (corte == fecha_corte_actual and r["pid"] < pid):
            candidatos.append((corte, r["pid"]))
    if not candidatos:
        return ""
    candidatos.sort()
    return candidatos[-1][0]


def _seleccionar_francos_ventana_cierre(conn, pid, legajos, departamento, fecha_corte):
    """Fuente única de verdad de la ventana incremental de un cierre:

        fecha_corte_anterior < cargado_en <= fecha_corte

    donde fecha_corte_anterior es el corte del cierre activo (no anulado)
    anterior del mismo departamento (ver _fecha_corte_anterior), sin límite
    inferior si no existe uno. cargado_en se compara normalizado vía
    _SQL_CARGADO_EN_NORM (separador 'T' -> espacio, sin truncar
    microsegundos), con la MISMA expresión para detectar vacíos y para
    ambos límites. La usan tanto _snapshot_francos_cierre como
    /admin/corregir-francos-cierre, para que ambos produzcan exactamente el
    mismo conjunto sin duplicar la consulta ni la lógica de ventana.

    Si `legajos` es None o está vacío, NUNCA se ejecuta una selección global:
    se devuelve (rows=[], fecha_corte_anterior, legacy=[]) sin tocar francos
    de ningún empleado — un cierre sin empleados no puede "arrastrar" francos
    ajenos.

    Los francos con cargado_en vacío/NULL/solo espacios ("legacy", sin dato
    de carga) NUNCA se asignan a una ventana por comparación lexicográfica:
    se devuelven aparte en `legacy` para revisión/tratamiento manual y no se
    incluyen en `rows`.

    Devuelve (rows, fecha_corte_anterior, legacy):
    - rows: filas (sqlite3.Row) de francos_tomados no anulados y NO cerrados
      todavía (un franco ya 'Cerrado' nunca se vuelve a incluir en un cierre
      posterior, evitando duplicarlo en francos_cierre_detalle), de los
      `legajos` dados, con cargado_en dentro de la ventana, ordenadas por
      legajo y fecha_desde.
    - fecha_corte_anterior: corte normalizado del cierre anterior, o "".
    - legacy: filas de francos_tomados no anuladas (incluye estado 'Cerrado'
      -- pueden haber sido incluidas históricamente bajo la lógica anterior y
      necesitar auditoría manual) de los mismos `legajos`, cuyo cargado_en
      está vacío/NULL/solo espacios. 'Anulado' nunca aparece aquí."""
    fecha_corte_anterior = _fecha_corte_anterior(conn, pid, departamento, fecha_corte)

    legajos_norm = [str(l) for l in (legajos or [])]
    if not legajos_norm:
        return [], fecha_corte_anterior, []

    cond_legajos = f"AND legajo IN ({','.join('?' * len(legajos_norm))})"
    params_legajos = tuple(legajos_norm)

    cond_inferior, params_inferior = "", ()
    if fecha_corte_anterior:
        cond_inferior = f"AND {_SQL_CARGADO_EN_NORM} > ?"
        params_inferior = (fecha_corte_anterior,)

    rows = conn.execute(f"""
        SELECT * FROM francos_tomados
        WHERE COALESCE(estado,'') NOT IN ('Anulado', 'Cerrado')
          AND {_SQL_CARGADO_EN_NORM} != ''
          AND {_SQL_CARGADO_EN_NORM} <= ?
          {cond_inferior}
          {cond_legajos}
        ORDER BY CAST(legajo AS INTEGER), fecha_desde
    """, (fecha_corte, *params_inferior, *params_legajos)).fetchall()

    legacy = conn.execute(f"""
        SELECT * FROM francos_tomados
        WHERE COALESCE(estado,'') != 'Anulado'
          AND {_SQL_CARGADO_EN_NORM} = ''
          {cond_legajos}
        ORDER BY CAST(legajo AS INTEGER), fecha_desde
    """, params_legajos).fetchall()

    return rows, fecha_corte_anterior, legacy


def _revertir_estado_francos_cierre(conn, pid):
    """Revierte a su estado original los francos_tomados que el cierre #pid
    marcó como 'Cerrado' al cerrarse (ver _snapshot_francos_cierre), usando
    el estado guardado en francos_cierre_detalle (capturado antes de
    cerrar, vía la fila de la SELECT previa al UPDATE).

    Sin esto, anular un cierre deja esos francos con estado='Cerrado' para
    siempre: _seleccionar_francos_ventana_cierre los excluye de por vida
    (WHERE estado NOT IN ('Anulado','Cerrado')), aunque el cierre que los
    "cerró" ya no exista. Solo toca filas que siguen en 'Cerrado' — no pisa
    cambios manuales posteriores (por ejemplo si alguien ya lo anuló a mano
    por otro motivo). Idempotente: llamarlo de nuevo sobre filas ya
    revertidas no hace nada (el WHERE exige estado='Cerrado')."""
    detalle = conn.execute(
        "SELECT * FROM francos_cierre_detalle WHERE periodo_id=?", (pid,)
    ).fetchall()
    revertidos = 0
    for d in detalle:
        francos_tomados_id = d["francos_tomados_id"] if "francos_tomados_id" in d.keys() else None
        if francos_tomados_id is not None:
            # Sin ambigüedad: apunta al francos_tomados de origen exacto.
            cur = conn.execute(
                "UPDATE francos_tomados SET estado=? WHERE id=? AND estado='Cerrado'",
                (d["estado"] or "Aprobado", francos_tomados_id)
            )
        else:
            # Filas históricas de antes de agregar la columna: fallback por
            # tupla de campos (puede no matchear si el franco fue editado
            # después de cerrarse, o dejar sin revertir en caso de duplicado
            # exacto -- ver plan de rediseño de cierre de francos).
            cur = conn.execute("""
                UPDATE francos_tomados
                SET estado = ?
                WHERE legajo=? AND tipo=? AND fecha_desde=? AND fecha_hasta=?
                  AND fechas_sueltas=? AND dias=? AND estado='Cerrado'
            """, (d["estado"] or "Aprobado", d["legajo"], d["tipo"] or "",
                  d["fecha_desde"] or "", d["fecha_hasta"] or "",
                  d["fechas_sueltas"] or "[]", d["dias"] or 0))
        revertidos += cur.rowcount
    return revertidos


def _delta_francos_cierre(conn, pid, legajos_cierre):
    """Generados y tomados que APORTA específicamente el cierre #pid, por legajo.

    A diferencia de _calcular_saldos() (saldo global "en vivo" para pantalla,
    que además suma francos_semana_parcial de semanas en curso de cualquier
    depto/mes no cerrado todavía), esto NUNCA mezcla datos de otro
    período/depto/semana en curso: solo lee lo que ya quedó grabado bajo
    este pid en periodo_empleados y francos_cierre_detalle. Se usa para
    actualizar francos_saldo_inicial al cerrar un período, para que el saldo
    grabado no dependa de qué otro parcial esté pendiente en ese momento."""
    generados = {str(r["legajo"]): (r["francos"] or 0) for r in conn.execute(
        "SELECT legajo, francos FROM periodo_empleados WHERE periodo_id=?", (pid,))}
    tomados = {r["legajo"]: (r["total"] or 0) for r in conn.execute(
        "SELECT legajo, SUM(dias) as total FROM francos_cierre_detalle "
        "WHERE periodo_id=? GROUP BY legajo", (pid,))}
    return {leg: {"generados_periodo": generados.get(leg, 0),
                  "tomados_periodo": tomados.get(leg, 0)} for leg in legajos_cierre}


def _snapshot_francos_cierre(conn, pid, fecha_corte, legajos=None, departamento=""):
    """Copia francos_tomados a francos_cierre_detalle y genera PDF.
    Incluye los francos no anulados CARGADOS en el sistema dentro de la
    ventana incremental de este cierre del departamento dado (ver
    _seleccionar_francos_ventana_cierre). Los francos sin cargado_en
    ("legacy") no se incluyen ni se modifican aquí — quedan señalados para
    revisión manual vía /admin/corregir-francos-cierre. No depende de la
    fecha tomada (fecha_desde/fecha_hasta) ni del período del cierre —
    filtra solo por legajos del departamento para no mezclar deptos."""
    if not fecha_corte:
        return
    rows, _fecha_corte_anterior, _legacy = _seleccionar_francos_ventana_cierre(
        conn, pid, legajos, departamento, fecha_corte)
    depto_visible = _nombre_departamento_visible(departamento) if departamento else ""
    for r in rows:
        conn.execute("""
            INSERT INTO francos_cierre_detalle
              (periodo_id, legajo, nombre, departamento, tipo, fecha_desde, fecha_hasta,
               fechas_sueltas, dias, estado, fecha_emision, autorizado_por, observaciones,
               francos_tomados_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (pid, r["legajo"], r["nombre"], depto_visible, r["tipo"],
              r["fecha_desde"] or "", r["fecha_hasta"] or "", r["fechas_sueltas"] or "[]",
              r["dias"], r["estado"] or "", r["fecha_emision"] or "",
              r["autorizado_por"] or "", r["observaciones"] or "", r["id"]))
        conn.execute(
            "UPDATE francos_tomados SET estado='Cerrado' "
            "WHERE id=? AND COALESCE(estado,'') NOT IN ('Anulado','Cerrado')",
            (r["id"],)
        )
    francos_list = [{**dict(r), "departamento": depto_visible} for r in rows]

    # Movimientos compensatorios: devoluciones por francos anulados de cierres
    # anteriores, todavía no incluidas en ningún cierre posterior. No suman al
    # cálculo de saldo (eso ya ocurrió al anular) -- son solo la constancia
    # de que la devolución figura en este cierre.
    devoluciones = []
    legajos_norm = [str(l) for l in (legajos or [])]
    if legajos_norm:
        ph = ",".join("?" * len(legajos_norm))
        devoluciones = conn.execute(f"""
            SELECT * FROM francos_anulaciones_cerrados
            WHERE periodo_aplicado_id IS NULL AND legajo IN ({ph})
            ORDER BY CAST(legajo AS INTEGER), anulado_en
        """, legajos_norm).fetchall()
        if devoluciones:
            ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ids = [d["id"] for d in devoluciones]
            ph_ids = ",".join("?" * len(ids))
            conn.execute(
                f"UPDATE francos_anulaciones_cerrados SET periodo_aplicado_id=?, aplicado_en=? WHERE id IN ({ph_ids})",
                (pid, ahora_str, *ids)
            )
    devoluciones_list = [dict(d) for d in devoluciones]

    _generar_pdf_francos_cierre(pid, francos_list, fecha_corte, devoluciones=devoluciones_list)
    return devoluciones_list


def _generar_pdf_francos_cierre(pid, francos, fecha_corte, reimpreso_el=None, escribir_archivo=True, devoluciones=None):
    """Genera PDF con detalle de francos tomados para el cierre.

    `fecha_corte` es la fecha/hora REAL de cierre (periodos.cerrado_en) y se
    muestra en el encabezado como "Cargados hasta el DD/MM/YYYY". Si se pasa
    `reimpreso_el`, se agrega debajo una línea "Reimpreso el DD/MM/YYYY" —
    la fecha actual NUNCA se usa como fecha de cierre.

    Devuelve los bytes del PDF (o None si falla). Si `escribir_archivo` es
    True (default), además lo guarda en reportes/francos_cierre_{pid}_{ts}.pdf
    (archivo histórico del snapshot al momento del cierre)."""
    try:
        from fpdf import FPDF
        FONTS_DIR = Path("recursos/fonts")
        FONT_REG  = FONTS_DIR / "DejaVuSans.ttf"
        FONT_BOLD = FONTS_DIR / "DejaVuSans-Bold.ttf"

        class PDF(FPDF):
            def __init__(self):
                super().__init__(orientation="L", unit="mm", format="A4")
                self.set_auto_page_break(auto=True, margin=15)
                self._unicode = False
                if FONT_REG.exists() and FONT_BOLD.exists():
                    try:
                        self.add_font("DejaVu", "",  str(FONT_REG),  uni=True)
                        self.add_font("DejaVu", "B", str(FONT_BOLD), uni=True)
                        self._unicode = True
                    except Exception:
                        pass
            def fam(self): return "DejaVu" if self._unicode else "Helvetica"
            def header(self):
                self.set_font(self.fam(), "B", 12)
                titulo = f"Francos Tomados — Cierre #{pid}"
                fecha_corte_fmt = _fmt_fecha_corte_pdf(fecha_corte)
                if fecha_corte_fmt:
                    titulo += f"   (Cargados hasta el {fecha_corte_fmt})"
                self.cell(0, 9, titulo, ln=1, align="C")
                if reimpreso_el:
                    self.set_font(self.fam(), "", 8)
                    self.set_text_color(110, 110, 110)
                    self.cell(0, 5, f"Reimpreso el {_fmt_fecha_corte_pdf(reimpreso_el)}", ln=1, align="C")
                    self.set_text_color(0, 0, 0)
                self.ln(2)
            def footer(self):
                self.set_y(-13)
                self.set_font(self.fam(), "", 7)
                self.cell(0, 8, f"Pag. {self.page_no()}  |  CM_HorasExtras", 0, 0, "C")

        pdf = PDF()
        pdf.add_page()
        f = pdf.fam()
        avail = round(pdf.w - pdf.l_margin - pdf.r_margin)
        COLS   = ["Leg.", "Nombre", "Tipo", "Fechas", "Días", "Estado", "Autorizado por", "Obs."]
        _base  = [12,    62,       15,     90,        10,     20,       38,               15]
        ANCHOS = _base[:]
        ANCHOS[3] += avail - sum(_base)  # Fechas absorbe el espacio disponible

        def cabecera():
            pdf.set_fill_color(200, 215, 240)
            pdf.set_font(f, "B", 8)
            for col, ancho in zip(COLS, ANCHOS):
                pdf.cell(ancho, 6, col, 1, 0, "C", fill=True)
            pdf.ln()

        cabecera()
        depto_act = None
        for r in francos:
            if pdf.get_y() + 7 > pdf.h - pdf.b_margin:
                pdf.add_page(); cabecera()
            dep = (r.get("departamento") or "").upper()
            if dep != depto_act:
                depto_act = dep
                pdf.set_fill_color(30, 58, 95)
                pdf.set_text_color(224, 234, 248)
                pdf.set_font(f, "B", 8)
                pdf.cell(sum(ANCHOS), 6, f"  {dep}", 1, 1, "L", fill=True)
                pdf.set_text_color(0, 0, 0)
            tipo = r.get("tipo", "")
            if tipo == "UNICO":
                fechas = r.get("fecha_desde", "")
            elif tipo == "RANGO":
                fechas = f"{r.get('fecha_desde','')} → {r.get('fecha_hasta','')}"
            else:
                try:
                    fl = json.loads(r.get("fechas_sueltas") or "[]")
                    fechas = "  ".join(f"{d[8:10]}/{d[5:7]}" for d in fl if len(d) >= 10)
                except Exception:
                    fechas = r.get("fecha_desde", "")
            pdf.set_font(f, "", 8)
            pdf.cell(ANCHOS[0], 6, str(r.get("legajo","")),        1, 0, "C")
            pdf.cell(ANCHOS[1], 6, r.get("nombre",""),              1, 0, "L")
            pdf.cell(ANCHOS[2], 6, tipo,                            1, 0, "C")
            pdf.cell(ANCHOS[3], 6, fechas,                          1, 0, "L")
            pdf.cell(ANCHOS[4], 6, str(r.get("dias", 0)),           1, 0, "C")
            pdf.cell(ANCHOS[5], 6, r.get("estado",""),              1, 0, "C")
            pdf.cell(ANCHOS[6], 6, r.get("autorizado_por","") or "",1, 0, "L")
            pdf.cell(ANCHOS[7], 6, r.get("observaciones","") or "", 1, 1, "L")

        # Fila de total general
        total_dias = sum(r.get("dias", 0) or 0 for r in francos)
        if pdf.get_y() + 8 > pdf.h - pdf.b_margin:
            pdf.add_page(); cabecera()
        pdf.set_fill_color(23, 32, 51)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(f, "B", 8)
        pdf.cell(ANCHOS[0]+ANCHOS[1]+ANCHOS[2]+ANCHOS[3], 7, "TOTAL", 1, 0, "R", fill=True)
        pdf.cell(ANCHOS[4], 7, str(total_dias), 1, 0, "C", fill=True)
        pdf.cell(sum(ANCHOS[5:]), 7, "", 1, 1, "C", fill=True)
        pdf.set_text_color(0, 0, 0)

        if devoluciones:
            pdf.ln(3)
            if pdf.get_y() + 14 > pdf.h - pdf.b_margin:
                pdf.add_page()
            pdf.set_font(f, "B", 9)
            pdf.cell(0, 7, "MOVIMIENTOS COMPENSATORIOS", ln=1)
            pdf.set_font(f, "", 8)
            total_dev = 0
            for d in devoluciones:
                if pdf.get_y() + 6 > pdf.h - pdf.b_margin:
                    pdf.add_page()
                dias_dev = d.get("dias", 0) or 0
                total_dev += dias_dev
                tipo = d.get("tipo", "")
                if tipo == "RANGO":
                    fechas = f"{d.get('fecha_desde','')} → {d.get('fecha_hasta','')}"
                elif tipo == "SUELTAS":
                    try:
                        fl = json.loads(d.get("fechas_sueltas") or "[]")
                        fechas = "  ".join(f"{x[8:10]}/{x[5:7]}" for x in fl if len(x) >= 10)
                    except Exception:
                        fechas = ""
                else:
                    fechas = d.get("fecha_desde", "")
                pdf.set_text_color(0, 120, 0)
                texto = (f"  {d.get('legajo','')} - {d.get('nombre','')}: "
                         f"Devolución por anulación de franco cerrado ({fechas}) "
                         f"+{dias_dev} — Motivo: {d.get('motivo','')}")
                pdf.multi_cell(0, 5, texto)
                pdf.set_text_color(0, 0, 0)
            pdf.set_font(f, "B", 8)
            pdf.set_text_color(0, 120, 0)
            pdf.cell(0, 6, f"  Total devolución por anulación de francos cerrados: +{total_dev}", ln=1)
            pdf.set_text_color(0, 0, 0)

        if escribir_archivo:
            Path("reportes").mkdir(exist_ok=True)
            # Solo dígitos: ':' (y 'T'/espacios) no son válidos en nombres de
            # archivo en Windows -- preserva la precisión (incl. microsegundos)
            # de fecha_corte sin separadores.
            ts = re.sub(r"[^0-9]", "", fecha_corte) if fecha_corte else datetime.now().strftime("%Y%m%d%H%M%S%f")
            pdf.output(str(Path(f"reportes/francos_cierre_{pid}_{ts}.pdf")))
        return _pdf_bytes(pdf)
    except Exception:
        return None


def _parse_td(s):
    if not s or s == "00:00:00":
        return timedelta(0)
    parts = s.split(":")
    h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
    return timedelta(hours=h, minutes=m, seconds=sec)

def _parse_hm(s):
    """'3h 45m' o '3h' → timedelta"""
    if not s or s == "0h":
        return timedelta(0)
    h = int(re.search(r'(\d+)h', s).group(1)) if re.search(r'(\d+)h', s) else 0
    m = int(re.search(r'(\d+)m', s).group(1)) if re.search(r'(\d+)m', s) else 0
    return timedelta(hours=h, minutes=m)

def _fmt_hm(td):
    total = int(td.total_seconds())
    if total <= 0:
        return "0h"
    h = total // 3600
    m = (total % 3600) // 60
    return f"{h}h {m:02d}m" if m else f"{h}h"

def _fmt_hhcmm(td):
    """Formatea timedelta a H:MM para informes de Contaduría (ej: 5:30)."""
    total_min = max(0, int(td.total_seconds() // 60))
    h, m = divmod(total_min, 60)
    return f"{h}:{m:02d}"

_MESES_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def _label_mes(fecha_str):
    try:
        d = datetime.strptime((fecha_str or "")[:10], "%Y-%m-%d")
        return f"{_MESES_ES[d.month-1]} {d.year}"
    except Exception:
        return (fecha_str or "")[:7]

def _pdf_bytes(pdf):
    data = pdf.output(dest="S")
    if isinstance(data, str):
        return data.encode("latin-1")
    return bytes(data)

def _pdf_cell_text(value):
    return "" if value is None else str(value)

def _leer_confirmaciones_cierre(periodo):
    archivo = periodo.get("archivo") or ""
    carpeta = CONFIRM_DIR / Path(archivo).stem
    items = []
    if carpeta.exists():
        for f in sorted(carpeta.glob("*.json")):
            try:
                items.append(json.loads(f.read_text(encoding="utf-8")))
            except Exception:
                continue
    # Deduplicar por (legajo, semana): si hay dos archivos para el mismo empleado y semana,
    # conservar el de mayor score (confirmado_en más reciente)
    dedup = {}
    for item in items:
        key = (str(item.get("legajo", "")), item.get("semana_depto", item.get("semana", 0)))
        prev = dedup.get(key)
        if prev is None or (item.get("confirmado_en") or "") > (prev.get("confirmado_en") or ""):
            dedup[key] = item
    return sorted(dedup.values(), key=lambda x: (
        _normalizar_departamento_web(x.get("departamento", "")),
        _legajo_key(x),
    ))

def _generar_pdf_confirmaciones_cierre(periodo, empleados, datos_csv=None):
    from pdf_generator import PDFGeneral

    confirmaciones = _leer_confirmaciones_cierre(periodo)

    # Índice: (depto_norm, legajo, semana_depto) → confirmación
    conf_idx = {}
    for c in confirmaciones:
        key = (
            _normalizar_departamento_web(c.get("departamento", "")),
            str(c.get("legajo", "")),
            int(c.get("semana_depto", c.get("semana", 0))),
        )
        prev = conf_idx.get(key)
        if prev is None or (c.get("confirmado_en") or "") > (prev.get("confirmado_en") or ""):
            conf_idx[key] = c

    pdf = PDFGeneral()
    pdf.titulo = "Confirmaciones del cierre"
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    fam = "DejaVu" if pdf._unicode else "Helvetica"

    semanas_periodo = sorted(set(s for e in empleados for s in e.get("semanas", [])))
    if not semanas_periodo:
        semanas_periodo = list(range(periodo["semana_desde"], periodo["semana_hasta"] + 1))
    cerrado = (periodo["cerrado_en"] or "")[:16].replace("T", " ")
    rango = f"{periodo['fecha_desde'] or ''} al {periodo['fecha_hasta'] or ''}".strip()
    estado = periodo["estado"] or "ACTIVO"

    pdf.set_font(fam, "", 9)
    pdf.cell(0, 6, f"Cierre ID: {periodo['id']}    Estado: {estado}    Cerrado: {cerrado}", ln=1)
    pdf.cell(0, 6, f"Semanas: {', '.join(str(s) for s in semanas_periodo)}    Periodo: {rango}", ln=1)
    pdf.ln(3)

    # Iterar por departamento → empleado → semana
    depto_actual = None
    for emp in sorted(empleados, key=lambda e: (_normalizar_departamento_web(e.get("departamento","")), _legajo_key(e))):
        depto = emp.get("departamento", "") or "-"
        depto_norm = _normalizar_departamento_web(depto)
        legajo = str(emp.get("legajo", ""))
        nombre = emp.get("nombre", "")
        semanas_emp = sorted(set(emp.get("semanas", semanas_periodo)))

        if depto != depto_actual:
            depto_actual = depto
            pdf.ln(1)
            pdf.set_fill_color(224, 231, 255)
            pdf.set_font(fam, "B", 10)
            pdf.cell(0, 7, _pdf_cell_text(depto), border=1, ln=1, fill=True)

        if pdf.get_y() + 14 > pdf.h - pdf.b_margin:
            pdf.add_page()

        pdf.set_font(fam, "B", 9)
        pdf.cell(0, 6, f"{legajo} - {_pdf_cell_text(nombre)}", ln=1)

        for sem in semanas_emp:
            c = conf_idx.get((depto_norm, legajo, sem))
            if pdf.get_y() + 10 > pdf.h - pdf.b_margin:
                pdf.add_page()

            if c is None:
                # Semana sin confirmar — mostrar datos del CSV si están disponibles
                csv_emp = (datos_csv or {}).get(sem, {}).get(legajo)
                if csv_emp:
                    tot_csv = csv_emp
                    excluido_csv = csv_emp.get("excluido_ot", False)
                    pdf.set_font(fam, "I", 8)
                    pdf.set_text_color(180, 0, 0)
                    pdf.cell(0, 4, f"  Semana {sem} — Sin confirmar  |  OT50: {tot_csv['ot50']}  OT100: {tot_csv['ot100']}  Comidas: {tot_csv['comidas']}  Francos: {tot_csv['francos']}  Tardanzas: {tot_csv['tardanzas']}", ln=1)
                    pdf.set_text_color(0, 0, 0)
                    dias_nc = csv_emp.get("dias", [])
                    if dias_nc:
                        pdf.set_font(fam, "B", 7)
                        pdf.cell(23, 5, "Fecha", 1)
                        pdf.cell(24, 5, "Tipo", 1)
                        pdf.cell(20, 5, "OT50", 1)
                        pdf.cell(20, 5, "OT100", 1)
                        pdf.cell(22, 5, "Marcas", 1)
                        pdf.cell(0, 5, "Descripcion", 1, ln=1)
                        pdf.set_font(fam, "", 7)
                        for d in dias_nc:
                            if pdf.get_y() + 5 > pdf.h - pdf.b_margin:
                                pdf.add_page()
                            marcas = []
                            if d.get("franco"): marcas.append("Franco")
                            if d.get("comida"): marcas.append("Comida")
                            if d.get("tarde"):  marcas.append("Tardanza")
                            x, y = pdf.get_x(), pdf.get_y()
                            ot50_val  = "" if excluido_csv else _pdf_cell_text(d.get("ot50", ""))
                            ot100_val = "" if excluido_csv else _pdf_cell_text(d.get("ot100", ""))
                            pdf.cell(23, 5, _pdf_cell_text(d.get("fecha", "")), 1)
                            pdf.cell(24, 5, _pdf_cell_text(d.get("tipo_dia", "normal")), 1)
                            pdf.cell(20, 5, ot50_val, 1)
                            pdf.cell(20, 5, ot100_val, 1)
                            pdf.cell(22, 5, ", ".join(marcas), 1)
                            pdf.multi_cell(0, 5, "Sin descripcion", 1)
                            if pdf.get_y() < y + 5:
                                pdf.set_y(y + 5)
                            pdf.set_x(x)
                else:
                    pdf.set_font(fam, "I", 8)
                    pdf.set_text_color(180, 0, 0)
                    pdf.cell(0, 5, f"  Semana {sem} — Sin confirmar", ln=1)
                    pdf.set_text_color(0, 0, 0)
            else:
                tot = c.get("totales", {})
                excluido = c.get("excluido_ot", False)
                pdf.set_font(fam, "", 8)
                conf_en = (c.get("confirmado_en") or "")[:16].replace("T", " ")
                pdf.cell(0, 4, f"  Semana {sem}  |  Confirmado: {conf_en}  |  OT50: {tot.get('ot50','0h')}  OT100: {tot.get('ot100','0h')}  Comidas: {tot.get('comidas',0)}  Francos: {tot.get('francos',0)}  Tardanzas: {tot.get('tardanzas',0)}", ln=1)

                dias = c.get("dias", [])
                dias_activos = [
                    d for d in dias
                    if d.get("franco") or d.get("comida") or d.get("tarde")
                    or (not excluido and (
                        _parse_td(d.get("ot50", "")) > timedelta(0)
                        or _parse_td(d.get("ot100", "")) > timedelta(0)
                    ))
                ]
                if dias_activos:
                    pdf.set_font(fam, "B", 7)
                    pdf.cell(23, 5, "Fecha", 1)
                    pdf.cell(24, 5, "Tipo", 1)
                    pdf.cell(20, 5, "OT50", 1)
                    pdf.cell(20, 5, "OT100", 1)
                    pdf.cell(22, 5, "Marcas", 1)
                    pdf.cell(0, 5, "Descripcion", 1, ln=1)
                    pdf.set_font(fam, "", 7)
                    for d in dias_activos:
                        if pdf.get_y() + 5 > pdf.h - pdf.b_margin:
                            pdf.add_page()
                        marcas = []
                        if d.get("franco"): marcas.append("Franco")
                        if d.get("comida"): marcas.append("Comida")
                        if d.get("tarde"):  marcas.append("Tardanza")
                        x, y = pdf.get_x(), pdf.get_y()
                        desc = _pdf_cell_text(d.get("descripcion") or "Sin descripcion")
                        ot50_val  = "" if excluido else _pdf_cell_text(d.get("ot50", ""))
                        ot100_val = "" if excluido else _pdf_cell_text(d.get("ot100", ""))
                        pdf.cell(23, 5, _pdf_cell_text(d.get("fecha", "")), 1)
                        pdf.cell(24, 5, _pdf_cell_text(d.get("tipo_dia", "normal")), 1)
                        pdf.cell(20, 5, ot50_val, 1)
                        pdf.cell(20, 5, ot100_val, 1)
                        pdf.cell(22, 5, ", ".join(marcas), 1)
                        pdf.multi_cell(0, 5, desc, 1)
                        if pdf.get_y() < y + 5:
                            pdf.set_y(y + 5)
                        pdf.set_x(x)

        # Total acumulado del período (desde periodo_empleados)
        if pdf.get_y() + 6 > pdf.h - pdf.b_margin:
            pdf.add_page()
        pdf.set_font(fam, "B", 8)
        pdf.set_fill_color(240, 240, 240)
        total_line = (f"  TOTAL PERÍODO — OT50: {emp.get('ot50','0h')}  "
                      f"OT100: {emp.get('ot100','0h')}  "
                      f"Comidas: {emp.get('comidas',0)}  "
                      f"Francos: {emp.get('francos',0)}  "
                      f"Tardanzas: {emp.get('tardanzas',0)}")
        pdf.cell(0, 5, total_line, border=1, ln=1, fill=True)
        pdf.ln(1)

    return _pdf_bytes(pdf)


def _generar_pdf_confirmaciones_parcial(todos, info):
    from pdf_generator import PDFGeneral

    pdf = PDFGeneral()
    pdf.titulo = "Horas extras parcial"
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    fam = "DejaVu" if pdf._unicode else "Helvetica"

    depto_label = info.get("departamento_label", "Todos")
    rango       = f"{info.get('fecha_desde', '')} al {info.get('fecha_hasta', '')}".strip()
    n_conf      = sum(1 for e in todos if e.get("confirmado"))
    n_pend      = len(todos) - n_conf

    pdf.set_font(fam, "B", 11)
    pdf.cell(0, 7, "Horas Extras - Periodo activo (parcial)", ln=1)
    pdf.set_font(fam, "", 8)
    pdf.cell(0, 5,
        f"Dpto: {depto_label}   Periodo: {rango}   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   "
        f"Total: {len(todos)}   Confirmados: {n_conf}   Pendientes: {n_pend}",
        ln=1)
    pdf.ln(2)

    # Sub-tabla días
    IND = 4
    WD  = {"fecha": 22, "tipo": 22, "ot50": 16, "ot100": 16, "marc": 20}

    def _dias_header():
        pdf.set_x(pdf.l_margin + IND)
        pdf.set_fill_color(240, 244, 255)
        pdf.set_font(fam, "B", 6)
        pdf.cell(WD["fecha"], 4, "Fecha",       1, 0, "C", True)
        pdf.cell(WD["tipo"],  4, "Tipo dia",    1, 0, "C", True)
        pdf.cell(WD["ot50"],  4, "OT50",        1, 0, "C", True)
        pdf.cell(WD["ot100"], 4, "OT100",       1, 0, "C", True)
        pdf.cell(WD["marc"],  4, "Marcas",      1, 0, "C", True)
        pdf.cell(0,           4, "Descripcion", 1, 1, "C", True)

    depto_actual = None
    for e in sorted(todos, key=lambda x: (x.get("departamento", ""), _legajo_key(x))):
        depto     = e.get("departamento", "") or "-"
        dias      = e.get("dias", [])
        confirmado = e.get("confirmado", False)

        if pdf.get_y() + 13 > pdf.h - pdf.b_margin:
            pdf.add_page()

        # Separador de departamento
        if depto != depto_actual:
            depto_actual = depto
            pdf.set_fill_color(224, 231, 255)
            pdf.set_font(fam, "B", 8)
            pdf.cell(0, 5, f"  {_pdf_cell_text(depto)}", 0, 1, "L", True)

        # Banda del empleado — color según estado
        sems    = ", ".join(str(s) for s in e.get("semanas", []))
        conf_en = (e.get("confirmado_en") or "")[:16].replace("T", " ")
        estado  = "CONFIRMADO" if confirmado else "PENDIENTE"
        if confirmado:
            pdf.set_fill_color(234, 242, 255)   # azul claro
        else:
            pdf.set_fill_color(255, 249, 235)   # amarillo claro

        pdf.set_font(fam, "B", 7)
        linea1 = (f"{e.get('legajo','')} - {_pdf_cell_text(e.get('nombre',''))}   "
                  f"Sems: {sems}   "
                  f"OT50: {e.get('ot50','0h')}   OT100: {e.get('ot100','0h')}   "
                  f"Com: {e.get('comidas',0)}   Fr: {e.get('francos',0)}")
        # Ancho disponible para la línea principal menos el badge de estado (30mm)
        W_estado = 30
        W_linea  = pdf.w - pdf.l_margin - pdf.r_margin - W_estado
        x0 = pdf.get_x()
        y0 = pdf.get_y()
        pdf.cell(W_linea,  5, linea1,  "LTB", 0, "L", True)
        pdf.set_font(fam, "B", 7)
        if confirmado:
            pdf.set_text_color(22, 101, 52)    # verde
        else:
            pdf.set_text_color(146, 64, 14)    # naranja
        pdf.cell(W_estado, 5, estado, "RTB", 1, "C", True)
        pdf.set_text_color(0, 0, 0)

        # Fecha de confirmación (segunda línea, solo si confirmado)
        if confirmado and conf_en:
            pdf.set_fill_color(234, 242, 255)
            pdf.set_font(fam, "", 6)
            pdf.cell(0, 3, f"    Confirmado el: {conf_en}", 0, 1, "L", True)

        # Detalle de días
        if dias:
            _dias_header()
            pdf.set_font(fam, "", 6)
            for d in dias:
                if pdf.get_y() + 4 > pdf.h - pdf.b_margin:
                    pdf.add_page()
                    _dias_header()
                marcas = []
                if d.get("franco"): marcas.append("Franco")
                if d.get("comida"): marcas.append("Comida")
                desc = _pdf_cell_text(d.get("descripcion") or "")
                pdf.set_x(pdf.l_margin + IND)
                y1 = pdf.get_y()
                pdf.cell(WD["fecha"], 4, _pdf_cell_text(d.get("fecha", "")),    1)
                pdf.cell(WD["tipo"],  4, _pdf_cell_text(d.get("tipo_dia", "")), 1)
                pdf.cell(WD["ot50"],  4, _pdf_cell_text(d.get("ot50", "")),     1, 0, "C")
                pdf.cell(WD["ot100"], 4, _pdf_cell_text(d.get("ot100", "")),    1, 0, "C")
                pdf.cell(WD["marc"],  4, ", ".join(marcas),                     1, 0, "C")
                pdf.multi_cell(0,     4, desc,                                  1)
                if pdf.get_y() < y1 + 4:
                    pdf.set_y(y1 + 4)
        pdf.ln(1)

    return _pdf_bytes(pdf)


def _pendientes_cierre(confirmaciones, empleados):
    confirmados = {
        (_normalizar_departamento_web(c.get("departamento", "")), str(c.get("legajo", "")))
        for c in confirmaciones
    }
    return [
        e for e in empleados
        if not e.get("confirmado")
        and (_normalizar_departamento_web(e.get("departamento", "")), str(e.get("legajo", ""))) not in confirmados
    ]

def _datos_semanas_periodo(periodo):
    """
    Carga y reprocesa los CSVs de cada semana del período.
    Retorna: {visible_sem: {legajo: {ot50, ot100, comidas, francos, tardanzas, dias, excluido_ot}}}
    Usado para mostrar datos de semanas sin confirmar en el PDF de confirmaciones.
    """
    mapa = _mapa_semanas_visibles_periodo(periodo)  # global → visible
    if not mapa:
        return {}
    departamento = _normalizar_departamento_web(periodo.get("departamento", "") or "")
    resultado = {}
    for global_n, visible_sem in mapa.items():
        csv_path = SEMANAS_DIR / f"semana_{global_n}.csv"
        if not csv_path.exists():
            continue
        try:
            df = _cargar_semana_csv(global_n)
            if df is None:
                continue
            empleados_raw, _, _ = _procesar_empleados(_normalizar_columnas(df))
            _aplicar_exclusiones_ot(empleados_raw)
            sem_data = {}
            for emp in empleados_raw:
                if departamento and _normalizar_departamento_web(emp.get("departamento", "")) != departamento:
                    continue
                leg = str(emp.get("legajo", ""))
                excluido = emp.get("excluido_ot", False)
                dias_prep = _preparar_dias(emp["registros"])
                ot50 = ot100 = timedelta(0)
                comidas = francos = tardanzas = 0
                for d in dias_prep:
                    ot50      += _parse_td(d["ot50"])
                    ot100     += _parse_td(d["ot100"])
                    comidas   += int(d.get("comida", 0))
                    francos   += int(d.get("franco", 0))
                    tardanzas += int(d.get("tarde", 0))
                dias_activos = [
                    d for d in dias_prep
                    if d.get("franco") or d.get("comida") or d.get("tarde")
                    or _parse_td(d.get("ot50",  "")) > timedelta(0)
                    or _parse_td(d.get("ot100", "")) > timedelta(0)
                ]
                sem_data[leg] = {
                    "ot50": _fmt_hm(ot50), "ot100": _fmt_hm(ot100),
                    "comidas": comidas, "francos": francos, "tardanzas": tardanzas,
                    "dias": dias_activos, "excluido_ot": excluido,
                }
            resultado[visible_sem] = sem_data
        except Exception:
            continue
    return resultado


def _mapa_semanas_visibles_periodo(periodo):
    archivo = periodo.get("archivo") or ""
    json_path = PERIODOS_DIR / archivo
    if not json_path.exists():
        return {}
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    semanas = data.get("semanas") or []
    try:
        visible_desde = int(data.get("semana_visible_desde")) if data.get("semana_visible_desde") else 1
    except Exception:
        visible_desde = 1
    return {int(global_n): visible_desde + i for i, global_n in enumerate(semanas)}

def _aplicar_semanas_visibles(empleados, periodo):
    mapa = _mapa_semanas_visibles_periodo(periodo)
    if not mapa:
        return empleados
    for e in empleados:
        e["semanas"] = [mapa.get(int(s), s) for s in e.get("semanas", [])]
    return empleados

def _restaurar_confirmaciones_desde_archivos(departamento=None):
    departamento = _normalizar_departamento_web(departamento)
    meta = _cargar_metadata()
    restauradas = 0
    mejores = {}
    for f in _archivos_confirmacion():
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        depto_conf = _normalizar_departamento_web(data.get("departamento", "") or "Todos")
        if departamento and depto_conf != departamento:
            continue
        sem_conf, sem_depto = _resolver_semana_confirmacion(data, meta)
        data["semana"] = sem_conf
        data["semana_depto"] = sem_depto
        legajo = str(data.get("legajo", ""))
        key = (depto_conf, legajo, sem_conf)
        actual = mejores.get(key)
        if actual is None or _score_confirmacion(data) > _score_confirmacion(actual):
            mejores[key] = data

    for data in mejores.values():
        depto_conf = _normalizar_departamento_web(data.get("departamento", "") or "Todos")
        sem_conf = data.get("semana", 0)
        sem_depto = data.get("semana_depto", sem_conf)
        legajo = str(data.get("legajo", ""))
        for d in _sesion.values():
            if str(d.get("legajo", "")) != legajo:
                continue
            if d.get("semana") != sem_conf:
                continue
            if _normalizar_departamento_web(d.get("departamento", "") or "Todos") != depto_conf:
                continue
            if d.get("confirmado"):
                break
            d["confirmado"] = True
            d["confirmado_en"] = data.get("confirmado_en")
            d["departamento"] = depto_conf
            d["semana_depto"] = sem_depto
            dias_por_fecha = {x.get("fecha"): x for x in data.get("dias", [])}
            for dia in d.get("dias", []):
                confirmado = dias_por_fecha.get(dia.get("fecha"))
                if confirmado:
                    dia["descripcion"] = confirmado.get("descripcion", dia.get("descripcion", ""))
            restauradas += 1
            break
    if restauradas:
        _guardar_sesion(_sesion)
    return restauradas

def _recalcular_totales_token(d):
    ot50 = ot100 = timedelta(0)
    comidas = francos = tardanzas = 0
    for dia in d.get("dias", []):
        ot50      += _parse_td(dia.get("ot50",  "00:00:00"))
        ot100     += _parse_td(dia.get("ot100", "00:00:00"))
        comidas   += int(dia.get("comida",  0))
        francos   += int(dia.get("franco",  0))
        tardanzas += int(dia.get("tarde",   0))
    d["totales"] = {
        "ot50": _fmt_hm(ot50),
        "ot100": _fmt_hm(ot100),
        "comidas": comidas,
        "francos": francos,
        "tardanzas": tardanzas,
    }

def _recortar_semana_lunes_domingo(n, departamento=None):
    departamento = _normalizar_departamento_web(departamento)
    tokens = [
        (t, d) for t, d in _sesion.items()
        if d.get("semana") == n
        and (not departamento or _normalizar_departamento_web(d.get("departamento", "") or "Todos") == departamento)
    ]
    fechas = [dia.get("fecha", "") for _, d in tokens for dia in d.get("dias", [])]
    fecha_desde, fecha_hasta = _rango_lunes_domingo(fechas)
    desde = _parse_fecha(fecha_desde)
    hasta = _parse_fecha(fecha_hasta)
    if not desde or not hasta:
        return {"ok": False, "error": "No hay fechas validas para recortar."}

    afectados = 0
    for _, d in tokens:
        dias = [
            dia for dia in d.get("dias", [])
            if (lambda f: f and desde <= f <= hasta)(_parse_fecha(dia.get("fecha", "")))
        ]
        if len(dias) != len(d.get("dias", [])):
            afectados += 1
        d["dias"] = dias
        _recalcular_totales_token(d)
    _guardar_sesion(_sesion)

    meta = _cargar_metadata()
    for s in meta.get("semanas", []):
        if s.get("numero") == n and (not departamento or _normalizar_departamento_web(s.get("departamento", "") or "Todos") == departamento):
            s["fecha_desde"] = fecha_desde
            s["fecha_hasta"] = fecha_hasta
    _guardar_metadata(meta)
    return {"ok": True, "semana": n, "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta, "tokens_afectados": afectados}

_sesion = _cargar_sesion()
_init_db()

# Recalcular totales de tokens existentes desde sus días (corrige francos mal contados)
_changed = False
for _tok_data in _sesion.values():
    _dias = _tok_data.get("dias", [])
    if not _dias:
        continue
    _ot50 = _ot100 = timedelta(0)
    _com = _fra = _tar = 0
    for _d in _dias:
        _ot50  += _parse_td(_d.get("ot50",  "00:00:00"))
        _ot100 += _parse_td(_d.get("ot100", "00:00:00"))
        _com   += int(_d.get("comida", 0))
        _fra   += int(_d.get("franco", 0))
        _tar   += int(_d.get("tarde",  0))
    _new_t = {"ot50": _fmt_hm(_ot50), "ot100": _fmt_hm(_ot100),
              "comidas": _com, "francos": _fra, "tardanzas": _tar}
    if _tok_data.get("totales") != _new_t:
        _tok_data["totales"] = _new_t
        _changed = True
if _changed:
    _guardar_sesion(_sesion)

# Si no hay semanas activas, limpiar tokens huérfanos de períodos ya cerrados
_meta_inicio = _cargar_metadata()
if not _meta_inicio.get("semanas"):
    if _sesion:
        _sesion.clear()
        _guardar_sesion(_sesion)


# ═══════════════════════════════════════════════
# UTILIDADES
# ═══════════════════════════════════════════════
def _preparar_dias(registros):
    dias_dict = {}
    dias_order = []
    for r in registros:
        fecha = r["Fecha"]
        if fecha not in dias_dict:
            dias_order.append(fecha)
            tiene_ot = (
                r["50%"]  != "00:00:00" or r["100%"] != "00:00:00" or
                int(r.get("FRANCO", 0)) > 0 or int(r.get("COMIDA", 0)) > 0
            )
            obs = r.get("Observaciones", "") or ""
            if   "FER"  in obs: tipo_dia = "feriado"
            elif "SAB"  in obs: tipo_dia = "sabado"
            elif "DOM"  in obs: tipo_dia = "domingo"
            elif "PARO" in obs: tipo_dia = "paro"
            else:               tipo_dia = "normal"
            dias_dict[fecha] = {
                "fecha": fecha, "fecha_fmt": fecha[5:],
                "dia_semana": _dia_semana(fecha),
                "tramos": [],
                "normales": r["Normales"], "ot50": r["50%"], "ot100": r["100%"],
                "comida": int(r.get("COMIDA",0)), "franco": int(r.get("FRANCO",0)),
                "tarde":  int(r.get("Tarde", 0)),
                "tiene_ot": tiene_ot, "tipo_dia": tipo_dia, "descripcion": "",
            }
        e, s = r.get("Entrada",""), r.get("Salida","")
        if e:
            dias_dict[fecha]["tramos"].append({"entrada": e[:5], "salida": s[:5] if s else "—"})
    return [dias_dict[f] for f in dias_order]

def _cargar_excluidos_sa():
    ruta = Path("recursos/excluidos_sa.json")
    try:
        return set(json.loads(ruta.read_text(encoding="utf-8")).get("legajos", []))
    except Exception:
        return set()

def _cargar_excluidos_ot():
    ruta = Path("recursos/excluidos_ot.json")
    try:
        return set(json.loads(ruta.read_text(encoding="utf-8")).get("legajos", []))
    except Exception:
        return set()

def _cargar_francos_deptos_ocultos():
    """Deptos cuyo módulo de Francos está oculto temporalmente (ver
    recursos/francos_deptos_ocultos.json). El depto sigue existiendo en el
    sistema (empleados, histórico); solo se excluye de pantallas, formularios
    de carga, cierres nuevos e informes de francos."""
    ruta = Path("recursos/francos_deptos_ocultos.json")
    try:
        deptos = json.loads(ruta.read_text(encoding="utf-8")).get("departamentos", [])
        return {_normalizar_departamento_web(d) for d in deptos}
    except Exception:
        return set()

def _depto_francos_oculto(depto):
    return _normalizar_departamento_web(depto or "") in _cargar_francos_deptos_ocultos()


_DEPARTAMENTOS_FRANCOS_BASE = {"Ingenieros"}


def _departamentos_francos_disponibles(empleados=None, incluir_ocultos=False):
    """Lista canónica para todos los selectores del módulo Francos.

    Ingenieros debe estar disponible incluso antes de registrar movimientos;
    los demás departamentos se incorporan desde el catálogo de empleados.
    """
    empleados = empleados if empleados is not None else _empleados_conocidos()
    departamentos = set(_DEPARTAMENTOS_FRANCOS_BASE)
    departamentos.update(
        _nombre_departamento_visible(e.get("departamento", ""))
        for e in empleados if e.get("departamento")
    )
    if not incluir_ocultos:
        departamentos = {d for d in departamentos if not _depto_francos_oculto(d)}
    return sorted(departamentos, key=lambda d: d.lower())

_OBS_EXCLUIDO_OT = "Horas registradas para control. No se liquidan para pago."

def _aplicar_exclusiones_ot(empleados):
    """Marca empleados cuya OT no se liquida.

    Las horas se conservan intactas. El flag excluido_ot=True y
    liquida_ot=False habilitan el indicador visual en pantallas y PDFs,
    pero no zerean ningún valor.
    """
    excluidos = _cargar_excluidos_ot()
    for emp in empleados:
        excl = str(emp["legajo"]) in excluidos
        emp["excluido_ot"] = excl
        if excl:
            emp["liquida_ot"] = False
            emp["observacion_liquidacion"] = _OBS_EXCLUIDO_OT
        else:
            emp.setdefault("liquida_ot", True)
            emp.setdefault("observacion_liquidacion", "")
    return empleados


# Normalización de departamentos: ver departamentos.py (compartido con
# reporte_saldos_francos.py, que corre como proceso aparte).
_normalizar_departamento_web = clave_canonica
_nombre_departamento_visible = nombre_visible

def _leer_archivo(fs):
    nombre = (fs.filename or "").lower()
    if nombre.endswith(".xlsx"): return pd.read_excel(fs, engine="openpyxl")
    if nombre.endswith(".xls"):  return pd.read_excel(fs, engine="xlrd")
    for enc in ("utf-8","latin-1","cp1252"):
        for sep in (";",","):
            try:
                fs.seek(0)
                return pd.read_csv(fs, sep=sep, encoding=enc)
            except Exception:
                continue
    raise ValueError("No se pudo leer el archivo.")

def _normalizar_columnas(df):
    aliases = {
        "Nro. de usuario":"Legajo","Fecha/Hora":"FechaHora","Tipo de registro":"Tipo",
        "Fecha_Hora":"FechaHora","Fecha/hora":"FechaHora","fecha_hora":"FechaHora",
        "Depto":"Departamento","depto":"Departamento","departamento":"Departamento",
        "nro_usuario":"Legajo","legajo":"Legajo","nombre":"Nombre","tipo":"Tipo",
    }
    return df.rename(columns={k:v for k,v in aliases.items() if k in df.columns})

def _procesar_empleados(df, departamentos=None):
    """Procesa DataFrame y devuelve (empleados, fecha_desde, fecha_hasta).
    departamentos: lista de nombres a incluir; None = todos."""
    resultados = procesar_fichadas(df, excluir_sa=_cargar_excluidos_sa())
    empleados  = aplanar_registros_por_tramo(resultados)
    if departamentos:
        empleados = [e for e in empleados if e.get("departamento","") in departamentos]
    empleados = _aplicar_exclusiones_ot(empleados)
    todas_fechas = [r["Fecha"] for emp in empleados for r in emp["registros"] if r["Fecha"]]
    fecha_desde, fecha_hasta = _rango_lunes_domingo(todas_fechas)
    if fecha_desde and fecha_hasta:
        empleados = _filtrar_empleados_por_rango(empleados, fecha_desde, fecha_hasta)
    empleados = sorted(empleados, key=_legajo_key)
    return empleados, fecha_desde, fecha_hasta

def _rango_lunes_domingo(fechas):
    """Ventana lunes-domingo que contiene la mayor cantidad de fechas.
    No usa "el primer lunes que aparezca": si el lunes real fue feriado (poca
    o ninguna fichada) puede no estar en los datos, y anclar por ese criterio
    elige la semana equivocada cuando el archivo incluye días sueltos de la
    semana siguiente."""
    fechas_dt = sorted(f for f in (_parse_fecha(x) for x in fechas) if f)
    if not fechas_dt:
        return "", ""
    inicios = [f - timedelta(days=f.weekday()) for f in fechas_dt]
    conteo = Counter(inicios)
    max_count = max(conteo.values())
    inicio = min(i for i, c in conteo.items() if c == max_count)
    fin = inicio + timedelta(days=6)
    return inicio.isoformat(), fin.isoformat()

def _filtrar_empleados_por_rango(empleados, fecha_desde, fecha_hasta):
    desde = _parse_fecha(fecha_desde)
    hasta = _parse_fecha(fecha_hasta)
    if not desde or not hasta:
        return empleados
    filtrados = []
    for emp in empleados:
        registros = [
            r for r in emp.get("registros", [])
            if (lambda f: f and desde <= f <= hasta)(_parse_fecha(r.get("Fecha", "")))
        ]
        if registros:
            emp = dict(emp)
            emp["registros"] = registros
            filtrados.append(emp)
    return filtrados

def _crear_tokens(empleados, semana_n, semana_depto, base_url):
    tokens_creados = []
    links = []
    for emp in empleados:
        if str(emp.get("legajo", "")) in LEGAJOS_EXCLUIR_INFORMES:
            continue
        token = secrets.token_urlsafe(10)
        tokens_creados.append(token)
        dias_prep = _preparar_dias(emp["registros"])
        ot50 = ot100 = timedelta(0)
        comidas = francos = tardanzas = 0
        for d in dias_prep:
            ot50      += _parse_td(d["ot50"])
            ot100     += _parse_td(d["ot100"])
            comidas   += int(d.get("comida", 0))
            francos   += int(d.get("franco", 0))
            tardanzas += int(d.get("tarde",  0))
        _sesion[token] = {
            "legajo": emp["legajo"], "nombre": emp["nombre"],
            "departamento": emp["departamento"],
            "excluido_ot": emp.get("excluido_ot", False),
            "liquida_ot": emp.get("liquida_ot", True),
            "observacion_liquidacion": emp.get("observacion_liquidacion", ""),
            "dias": dias_prep,
            "totales": {
                "ot50": _fmt_hm(ot50), "ot100": _fmt_hm(ot100),
                "comidas": comidas, "francos": francos, "tardanzas": tardanzas,
            },
            "confirmado": False, "confirmado_en": None,
            "semana": semana_n,
            "semana_depto": semana_depto,
        }
        emp_url = f"{base_url}/e/{token}"
        links.append({"legajo": emp["legajo"], "nombre": emp["nombre"],
                       "url": emp_url,
                       "wa_url": _wa_url(emp["legajo"], emp["nombre"], emp_url,
                                         totales=_sesion[token].get("totales"),
                                         dias=_sesion[token].get("dias"))})
    return tokens_creados, links


# ═══════════════════════════════════════════════
# RUTAS PÚBLICAS (empleados)
# ═══════════════════════════════════════════════
@app.route("/e/<token>")
def empleado(token):
    data = _sesion.get(token)
    if not data:
        return render_template("error.html", mensaje="Link inválido o expirado."), 404
    return render_template("empleado.html", data=data, token=token)

@app.route("/e/<token>/confirmar", methods=["POST"])
def confirmar(token):
    data = _sesion.get(token)
    if not data:
        return render_template("error.html", mensaje="Token inválido."), 404
    for dia in data["dias"]:
        if dia["tiene_ot"]:
            dia["descripcion"] = request.form.get(f"desc_{dia['fecha']}", "").strip()
    data["confirmado"]    = True
    data["confirmado_en"] = datetime.now().isoformat()
    _guardar_sesion(_sesion)
    CONFIRM_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    (CONFIRM_DIR / f"{data['legajo']}_{ts}.json").write_text(
        json.dumps({
            "legajo": data["legajo"], "nombre": data["nombre"],
            "departamento": data["departamento"],
            "confirmado_en": data["confirmado_en"],
            "semana": data.get("semana", 0),
            "semana_depto": data.get("semana_depto", data.get("semana", 0)),
            "totales": data["totales"],
            "dias": [
                {"fecha": d["fecha"], "ot50": d["ot50"], "ot100": d["ot100"],
                 "franco": d.get("franco",0), "comida": d.get("comida",0),
                 "tipo_dia": d.get("tipo_dia","normal"), "descripcion": d.get("descripcion",""),
                 "tramos": d.get("tramos", [])}
                for d in data["dias"] if d["tiene_ot"]
            ],
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return render_template("confirmado.html", nombre=data["nombre"])


# ═══════════════════════════════════════════════
# LOGIN / LOGOUT
# ═══════════════════════════════════════════════
@app.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == SUPERVISOR_PASS:
            session["auth"] = True
            return redirect(request.args.get("next") or url_for("index"))
        error = "Contraseña incorrecta."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ═══════════════════════════════════════════════
# RUTAS SUPERVISOR (protegidas)
# ═══════════════════════════════════════════════
@app.route("/")
@app.route("/supervisor")
def index():
    if not _autenticado(): return _requiere_auth()
    return render_template("supervisor.html")


@app.route("/manual")
def manual_usuario():
    if not _autenticado(): return _requiere_auth()
    ruta = Path(__file__).parent / "manual_usuario.html"
    return send_file(ruta, mimetype="text/html")


@app.route("/telefonos/upload", methods=["POST"])
def telefonos_upload():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    if "archivo" not in request.files: return jsonify({"error":"No se recibió archivo"}), 400
    try:
        df = _leer_archivo(request.files["archivo"])
        df.columns = [str(c).strip() for c in df.columns]
        leg_col = next((c for c in df.columns if "legajo" in c.lower()), df.columns[0])
        tel_col = next((c for c in df.columns if "tel" in c.lower() or "cel" in c.lower()), df.columns[-1])
        telefonos_nuevos = {}
        for _, row in df.iterrows():
            leg = _normalizar_valor_excel(row[leg_col])
            tel = _normalizar_valor_excel(row[tel_col])
            if leg and tel:
                telefonos_nuevos[leg] = tel
        if not telefonos_nuevos:
            return jsonify({"error": "No se encontraron telefonos validos en el archivo."}), 400
        telefonos = _cargar_telefonos()
        telefonos.update(telefonos_nuevos)
        _guardar_telefonos(telefonos)
        return jsonify({"ok": True, "total": len(telefonos), "actualizados": len(telefonos_nuevos)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/detectar-departamentos", methods=["POST"])
def detectar_departamentos():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    if "csv" not in request.files: return jsonify({"error":"No se recibió archivo"}), 400
    try:
        df = _normalizar_columnas(_leer_archivo(request.files["csv"]))
        resultados = procesar_fichadas(df, excluir_sa=_cargar_excluidos_sa())
        empleados  = aplanar_registros_por_tramo(resultados)
        deptos = sorted(set(e.get("departamento","—") for e in empleados))
        return jsonify({"departamentos": deptos})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/procesar", methods=["POST"])
def procesar():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    if "csv" not in request.files: return jsonify({"error":"No se recibió archivo"}), 400
    try:
        df = _normalizar_columnas(_leer_archivo(request.files["csv"]))
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    departamentos = request.form.getlist("departamentos")
    if len(departamentos) != 1:
        return jsonify({"error": "Seleccioná un departamento."}), 400
    depto_original = departamentos[0].strip()
    depto_label = _normalizar_departamento_web(depto_original)
    if not depto_label or depto_label == "todos":
        return jsonify({"error": "Seleccioná un departamento válido."}), 400

    try:
        empleados, fecha_desde, fecha_hasta = _procesar_empleados(df, departamentos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not empleados:
        return jsonify({"error": "No se encontraron empleados para los departamentos seleccionados."}), 400

    for emp in empleados:
        emp["departamento"] = depto_label

    meta = _cargar_metadata()

    # Detectar duplicado por fechas + departamento
    if not request.form.get("force"):
        for s in meta.get("semanas", []):
            if s.get("fecha_desde") == fecha_desde and s.get("fecha_hasta") == fecha_hasta \
               and s.get("departamento", "Todos") == depto_label:
                nd = s.get("num_depto", s["numero"])
                return jsonify({
                    "duplicado": True,
                    "semana_existente": s["numero"],
                    "num_depto_existente": nd,
                    "fecha_desde": fecha_desde,
                    "fecha_hasta": fecha_hasta,
                    "msg": f"Ya existe la Semana {nd} con las mismas fechas y departamento ({fecha_desde} → {fecha_hasta} / {depto_label}).",
                })

    n = meta["semana_actual"] + 1
    meta["semana_actual"] = n

    # Número de semana por departamento (independiente del global)
    num_depto = sum(1 for s in meta.get("semanas", [])
                    if _normalizar_departamento_web(s.get("departamento", "Todos"))
                    == _normalizar_departamento_web(depto_label)) + 1

    _guardar_semana_csv(n, df)
    base_url = request.host_url.rstrip("/")
    tokens_creados, links = _crear_tokens(empleados, n, num_depto, base_url)
    _guardar_sesion(_sesion)

    meta["semanas"].append({
        "numero": n, "num_depto": num_depto,
        "fecha_upload": datetime.now().isoformat(),
        "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
        "departamento": depto_label,
        "tokens": tokens_creados,
        "legajos": [emp["legajo"] for emp in empleados],
    })
    _guardar_metadata(meta)

    return jsonify({"empleados": links, "semana": n, "num_depto": num_depto,
                    "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
                    "departamento": depto_label})


@app.route("/semanas/<int:n>/reprocesar", methods=["POST"])
def reprocesar_semana(n):
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    if "csv" not in request.files: return jsonify({"error":"No se recibió archivo"}), 400

    meta = _cargar_metadata()
    semana_existente = next((s for s in meta.get("semanas", []) if s.get("numero") == n), None)
    if not semana_existente:
        return jsonify({"error": f"No existe la semana {n}."}), 404

    # Legajos a actualizar (vacío = todos)
    solo_legajos_raw = request.form.get("solo_legajos", "").strip()
    solo_legajos = {s.strip() for s in solo_legajos_raw.split(",") if s.strip()} if solo_legajos_raw else set()

    try:
        df = _normalizar_columnas(_leer_archivo(request.files["csv"]))
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    depto_label = semana_existente.get("departamento", "Todos")
    try:
        empleados_nuevos, fecha_desde, fecha_hasta = _procesar_empleados(df, [depto_label])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not empleados_nuevos:
        return jsonify({"error": "No se encontraron empleados para el departamento."}), 400

    for emp in empleados_nuevos:
        emp["departamento"] = depto_label

    base_url = request.host_url.rstrip("/")
    num_depto = semana_existente.get("num_depto", n)

    # Mapa legajo → nuevo empleado procesado
    nuevos_por_legajo = {str(e["legajo"]): e for e in empleados_nuevos}

    # Mapa legajo → token existente
    token_por_legajo = {
        str(_sesion[t]["legajo"]): t
        for t in semana_existente.get("tokens", [])
        if t in _sesion
    }

    links = []
    tokens_finales = list(semana_existente.get("tokens", []))  # preservar orden original

    # Actualizar solo los legajos indicados (o todos si no se especificaron)
    legajos_a_actualizar = solo_legajos if solo_legajos else set(nuevos_por_legajo.keys())

    # Reprocesar completo: eliminar tokens de empleados que ya no están en el CSV
    if not solo_legajos:
        for t in list(tokens_finales):
            if t not in _sesion:
                continue
            leg_t = str(_sesion[t].get("legajo", ""))
            if leg_t not in nuevos_por_legajo:
                if leg_t in _sesion:
                    del _sesion[leg_t]
                if t in _sesion:
                    del _sesion[t]
                tokens_finales.remove(t)

    for leg in legajos_a_actualizar:
        emp = nuevos_por_legajo.get(leg)
        if not emp:
            continue  # legajo no encontrado en el CSV nuevo, no tocar nada

        dias_prep = _preparar_dias(emp["registros"])
        ot50 = ot100 = timedelta(0)
        comidas = francos = tardanzas = 0
        for d in dias_prep:
            ot50      += _parse_td(d["ot50"])
            ot100     += _parse_td(d["ot100"])
            comidas   += int(d.get("comida", 0))
            francos   += int(d.get("franco", 0))
            tardanzas += int(d.get("tarde",  0))
        totales = {"ot50": _fmt_hm(ot50), "ot100": _fmt_hm(ot100),
                   "comidas": comidas, "francos": francos, "tardanzas": tardanzas}

        if leg in token_por_legajo:
            # Actualizar datos en el token existente, preservar link y confirmado
            token = token_por_legajo[leg]
            _sesion[token] = {
                **_sesion[token],
                "dias": dias_prep,
                "totales": totales,
                "excluido_ot": emp.get("excluido_ot", False),
                "liquida_ot": emp.get("liquida_ot", True),
                "observacion_liquidacion": emp.get("observacion_liquidacion", ""),
                "nombre": emp["nombre"],
                "departamento": depto_label,
            }
        else:
            # Empleado nuevo en el CSV: crear token
            token = secrets.token_urlsafe(10)
            _sesion[token] = {
                "legajo": leg, "nombre": emp["nombre"],
                "departamento": depto_label,
                "excluido_ot": emp.get("excluido_ot", False),
                "liquida_ot": emp.get("liquida_ot", True),
                "observacion_liquidacion": emp.get("observacion_liquidacion", ""),
                "dias": dias_prep, "totales": totales,
                "confirmado": False, "confirmado_en": None,
                "semana": n, "semana_depto": num_depto,
            }
            tokens_finales.append(token)
            token_por_legajo[leg] = token

    # Armar lista de links de todos los empleados de la semana (actualizados y sin tocar)
    for t in tokens_finales:
        if t not in _sesion:
            continue
        d = _sesion[t]
        emp_url = f"{base_url}/e/{t}"
        links.append({"legajo": d["legajo"], "nombre": d["nombre"],
                      "url": emp_url,
                      "wa_url": _wa_url(d["legajo"], d["nombre"], emp_url,
                                        totales=d.get("totales"), dias=d.get("dias"))})

    _guardar_sesion(_sesion)

    # Solo pisar el CSV si se actualizaron todos (parcial no reemplaza el archivo)
    if not solo_legajos:
        _guardar_semana_csv(n, df)
        semana_existente["fecha_desde"] = fecha_desde
        semana_existente["fecha_hasta"] = fecha_hasta
        semana_existente["legajos"] = [emp["legajo"] for emp in empleados_nuevos]

    semana_existente["tokens"] = tokens_finales
    semana_existente["fecha_upload"] = datetime.now().isoformat()
    _guardar_metadata(meta)

    actualizados = list(legajos_a_actualizar & set(nuevos_por_legajo.keys()))
    return jsonify({"empleados": links, "semana": n, "num_depto": num_depto,
                    "fecha_desde": semana_existente.get("fecha_desde", ""),
                    "fecha_hasta": semana_existente.get("fecha_hasta", ""),
                    "departamento": depto_label,
                    "actualizados": actualizados})


@app.route("/semanas")
def semanas():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    meta = _cargar_metadata()
    base_url = request.host_url.rstrip("/")
    resultado = []
    for s in meta.get("semanas", []):
        n = s["numero"]
        tokens = [t for t in s.get("tokens",[]) if t in _sesion]
        total      = len(tokens)
        confirmados = sum(1 for t in tokens if _sesion[t].get("confirmado"))
        resultado.append({
            "numero":       n,
            "num_depto":    s.get("num_depto", n),
            "departamento": s.get("departamento","Todos"),
            "fecha_desde":  s.get("fecha_desde",""),
            "fecha_hasta":  s.get("fecha_hasta",""),
            "fecha_upload": s.get("fecha_upload","")[:10],
            "total":        total,
            "confirmados":  confirmados,
        })
    return jsonify(resultado)


@app.route("/semanas/<int:n>/links")
def semana_links(n):
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    base_url = request.host_url.rstrip("/")
    links = [
        {"legajo": d["legajo"], "nombre": d["nombre"],
         "confirmado": d["confirmado"], "url": f"{base_url}/e/{t}",
         "wa_url": _wa_url(d["legajo"], d["nombre"], f"{base_url}/e/{t}",
                           totales=d.get("totales"), dias=d.get("dias"))}
        for t, d in _sesion.items() if d.get("semana") == n
    ]
    links.sort(key=lambda x: (x["confirmado"], x["nombre"]))
    return jsonify(links)


@app.route("/semanas/<int:n>/eliminar", methods=["POST"])
def eliminar_semana(n):
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    meta = _cargar_metadata()
    sem = next((s for s in meta["semanas"] if s["numero"] == n), None)
    if not sem:
        return jsonify({"error": f"Semana {n} no encontrada"}), 404

    # Eliminar TODOS los tokens de esta semana (incluyendo huérfanos)
    tokens_a_borrar = [t for t, d in _sesion.items() if d.get("semana") == n]
    for t in tokens_a_borrar:
        del _sesion[t]
    _guardar_sesion(_sesion)

    # Eliminar CSV guardado
    csv_path = SEMANAS_DIR / f"semana_{n}.csv"
    if csv_path.exists():
        csv_path.unlink()

    # Borrar francos parciales de esta semana
    with _get_db() as conn:
        conn.execute("DELETE FROM francos_semana_parcial WHERE semana_num=?", (n,))
        conn.commit()

    # Quitar de metadata y reordenar contador si era la última
    meta["semanas"] = [s for s in meta["semanas"] if s["numero"] != n]
    if meta["semana_actual"] == n:
        meta["semana_actual"] = max((s["numero"] for s in meta["semanas"]), default=0)
    _guardar_metadata(meta)

    return jsonify({"ok": True})


@app.route("/semanas/<int:n>/guardar-francos", methods=["POST"])
def guardar_francos_semana(n):
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    ahora = datetime.now().isoformat(timespec="seconds")
    tokens_semana = [(t, d) for t, d in _sesion.items() if d.get("semana") == n]
    if not tokens_semana:
        return jsonify({"error": f"No hay datos de semana {n} en sesión"}), 404
    with _get_db() as conn:
        conn.execute("DELETE FROM francos_semana_parcial WHERE semana_num=?", (n,))
        for _, d in tokens_semana:
            dias = int(d.get("totales", {}).get("francos", 0))
            conn.execute(
                "INSERT INTO francos_semana_parcial (legajo, nombre, departamento, semana_num, dias, guardado_en) VALUES (?,?,?,?,?,?)",
                (str(d.get("legajo", "")), d.get("nombre", ""), d.get("departamento", ""), n, dias, ahora)
            )
        conn.commit()
    guardados = len(tokens_semana)
    return jsonify({"ok": True, "semana": n, "empleados": guardados})


@app.route("/semanas/guardar-francos-todos", methods=["POST"])
def guardar_francos_todos():
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    semanas = {d.get("semana") for d in _sesion.values() if d.get("semana") is not None}
    if not semanas:
        return jsonify({"error": "No hay semanas en sesión"}), 404
    ahora = datetime.now().isoformat(timespec="seconds")
    total_empleados = 0
    with _get_db() as conn:
        # Limpiar todo para evitar que queden semanas eliminadas
        conn.execute("DELETE FROM francos_semana_parcial")
        for n in semanas:
            tokens_semana = [(t, d) for t, d in _sesion.items() if d.get("semana") == n]
            for _, d in tokens_semana:
                dias = int(d.get("totales", {}).get("francos", 0))
                conn.execute(
                    "INSERT INTO francos_semana_parcial (legajo, nombre, departamento, semana_num, dias, guardado_en) VALUES (?,?,?,?,?,?)",
                    (str(d.get("legajo", "")), d.get("nombre", ""), d.get("departamento", ""), n, dias, ahora)
                )
            total_empleados += len(tokens_semana)
        conn.commit()
    return jsonify({"ok": True, "semanas": len(semanas), "empleados": total_empleados})


@app.route("/semanas/<int:n>/regenerar", methods=["POST"])
def regenerar_semana(n):
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    df = _cargar_semana_csv(n)
    if df is None: return jsonify({"error": f"No hay datos guardados para la semana {n}"}), 404

    meta = _cargar_metadata()
    sem_meta = next((s for s in meta["semanas"] if s["numero"] == n), None)
    num_depto = sem_meta.get("num_depto", n) if sem_meta else n
    depto_label = sem_meta.get("departamento", "") if sem_meta else ""

    try:
        empleados, _, _ = _procesar_empleados(_normalizar_columnas(df))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Filtrar a los legajos que realmente pertenecen a esta semana.
    # Necesario porque el CSV guardado puede contener empleados de otros dtos.
    legajos_semana = set(sem_meta.get("legajos", [])) if sem_meta else set()
    if not legajos_semana:
        # Fallback para semanas antiguas sin campo "legajos": derivar desde
        # los tokens activos en sesión ANTES de borrar los pendientes.
        legajos_semana = {d["legajo"] for t, d in _sesion.items()
                          if d.get("semana") == n}
    if legajos_semana:
        empleados = [e for e in empleados if e["legajo"] in legajos_semana]

    if depto_label:
        for emp in empleados:
            emp["departamento"] = depto_label

    # Legajos que ya confirmaron esta semana → no regenerar
    ya_confirmados = {
        d["legajo"] for t, d in _sesion.items()
        if d.get("semana") == n and d.get("confirmado")
    }
    # Eliminar tokens pendientes de esta semana
    pendientes = [t for t, d in _sesion.items()
                  if d.get("semana") == n and not d.get("confirmado")]
    for t in pendientes:
        del _sesion[t]

    empleados_pendientes = [e for e in empleados if e["legajo"] not in ya_confirmados]
    base_url = request.host_url.rstrip("/")
    tokens_nuevos, links = _crear_tokens(empleados_pendientes, n, num_depto, base_url)
    _guardar_sesion(_sesion)

    # Actualizar tokens en metadata (y backfill legajos si faltaba)
    meta = _cargar_metadata()
    for s in meta["semanas"]:
        if s["numero"] == n:
            tokens_vivos = [t for t in s.get("tokens",[])
                            if t in _sesion and _sesion[t].get("confirmado")]
            s["tokens"] = tokens_vivos + tokens_nuevos
            if not s.get("legajos") and legajos_semana:
                s["legajos"] = list(legajos_semana)
            break
    _guardar_metadata(meta)

    return jsonify({"empleados": links, "semana": n,
                    "ya_confirmados": len(ya_confirmados)})


@app.route("/semanas/<int:n>/pdf")
def semana_pdf(n):
    """Genera PDF detallado (día a día) de una semana ya procesada."""
    if not _autenticado(): return _requiere_auth()
    df = _cargar_semana_csv(n)
    if df is None:
        return "No hay datos para esta semana.", 404

    meta       = _cargar_metadata()
    sem_meta   = next((s for s in meta.get("semanas", []) if s["numero"] == n), None)
    num_depto  = sem_meta.get("num_depto", n) if sem_meta else n
    depto_label = sem_meta.get("departamento", "") if sem_meta else ""
    fecha_desde = sem_meta.get("fecha_desde", "") if sem_meta else ""
    fecha_hasta = sem_meta.get("fecha_hasta", "") if sem_meta else ""
    legajos_semana = set(str(l) for l in (sem_meta.get("legajos") or [])) if sem_meta else set()

    try:
        empleados, _, _ = _procesar_empleados(_normalizar_columnas(df))
    except Exception as e:
        return f"Error al procesar datos: {e}", 500

    if legajos_semana:
        empleados = [e for e in empleados if str(e["legajo"]) in legajos_semana]
    if depto_label:
        for emp in empleados:
            emp["departamento"] = depto_label
    if not empleados:
        return "No se encontraron empleados para esta semana.", 404

    depto_visible = _nombre_departamento_visible(depto_label)
    mes_label = f"Semana {num_depto} — {depto_visible} — {fecha_desde} al {fecha_hasta}"
    feriados  = list(_cargar_feriados_config())

    from pdf_generator import generar_pdf_general
    try:
        pdf_data = generar_pdf_general(empleados, mes_label, feriados=feriados)
    except Exception as e:
        return f"Error generando PDF: {e}", 500

    nombre = f"semana_{num_depto}_{depto_label}_{fecha_desde}_{fecha_hasta}.pdf"
    return send_file(BytesIO(pdf_data), mimetype="application/pdf",
                     as_attachment=False, download_name=nombre)


@app.route("/semanas/acumulado/pdf")
def semanas_acumulado_pdf():
    """Genera PDF detallado acumulado de varias semanas (para informar al supervisor al viernes)."""
    if not _autenticado(): return _requiere_auth()
    try:
        desde = int(request.args.get("desde", 1))
        hasta = int(request.args.get("hasta", 1))
    except (TypeError, ValueError):
        return "Parámetros desde/hasta inválidos.", 400

    depto_param   = _normalizar_departamento_web(request.args.get("depto", "").strip())
    meta          = _cargar_metadata()
    semanas_rango = [s for s in meta.get("semanas", [])
                     if desde <= s["numero"] <= hasta
                     and (not depto_param or _normalizar_departamento_web(s.get("departamento","")) == depto_param)]
    if not semanas_rango:
        return "No hay semanas en ese rango.", 404

    feriados        = list(_cargar_feriados_config())
    todos_empleados = {}

    for sem in semanas_rango:
        n_sem       = sem["numero"]
        depto_label  = sem.get("departamento", "")
        legajos_sem  = set(str(l) for l in (sem.get("legajos") or []))
        df = _cargar_semana_csv(n_sem)
        if df is None:
            continue
        try:
            empleados, _, _ = _procesar_empleados(_normalizar_columnas(df))
        except Exception:
            continue
        if legajos_sem:
            empleados = [e for e in empleados if str(e["legajo"]) in legajos_sem]
        if depto_label:
            for emp in empleados:
                emp["departamento"] = depto_label
        for emp in empleados:
            key = str(emp["legajo"])
            if key not in todos_empleados:
                todos_empleados[key] = {
                    "legajo":       emp["legajo"],
                    "nombre":       emp["nombre"],
                    "departamento": emp.get("departamento", ""),
                    "registros":    [],
                    "excluido_ot":  emp.get("excluido_ot", False),
                }
            todos_empleados[key]["registros"].extend(emp.get("registros", []))

    if not todos_empleados:
        return "No se pudieron obtener datos de las semanas seleccionadas.", 500

    data    = list(todos_empleados.values())
    sem_pri = min(semanas_rango, key=lambda s: s["numero"])
    sem_ult = max(semanas_rango, key=lambda s: s["numero"])
    fd      = sem_pri.get("fecha_desde", "")
    fh      = sem_ult.get("fecha_hasta", "")
    depto_label_0 = semanas_rango[0].get("departamento", "")
    depto_visible = _nombre_departamento_visible(depto_label_0)
    mes_label     = f"Acumulado — {depto_visible} — {fd} al {fh}"

    from pdf_generator import generar_pdf_general
    try:
        pdf_data = generar_pdf_general(data, mes_label, feriados=feriados)
    except Exception as e:
        return f"Error generando PDF: {e}", 500

    nombre = f"acumulado_{depto_label_0}_{fd}_{fh}.pdf"
    return send_file(BytesIO(pdf_data), mimetype="application/pdf",
                     as_attachment=False, download_name=nombre)


@app.route("/estado")
def estado():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    meta = _cargar_metadata()
    semanas_activas = {s.get("numero") for s in meta.get("semanas", [])}
    resultado = [
        {"legajo": d["legajo"], "nombre": d["nombre"],
         "departamento": d.get("departamento",""),
         "semana": d.get("semana_depto", d.get("semana",0)),
         "confirmado": d["confirmado"],
         "dias": [
             {"fecha": x["fecha"], "ot50": x["ot50"], "ot100": x["ot100"],
              "descripcion": x.get("descripcion","")}
             for x in d["dias"] if x["tiene_ot"]
         ] if d["confirmado"] else []}
        for d in _sesion.values()
        if not semanas_activas or d.get("semana") in semanas_activas
    ]
    resultado.sort(key=lambda x: (x["semana"], not x["confirmado"], x["nombre"]))
    return jsonify(resultado)


@app.route("/admin/restaurar-confirmaciones", methods=["POST"])
def admin_restaurar_confirmaciones():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    departamento = request.form.get("departamento", "")
    restauradas = _restaurar_confirmaciones_desde_archivos(departamento)
    return jsonify({"ok": True, "restauradas": restauradas})


@app.route("/admin/recortar-semana", methods=["POST"])
def admin_recortar_semana():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    departamento = request.form.get("departamento", "")
    try:
        n = int(request.form.get("semana", "0"))
    except ValueError:
        return jsonify({"error": "Semana invalida."}), 400
    try:
        semana_depto = int(request.form.get("semana_depto", "0"))
    except ValueError:
        semana_depto = 0
    if n <= 0 and semana_depto > 0 and departamento:
        depto = _normalizar_departamento_web(departamento)
        meta = _cargar_metadata()
        sem = next((
            s for s in meta.get("semanas", [])
            if _normalizar_departamento_web(s.get("departamento", "") or "Todos") == depto
            and int(s.get("num_depto", s.get("numero", 0))) == semana_depto
        ), None)
        if sem:
            n = int(sem.get("numero", 0))
    if n <= 0:
        return jsonify({"error": "Semana requerida."}), 400
    resultado = _recortar_semana_lunes_domingo(n, departamento)
    status = 200 if resultado.get("ok") else 400
    return jsonify(resultado), status


@app.route("/historial")
def historial():
    if not _autenticado(): return _requiere_auth()
    departamento = _normalizar_departamento_web(request.args.get("departamento", ""))
    semana_raw = request.args.get("semana", "").strip()
    try:
        semana = int(semana_raw) if semana_raw else None
    except ValueError:
        semana = None
    todos = _leer_historial()
    deptos_map = {}
    semanas_map = {}
    for item in todos:
        valor = _normalizar_departamento_web(item.get("departamento", ""))
        if not valor or valor == "todos":
            continue
        deptos_map[valor] = _nombre_departamento_visible(item.get("departamento", ""))
        sem_depto_val = item.get("semana_depto") or item.get("semana")
        if sem_depto_val:
            semanas_map[sem_depto_val] = sem_depto_val
    departamentos = [
        {"valor": valor, "nombre": nombre}
        for valor, nombre in sorted(deptos_map.items(), key=lambda item: item[1])
    ]
    semanas = [
        {"valor": valor, "nombre": nombre}
        for valor, nombre in sorted(semanas_map.items(), key=lambda item: item[1])
    ]
    items = [i for i in _leer_historial(semana=semana, departamento=departamento)
              if str(i.get("legajo", "")) not in LEGAJOS_EXCLUIR_INFORMES]
    todas_fechas = [d["fecha"] for item in items for d in item.get("dias", []) if d.get("fecha")]
    fecha_desde = fecha_hasta = None
    if todas_fechas:
        f_min, f_max = min(todas_fechas), max(todas_fechas)
        fecha_desde = f"{f_min[8:10]}/{f_min[5:7]}/{f_min[:4]}"
        fecha_hasta = f"{f_max[8:10]}/{f_max[5:7]}/{f_max[:4]}"
    return render_template("historial.html", items=items,
                           departamentos=departamentos,
                           departamento_actual=departamento,
                           semanas=semanas,
                           semana_actual=semana,
                           fecha_desde=fecha_desde,
                           fecha_hasta=fecha_hasta)


@app.route("/historial/acumulado")
def historial_acumulado():
    if not _autenticado(): return _requiere_auth()
    departamento = _normalizar_departamento_web(request.args.get("departamento", ""))

    todos_sin_filtro = _leer_historial()
    deptos_map = {}
    for item in todos_sin_filtro:
        valor = _normalizar_departamento_web(item.get("departamento", ""))
        if not valor or valor == "todos":
            continue
        deptos_map[valor] = _nombre_departamento_visible(item.get("departamento", ""))
    departamentos = [
        {"valor": v, "nombre": n}
        for v, n in sorted(deptos_map.items(), key=lambda x: x[1])
    ]

    todos = _leer_historial(departamento=departamento if departamento else None)

    meta = _cargar_metadata()
    sem_fechas = {}
    for s in meta.get("semanas", []):
        depto_s = _normalizar_departamento_web(s.get("departamento", "") or "Todos")
        if not departamento or depto_s == departamento or depto_s == "todos":
            nd = s.get("num_depto", s.get("numero"))
            if nd is not None and nd not in sem_fechas:
                sem_fechas[nd] = {
                    "desde": s.get("fecha_desde", ""),
                    "hasta": s.get("fecha_hasta", ""),
                }

    excluidos = _cargar_excluidos_ot()

    # Mapa (legajo, fecha) → tramos desde _sesion para enriquecer archivos viejos sin tramos
    tramos_sesion = {}
    for d in _sesion.values():
        leg = str(d.get("legajo", ""))
        for dia in d.get("dias", []):
            t = dia.get("tramos", [])
            if leg and dia.get("fecha") and t:
                tramos_sesion[(leg, dia["fecha"])] = t

    por_empleado = {}
    semanas_set = set()

    for item in todos:
        legajo = str(item["legajo"])
        if legajo in LEGAJOS_EXCLUIR_INFORMES:
            continue
        depto = _normalizar_departamento_web(item.get("departamento", "") or "Todos")
        sem = item.get("semana_depto") or item.get("semana", 0)
        semanas_set.add(sem)

        clave = (depto, legajo)
        excluido = item.get("excluido_ot") or legajo in excluidos
        if clave not in por_empleado:
            por_empleado[clave] = {
                "legajo": legajo,
                "nombre": item["nombre"],
                "departamento": _nombre_departamento_visible(depto),
                "departamento_key": depto,
                "por_semana": {},
                "total": {"ot50": timedelta(0), "ot100": timedelta(0),
                          "comidas": 0, "francos": 0, "tardanzas": 0},
                "excluido_ot": excluido,
                "liquida_ot": item.get("liquida_ot", not excluido),
                "observacion_liquidacion": item.get(
                    "observacion_liquidacion",
                    _OBS_EXCLUIDO_OT if excluido else ""
                ),
            }

        e = por_empleado[clave]
        ot50_td  = _parse_hm(item["totales"]["ot50"])
        ot100_td = _parse_hm(item["totales"]["ot100"])
        comidas   = item["totales"].get("comidas", 0)
        francos   = sum(1 for d in item.get("dias", []) if d.get("franco"))
        tardanzas = item["totales"].get("tardanzas", 0)
        comentarios = [
            {"fecha": d["fecha"], "texto": d["descripcion"]}
            for d in item.get("dias", [])
            if d.get("descripcion", "").strip()
        ]

        dias_raw = sorted(item.get("dias", []), key=lambda d: d["fecha"])
        # Enriquecer con tramos de _sesion si el archivo de confirmación no los tiene
        dias_raw = [
            dict(d, tramos=d.get("tramos") or tramos_sesion.get((legajo, d["fecha"]), []))
            for d in dias_raw
        ]
        e["por_semana"][sem] = {
            "ot50": _fmt_hm(ot50_td),
            "ot100": _fmt_hm(ot100_td),
            "comidas": comidas,
            "francos": francos,
            "tardanzas": tardanzas,
            "comentarios": comentarios,
            "dias_list": dias_raw,
        }
        e["total"]["ot50"]      += ot50_td
        e["total"]["ot100"]     += ot100_td
        e["total"]["comidas"]   += comidas
        e["total"]["francos"]   += francos
        e["total"]["tardanzas"] += tardanzas

    empleados_lista = []
    for clave, e in sorted(por_empleado.items(),
                           key=lambda x: _legajo_key({"legajo": x[1]["legajo"]})):
        e["total"]["ot50"] = _fmt_hm(e["total"]["ot50"])
        e["total"]["ot100"] = _fmt_hm(e["total"]["ot100"])
        empleados_lista.append(e)

    semanas = sorted(semanas_set)

    deptos_resultado = {}
    for e in empleados_lista:
        dk = e["departamento_key"]
        if dk not in deptos_resultado:
            deptos_resultado[dk] = {"nombre": e["departamento"], "empleados": []}
        deptos_resultado[dk]["empleados"].append(e)

    return render_template("historial_acumulado.html",
                           deptos=list(deptos_resultado.values()),
                           semanas=semanas,
                           sem_fechas=sem_fechas,
                           departamentos=departamentos,
                           departamento_actual=departamento)


@app.route("/periodo")
def periodo():
    if not _autenticado(): return _requiere_auth()
    meta = _cargar_metadata()
    # Una entrada por semana/upload para que el selector las distinga
    semanas_selector = sorted([
        {"numero": s["numero"],
         "num_depto": s.get("num_depto", s["numero"]),
         "departamento": _normalizar_departamento_web(s.get("departamento", "")),
         "departamento_visible": _nombre_departamento_visible(s.get("departamento", "")),
         "fecha_desde": s.get("fecha_desde", ""),
         "fecha_hasta": s.get("fecha_hasta", ""),
         "archivo": s.get("archivo", "")}
        for s in meta.get("semanas", [])
    ], key=lambda x: x["numero"])
    deptos_map = {}
    for s in meta.get("semanas", []):
        valor = _normalizar_departamento_web(s.get("departamento", ""))
        if not valor or valor == "todos":
            continue
        deptos_map[valor] = _nombre_departamento_visible(s.get("departamento", ""))
    departamentos = [
        {"valor": valor, "nombre": nombre}
        for valor, nombre in sorted(deptos_map.items(), key=lambda item: item[1])
    ]
    return render_template("periodo.html", semanas=semanas_selector,
                           departamentos=departamentos,
                           firma=FIRMA_SUPERVISOR)


def _calcular_periodo(desde, hasta, departamento=None):
    """Acumula totales del período incluyendo no confirmados.

    Excluye LEGAJOS_EXCLUIR_INFORMES (Gatti y Mancioni): sus francos pasaron
    a cargarse manualmente, igual que Guardias/Internet/Telefonía, por lo que
    no deben generar periodo_empleados ni contar de nuevo en los informes,
    exportaciones ni en el cierre."""
    departamento = _normalizar_departamento_web(departamento)

    # Mapa semana global → num_depto para mostrar el número visible al usuario
    meta = _cargar_metadata()
    semana_visible = {}
    for s in meta.get("semanas", []):
        depto_s = _normalizar_departamento_web(s.get("departamento", "") or "Todos")
        if not departamento or depto_s == departamento or depto_s == "todos":
            num_g = s.get("numero", 0)
            semana_visible[num_g] = s.get("num_depto", num_g)

    por_empleado = {}

    for c in _leer_historial():
        sem = c.get("semana", 0)
        if not (desde <= sem <= hasta):
            continue
        depto = _normalizar_departamento_web(c.get("departamento", "") or "Todos")
        if departamento and depto != departamento:
            continue
        legajo = str(c["legajo"])
        clave = (depto, legajo)
        if clave not in por_empleado:
            por_empleado[clave] = {
                "legajo": legajo, "nombre": c["nombre"],
                "departamento": _nombre_departamento_visible(depto),
                "ot50": timedelta(0), "ot100": timedelta(0),
                "comidas": 0, "francos": 0, "tardanzas": 0,
                "semanas": [], "dias": [],
                "conf_sem": set(), "pend_sem": set(),
                "excluido_ot": c.get("excluido_ot", False),
                "liquida_ot": c.get("liquida_ot", True),
                "observacion_liquidacion": c.get("observacion_liquidacion", ""),
            }
        e = por_empleado[clave]
        e["ot50"]      += _parse_hm(c["totales"]["ot50"])
        e["ot100"]     += _parse_hm(c["totales"]["ot100"])
        e["comidas"]   += c["totales"].get("comidas", 0)
        e["francos"]   += sum(1 for dia in c.get("dias", []) if dia.get("franco"))
        e["tardanzas"] += c["totales"].get("tardanzas", 0)
        sem_vis = semana_visible.get(sem, sem)
        if sem_vis not in e["semanas"]: e["semanas"].append(sem_vis)
        e["dias"].extend(c.get("dias", []))
        e["conf_sem"].add(sem)

    for token, d in _sesion.items():
        sem = d.get("semana", 0)
        if not (desde <= sem <= hasta) or d.get("confirmado"):
            continue
        depto = _normalizar_departamento_web(d.get("departamento", "") or "Todos")
        if departamento and depto != departamento:
            continue
        legajo = str(d["legajo"])
        clave = (depto, legajo)
        tot = d.get("totales", {})
        if clave not in por_empleado:
            por_empleado[clave] = {
                "legajo": legajo, "nombre": d["nombre"],
                "departamento": _nombre_departamento_visible(depto),
                "ot50": timedelta(0), "ot100": timedelta(0),
                "comidas": 0, "francos": 0, "tardanzas": 0,
                "semanas": [], "dias": [],
                "conf_sem": set(), "pend_sem": set(),
                "excluido_ot": d.get("excluido_ot", False),
                "liquida_ot": d.get("liquida_ot", True),
                "observacion_liquidacion": d.get("observacion_liquidacion", ""),
            }
        e = por_empleado[clave]
        if sem in e["conf_sem"] or sem in e["pend_sem"]:
            continue
        e["ot50"]      += _parse_hm(tot.get("ot50", "0h"))
        e["ot100"]     += _parse_hm(tot.get("ot100", "0h"))
        e["comidas"]   += tot.get("comidas", 0)
        e["francos"]   += sum(1 for dia in d.get("dias", []) if dia.get("franco"))
        e["tardanzas"] += tot.get("tardanzas", 0)
        sem_vis = semana_visible.get(sem, sem)
        if sem_vis not in e["semanas"]: e["semanas"].append(sem_vis)
        e["pend_sem"].add(sem)


    excluidos  = _cargar_excluidos_ot()
    resultado = sorted([
        {
            "legajo":      e["legajo"],
            "nombre":      e["nombre"],
            "departamento": e["departamento"],
            "ot50":        _fmt_hm(e["ot50"]),
            "ot100":       _fmt_hm(e["ot100"]),
            "comidas":     e["comidas"],
            "francos":     e["francos"],
            "tardanzas":   e["tardanzas"],
            "semanas":     sorted(set(e["semanas"])),
            "confirmado":  len(e["pend_sem"]) == 0,
            "pendientes":  sorted(e["pend_sem"]),
            "dias":        sorted(e["dias"], key=lambda d: d["fecha"]),
            "excluido_ot": (
                e.get("excluido_ot", False)
                or str(e["legajo"]) in excluidos
            ),
            "liquida_ot":  str(e["legajo"]) not in excluidos,
            "observacion_liquidacion": (
                _OBS_EXCLUIDO_OT if str(e["legajo"]) in excluidos
                else e.get("observacion_liquidacion", "")
            ),
        }
        for e in por_empleado.values()
        if str(e["legajo"]) not in LEGAJOS_EXCLUIR_INFORMES
    ], key=lambda x: _legajo_key(x))

    # Agregar empleados conocidos del depto sin actividad en el período (aparecen con ceros)
    if departamento:
        presentes = {e["legajo"] for e in resultado}
        for emp in _empleados_conocidos():
            if _normalizar_departamento_web(emp["departamento"]) != departamento:
                continue
            if emp["legajo"] in presentes or emp["legajo"] in LEGAJOS_EXCLUIR_INFORMES:
                continue
            excl = emp["legajo"] in excluidos
            resultado.append({
                "legajo":      emp["legajo"],
                "nombre":      emp["nombre"],
                "departamento": emp["departamento"],
                "ot50":        "0h0m",
                "ot100":       "0h0m",
                "comidas":     0,
                "francos":     0,
                "tardanzas":   0,
                "semanas":     [],
                "confirmado":  True,
                "pendientes":  [],
                "dias":        [],
                "excluido_ot": excl,
                "liquida_ot":  not excl,
                "observacion_liquidacion": _OBS_EXCLUIDO_OT if excl else "",
            })
        resultado.sort(key=lambda x: _legajo_key(x))

    return resultado


def _recalcular_horas_periodo(pid):
    """Recalcula OT50/OT100/comidas/francos/tardanzas de un cierre ya hecho,
    directo desde los CSV de las semanas incluidas (igual que la sección de
    detalle día a día de _generar_pdf_cierre_completo), sin pasar por
    confirmaciones/historial.

    Sirve para diagnosticar y corregir un cierre cuyo resumen quedó con
    datos viejos — por ejemplo si una confirmación archivada de un período
    ANULADO fue "readoptada" por fecha (_resolver_semana_confirmacion) antes
    de que existiera el filtro de _archivos_confirmacion que excluye
    períodos anulados. Los tokens del cierre ya no existen en sesión.json
    (se borran al cerrar), así que ésta es la única fuente confiable una vez
    cerrado. Retorna None si no hay CSV disponibles para ninguna semana."""
    with _get_db() as conn:
        periodo = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not periodo:
            return None
        emp_rows = conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=?", (pid,)
        ).fetchall()
    if not emp_rows:
        return None
    legajos_cierre = {str(r["legajo"]) for r in emp_rows}

    periodo_json_path = PERIODOS_DIR / (periodo["archivo"] or "NADA")
    semanas_numeros = []
    if periodo_json_path.exists():
        try:
            pdata = json.loads(periodo_json_path.read_text(encoding="utf-8"))
            semanas_numeros = pdata.get("semanas", [])
        except Exception:
            pass
    if not semanas_numeros:
        semanas_numeros = list(range(periodo["semana_desde"] or 1, (periodo["semana_hasta"] or 1) + 1))

    totales = {}
    huno_csv = False
    for n_sem in semanas_numeros:
        df_sem = _cargar_semana_csv(n_sem)
        if df_sem is None:
            continue
        huno_csv = True
        try:
            emps_proc, _, _ = _procesar_empleados(_normalizar_columnas(df_sem))
        except Exception:
            continue
        for emp in emps_proc:
            leg = str(emp["legajo"])
            if leg not in legajos_cierre:
                continue
            dias_prep = _preparar_dias(emp["registros"])
            t = totales.setdefault(leg, {
                "nombre": emp["nombre"], "ot50": timedelta(0), "ot100": timedelta(0),
                "comidas": 0, "francos": 0, "tardanzas": 0,
            })
            for d in dias_prep:
                t["ot50"]      += _parse_td(d["ot50"])
                t["ot100"]     += _parse_td(d["ot100"])
                t["comidas"]   += int(d.get("comida", 0))
                t["francos"]   += int(d.get("franco", 0))
                t["tardanzas"] += int(d.get("tarde", 0))
    if not huno_csv:
        return None
    return totales


@app.route("/admin/semanas-de-periodo/<int:pid>")
def admin_semanas_de_periodo(pid):
    """Solo lectura: lista los números de semana GLOBALES (los que usa
    semanas/semana_N.csv en disco) que componen el cierre #pid, con sus
    fechas, leídos del respaldo JSON del período. Sirve para saber qué
    archivo semana_N.csv hay que reemplazar con /admin/reemplazar-csv-semana
    para corregir un dato de origen en un cierre ya cerrado (las semanas ya
    no figuran "activas" en el panel una vez cerrado el período, pero el
    CSV sigue en disco y _recalcular_horas_periodo lo sigue leyendo por
    número)."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
    semanas_numeros = []
    periodo_json_path = PERIODOS_DIR / (p["archivo"] or "NADA")
    if periodo_json_path.exists():
        try:
            pdata = json.loads(periodo_json_path.read_text(encoding="utf-8"))
            semanas_numeros = pdata.get("semanas", [])
        except Exception:
            pass
    if not semanas_numeros:
        semanas_numeros = list(range(p["semana_desde"] or 1, (p["semana_hasta"] or 1) + 1))

    semanas = []
    for n in semanas_numeros:
        existe_csv = (SEMANAS_DIR / f"semana_{n}.csv").exists()
        semanas.append({"numero": n, "csv_en_disco": existe_csv})
    return jsonify({"periodo_id": pid, "fecha_desde": p["fecha_desde"], "fecha_hasta": p["fecha_hasta"], "semanas": semanas})


@app.route("/admin/verificar-cadena-saldos-francos")
def admin_verificar_cadena_saldos_francos():
    """Solo lectura: para cada departamento, recorre los cierres de periodos
    ACTIVOS (no anulados) ordenados por cerrado_en y, para cada legajo
    presente en dos cierres consecutivos del mismo depto, recalcula el
    "saldo final" que ese cierre debería haber dejado (usando SOLO datos
    guardados de forma histórica e inmutable: periodos.saldo_anterior del
    cierre y su propio _delta_francos_cierre) y lo compara contra el
    saldo_anterior guardado del cierre siguiente.

    No depende del estado actual/en vivo de francos_generados ni de
    francos_semana_parcial -- por eso puede detectar si algún cierre
    histórico rompió la cadena, sin que un ajuste posterior lo enmascare.
    Reporta cada desajuste encontrado; lista vacía significa que la cadena
    está sana en todos los cierres activos."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    with _get_db() as conn:
        rows = conn.execute("""
            SELECT DISTINCT p.id, p.cerrado_en, p.saldo_anterior, pe.departamento
            FROM periodos p
            JOIN periodo_empleados pe ON pe.periodo_id = p.id
            WHERE COALESCE(p.estado,'ACTIVO') <> 'ANULADO'
        """).fetchall()

        por_depto = {}
        for r in rows:
            depto = _normalizar_departamento_web(r["departamento"] or "")
            por_depto.setdefault(depto, {})[r["id"]] = {
                "cerrado_en": r["cerrado_en"] or "",
                "saldo_anterior": json.loads(r["saldo_anterior"] or "{}"),
            }

        desajustes = []
        cierres_revisados = 0
        for depto, cierres in por_depto.items():
            orden = sorted(cierres.keys(), key=lambda pid: cierres[pid]["cerrado_en"])
            for i in range(len(orden) - 1):
                pid_actual, pid_sig = orden[i], orden[i + 1]
                sa_actual = cierres[pid_actual]["saldo_anterior"]
                sa_sig    = cierres[pid_sig]["saldo_anterior"]
                legajos_comunes = set(sa_actual.keys()) & set(sa_sig.keys())
                if not legajos_comunes:
                    continue
                cierres_revisados += 1
                deltas = _delta_francos_cierre(conn, pid_actual, legajos_comunes)
                for leg in sorted(legajos_comunes):
                    base = sa_actual[leg]
                    despues = sa_sig[leg]
                    d = deltas.get(leg, {"generados_periodo": 0, "tomados_periodo": 0})
                    gen_extra_delta = despues.get("gen_extra_al_corte", 0) - base.get("gen_extra_al_corte", 0)
                    saldo_final_recalculado = (
                        base.get("saldo", 0) + d["generados_periodo"] + gen_extra_delta - d["tomados_periodo"]
                    )
                    if saldo_final_recalculado != despues.get("saldo", 0):
                        desajustes.append({
                            "departamento": depto,
                            "legajo": leg,
                            "cierre_anterior_id": pid_actual,
                            "cierre_siguiente_id": pid_sig,
                            "saldo_final_recalculado": saldo_final_recalculado,
                            "saldo_anterior_declarado_en_siguiente": despues.get("saldo", 0),
                        })

    return jsonify({
        "cierres_consecutivos_revisados": cierres_revisados,
        "cadena_sana": not desajustes,
        "desajustes": desajustes,
    })


@app.route("/admin/reemplazar-csv-semana/<int:n>", methods=["GET", "POST"])
def admin_reemplazar_csv_semana(n):
    """Reemplaza semanas/semana_N.csv (el CSV interno que el sistema ya
    guardó al procesar esa semana) con un Excel/CSV corregido, sin importar
    si esa semana sigue "activa" en el panel. Sirve para corregir un dato de
    fichadas de una semana que ya forma parte de un cierre cerrado: después
    de subir el archivo corregido acá, hay que volver a correr
    /admin/recalcular-horas-cierre/<pid>?confirmar=si del cierre que
    incluye esta semana para que tome el dato nuevo.

    GET muestra un formulario simple de carga. POST recibe el archivo
    (campo "csv") y sobreescribe semana_N.csv. No toca metadata.json,
    periodo_empleados, francos_cierre_detalle ni ninguna otra tabla —
    solo el CSV en disco."""
    if not _autenticado(): return _requiere_auth()
    if request.method == "GET":
        return f"""
        <html><body style="font-family:sans-serif;max-width:480px;margin:40px auto">
        <h3>Reemplazar semana_{n}.csv</h3>
        <p>Sube el Excel/CSV corregido de esta semana. Va a sobreescribir
        <code>semanas/semana_{n}.csv</code> tal cual. No cierra nada ni toca
        ningún cierre por sí solo.</p>
        <form method="post" enctype="multipart/form-data">
          <input type="file" name="csv" required>
          <button type="submit">Reemplazar</button>
        </form>
        </body></html>
        """
    if "csv" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    try:
        df = _normalizar_columnas(_leer_archivo(request.files["csv"]))
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    _guardar_semana_csv(n, df)
    return jsonify({"ok": True, "semana": n, "filas": len(df)})


@app.route("/admin/recalcular-horas-cierre/<int:pid>", methods=["GET", "POST"])
def admin_recalcular_horas_cierre(pid):
    """Recalcula las horas de un cierre directo desde los CSV de sus
    semanas y las compara contra lo guardado en periodo_empleados.

    GET = diagnóstico de solo lectura (diferencias por legajo, nada se
    escribe). POST con ?confirmar=si = aplica: UPDATE de periodo_empleados
    (ot50, ot100, comidas, francos, tardanzas) con los valores recalculados
    para los legajos con diferencia. No toca francos_cierre_detalle,
    confirmaciones, semanas ni francos_saldo_inicial."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    with _get_db() as conn:
        periodo = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not periodo:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        emp_rows = conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=?", (pid,)
        ).fetchall()
    if not emp_rows:
        return jsonify({"error": f"Período #{pid} no tiene empleados asociados"}), 400

    frescos = _recalcular_horas_periodo(pid)
    if frescos is None:
        return jsonify({"error": "No se pudo recalcular: no hay CSV de semanas disponibles para este cierre."}), 400

    diffs = []
    for r in emp_rows:
        leg = str(r["legajo"])
        fresco = frescos.get(leg)
        if fresco is None:
            continue
        ot50_actual  = _parse_hm(r["ot50"] or "0h")
        ot100_actual = _parse_hm(r["ot100"] or "0h")
        if (ot50_actual != fresco["ot50"] or ot100_actual != fresco["ot100"]
                or (r["comidas"] or 0) != fresco["comidas"]
                or (r["francos"] or 0) != fresco["francos"]
                or (r["tardanzas"] or 0) != fresco["tardanzas"]):
            diffs.append({
                "legajo": leg, "nombre": r["nombre"],
                "ot50_actual": r["ot50"], "ot50_correcto": _fmt_hm(fresco["ot50"]),
                "ot100_actual": r["ot100"], "ot100_correcto": _fmt_hm(fresco["ot100"]),
                "comidas_actual": r["comidas"] or 0, "comidas_correcto": fresco["comidas"],
                "francos_actual": r["francos"] or 0, "francos_correcto": fresco["francos"],
                "tardanzas_actual": r["tardanzas"] or 0, "tardanzas_correcto": fresco["tardanzas"],
            })

    if request.args.get("confirmar") != "si":
        return jsonify({"periodo_id": pid, "requiere_correccion": bool(diffs), "diferencias": diffs})

    if not diffs:
        return jsonify({"ok": True, "mensaje": "No había diferencias, no se modificó nada."})

    with _get_db() as conn:
        for d in diffs:
            fresco = frescos[d["legajo"]]
            conn.execute(
                "UPDATE periodo_empleados SET ot50=?, ot100=?, comidas=?, francos=?, tardanzas=? "
                "WHERE periodo_id=? AND legajo=?",
                (_fmt_hm(fresco["ot50"]), _fmt_hm(fresco["ot100"]), fresco["comidas"],
                 fresco["francos"], fresco["tardanzas"], pid, d["legajo"])
            )
        conn.commit()

    return jsonify({"ok": True, "corregidos": diffs})


@app.route("/periodo/resumen")
def periodo_resumen():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    desde = int(request.args.get("desde", 1))
    hasta = int(request.args.get("hasta", 1))
    departamento = _normalizar_departamento_web(request.args.get("departamento", ""))
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    if not departamento or not fecha_desde or not fecha_hasta:
        return jsonify({"error": "Departamento y rango de fechas requeridos."}), 400
    return jsonify(_calcular_periodo(desde, hasta, departamento))


@app.route("/periodo/exportar")
def periodo_exportar():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    import csv, io
    from flask import Response
    desde = int(request.args.get("desde", 1))
    hasta = int(request.args.get("hasta", 1))
    departamento = _normalizar_departamento_web(request.args.get("departamento", ""))
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    if not departamento or not fecha_desde or not fecha_hasta:
        return jsonify({"error": "Departamento y rango de fechas requeridos."}), 400
    resultado = _calcular_periodo(desde, hasta, departamento)

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Departamento","Legajo","Nombre","OT 50%","OT 100%",
                "Comidas","Francos","Tardanzas","Semanas","Estado"])
    dep_actual = None
    for e in sorted(resultado, key=lambda x: (x["departamento"] or "—", x["nombre"])):
        dep = e["departamento"] or "—"
        if dep != dep_actual:
            dep_actual = dep
            w.writerow([f"--- {dep} ---", "", "", "", "", "", "", "", "", ""])
        w.writerow([
            dep, e["legajo"], e["nombre"],
            e["ot50"], e["ot100"],
            e["comidas"], e["francos"], e["tardanzas"],
            " ".join(str(s) for s in e["semanas"]),
            "Confirmado" if e["confirmado"] else "Pendiente",
        ])
    out.seek(0)
    nombre_archivo = f"periodo_sem{desde}-{hasta}.csv"
    return Response(
        "﻿" + out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@app.route("/periodo/confirmaciones_pdf")
def periodo_confirmaciones_pdf():
    if not _autenticado(): return _requiere_auth()
    try:
        desde = int(request.args.get("desde", 1))
        hasta = int(request.args.get("hasta", 1))
    except ValueError:
        return "Parámetros inválidos.", 400
    departamento = _normalizar_departamento_web(request.args.get("departamento", ""))
    fecha_desde  = request.args.get("fecha_desde", "").strip()
    fecha_hasta  = request.args.get("fecha_hasta", "").strip()
    if not departamento or not fecha_desde or not fecha_hasta:
        return "Departamento y rango de fechas requeridos.", 400

    todos = _calcular_periodo(desde, hasta, departamento)

    # Fecha confirmación más reciente por empleado (del historial)
    hist = [c for c in _leer_historial(departamento=departamento)
            if desde <= c.get("semana", 0) <= hasta]
    conf_en_map = {}
    for c in hist:
        leg   = str(c.get("legajo", ""))
        fecha = c.get("confirmado_en") or ""
        if fecha > conf_en_map.get(leg, ""):
            conf_en_map[leg] = fecha

    # Días por legajo desde _sesion — solo filtramos por rango de semanas,
    # sin filtrar departamento (evita desajustes de normalización).
    # Dedup por fecha dentro de cada legajo.
    dias_por_leg = {}
    for d in _sesion.values():
        sem = d.get("semana", 0)
        if not (desde <= sem <= hasta):
            continue
        leg = str(d.get("legajo", ""))
        by_fecha = dias_por_leg.setdefault(leg, {})
        for x in d.get("dias", []):
            if not (x.get("tiene_ot") or x.get("comida") or x.get("franco")):
                continue
            f = x.get("fecha", "")
            # Si ya existe la fecha, mantener el registro con más información
            if f not in by_fecha or x.get("tiene_ot"):
                by_fecha[f] = x

    for e in todos:
        leg = str(e.get("legajo", ""))
        e["confirmado_en"] = conf_en_map.get(leg, "")
        e["dias"] = sorted(
            dias_por_leg.get(leg, {}).values(),
            key=lambda x: x.get("fecha", "")
        )

    info = {
        "departamento_label": _nombre_departamento_visible(departamento),
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }
    try:
        pdf_data = _generar_pdf_confirmaciones_parcial(todos, info)
        nombre = f"confirmaciones_{departamento}_{fecha_desde}_{fecha_hasta}.pdf"
        return send_file(BytesIO(pdf_data), mimetype="application/pdf",
                         as_attachment=False, download_name=nombre)
    except Exception as exc:
        return f"Error generando PDF: {exc}", 500


@app.route("/periodo/cerrar", methods=["POST"])
def periodo_cerrar():
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    desde = int(request.form.get("desde", 1))
    hasta = int(request.form.get("hasta", 1))
    departamento = _normalizar_departamento_web(request.form.get("departamento", ""))
    fecha_desde_req = request.form.get("fecha_desde", "").strip()
    fecha_hasta_req = request.form.get("fecha_hasta", "").strip()
    semana_visible_desde = request.form.get("semana_visible_desde", "").strip()
    semana_visible_hasta = request.form.get("semana_visible_hasta", "").strip()
    if not departamento or not fecha_desde_req or not fecha_hasta_req:
        return jsonify({"error": "SeleccionÃ¡ un departamento para cerrar el perÃ­odo."}), 400
    with _get_db() as conn:
        duplicado = conn.execute("""
            SELECT 1
            FROM periodos p
            JOIN periodo_empleados pe ON pe.periodo_id = p.id
            WHERE p.fecha_desde = ?
              AND p.fecha_hasta = ?
              AND pe.departamento = ?
              AND COALESCE(p.estado, 'ACTIVO') <> 'ANULADO'
            LIMIT 1
        """, (fecha_desde_req, fecha_hasta_req, _nombre_departamento_visible(departamento))).fetchone()
    if duplicado:
        return jsonify({"error": "Ya existe un cierre para ese departamento y rango."}), 400

    resumen = _calcular_periodo(desde, hasta, departamento)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ahora = datetime.now().isoformat()
    # Fecha/hora real del cierre, sin truncar microsegundos -- mismo valor
    # que se guarda en periodos.cerrado_en, para que _fecha_corte_anterior
    # compare cortes consistentes entre cierres del mismo segundo
    # (criterio "ya estaba cargado antes del cierre")
    fecha_corte = _normalizar_cargado_en(ahora)
    archivo = f"periodo_{ts}.json"

    # Extraer rango de fechas de las semanas incluidas
    meta = _cargar_metadata()
    _fdlist, _fhlist = [], []
    semanas_cerradas = []
    for s in meta.get("semanas", []):
        depto = _normalizar_departamento_web(s.get("departamento", "") or "Todos")
        if depto != departamento:
            continue
        try:
            n_global = int(s.get("numero", 0))
            if desde <= n_global <= hasta:
                semanas_cerradas.append(n_global)
                if s.get("fecha_desde"): _fdlist.append(s["fecha_desde"])
                if s.get("fecha_hasta"): _fhlist.append(s["fecha_hasta"])
        except (TypeError, ValueError):
            pass
    fecha_desde_p = fecha_desde_req or (min(_fdlist) if _fdlist else "")
    fecha_hasta_p = fecha_hasta_req or (max(_fhlist) if _fhlist else "")
    if not semanas_cerradas:
        return jsonify({"error": "No hay semanas activas de ese departamento en el rango seleccionado."}), 400
    try:
        visible_desde_int = int(semana_visible_desde) if semana_visible_desde else 1
    except Exception:
        visible_desde_int = 1
    semanas_visibles_mapa = {
        int(n): visible_desde_int + i
        for i, n in enumerate(semanas_cerradas)
    }

    # Guardar JSON de respaldo
    PERIODOS_DIR.mkdir(exist_ok=True)
    (PERIODOS_DIR / archivo).write_text(
        json.dumps({"cerrado_en": ahora,
                    "estado": "ACTIVO",
                    "departamento": departamento,
                    "semana_visible_desde": semana_visible_desde,
                    "semana_visible_hasta": semana_visible_hasta,
                    "semanas": semanas_cerradas,
                    "fecha_desde": fecha_desde_p, "fecha_hasta": fecha_hasta_p,
                    "empleados": resumen},
                   ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # Guardar en base de datos
    with _get_db() as conn:
        cur = conn.execute(
            "INSERT INTO periodos (cerrado_en, semana_desde, semana_hasta, archivo, fecha_desde, fecha_hasta, estado) VALUES (?,?,?,?,?,?,?)",
            (ahora, desde, hasta, archivo, fecha_desde_p, fecha_hasta_p, "ACTIVO")
        )
        pid = cur.lastrowid
        for e in resumen:
            conn.execute(
                "INSERT INTO periodo_empleados (periodo_id,legajo,nombre,departamento,ot50,ot100,comidas,francos,tardanzas,semanas,confirmado) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (pid, str(e.get("legajo","")), e.get("nombre",""), e.get("departamento",""),
                 e.get("ot50","0h"), e.get("ot100","0h"),
                 e.get("comidas",0), e.get("francos",0), e.get("tardanzas",0),
                 json.dumps(sorted(semanas_visibles_mapa.values())),
                 1 if e.get("confirmado") else 0)
            )
        # Snapshot de francos tomados — legajos del departamento cerrado,
        # incluyendo empleados sin fichadas (ej. Karen Soto en Redes) para que
        # sus francos también entren al snapshot de su propio cierre y a la
        # actualización de saldo (antes se agregaban recién en el bloque de
        # saldo, después de este snapshot, y quedaban afuera).
        legajos_cierre = [str(e.get("legajo", "")) for e in resumen]
        _set_legajos_cierre = set(legajos_cierre)
        for _row in conn.execute("SELECT legajo, departamento FROM empleados_extra"):
            if _normalizar_departamento_web(_row["departamento"] or "") == departamento:
                _leg = str(_row["legajo"])
                if _leg not in _set_legajos_cierre:
                    legajos_cierre.append(_leg)
                    _set_legajos_cierre.add(_leg)
        _snapshot_francos_cierre(conn, pid, fecha_corte,
                                 legajos=legajos_cierre, departamento=departamento)
        # Borrar parciales de francos de las semanas cerradas
        conn.execute(
            "DELETE FROM francos_semana_parcial WHERE semana_num BETWEEN ? AND ?",
            (desde, hasta)
        )
        conn.commit()

    # Limpiar solo los tokens del rango cerrado
    tokens_borrar = [t for t, d in _sesion.items()
                     if desde <= d.get("semana", 0) <= hasta
                     and _normalizar_departamento_web(d.get("departamento", "") or "Todos") == departamento]
    for t in tokens_borrar:
        del _sesion[t]
    _guardar_sesion(_sesion)

    # Archivar solo las confirmaciones del rango cerrado
    archivo_dir = CONFIRM_DIR / f"periodo_{ts}"
    if CONFIRM_DIR.exists():
        archivo_dir.mkdir(exist_ok=True)
        for f in CONFIRM_DIR.glob("*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if (desde <= data.get("semana", 0) <= hasta
                        and _normalizar_departamento_web(data.get("departamento", "") or "Todos") == departamento):
                    f.rename(archivo_dir / f.name)
            except Exception:
                pass

    # Actualizar metadata: solo quitar las semanas cerradas
    meta = _cargar_metadata()
    meta["semanas"] = [
        s for s in meta.get("semanas", [])
        if not (
            desde <= s.get("numero", 0) <= hasta
            and _normalizar_departamento_web(s.get("departamento", "") or "Todos") == departamento
        )
    ]
    meta["semana_actual"] = max((s.get("numero", 0) for s in meta["semanas"]), default=0)
    _guardar_metadata(meta)

    # ── Actualización automática de saldo_inicial ──────────────────────────
    # Se ejecuta DESPUÉS de limpiar sesión y parciales. Usa _delta_francos_cierre
    # (solo lo que ESTE cierre #pid aportó a periodo_empleados y a su propio
    # francos_cierre_detalle) en vez de _calcular_saldos() (saldo "en vivo"
    # global que también suma francos_semana_parcial de cualquier semana/depto
    # en curso sin cerrar todavía). Así el saldo grabado nunca mezcla datos de
    # otro período/mes/depto que esté en curso al momento de este cierre —
    # ver CLAUDE.md, regla de la cadena saldo_final == saldo_inicial siguiente.
    try:
        legajos_cierre_set = set(legajos_cierre)
        ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nombres_conocidos = {str(e["legajo"]): e["nombre"] for e in _empleados_conocidos()}

        with _get_db() as conn:
            # Guardar snapshot de saldos anteriores para poder revertir al anular
            saldo_ant = {}
            for leg in legajos_cierre_set:
                row = conn.execute(
                    "SELECT saldo, tomados_al_corte, gen_extra_al_corte, fecha_corte "
                    "FROM francos_saldo_inicial WHERE legajo=?", (leg,)
                ).fetchone()
                if row:
                    saldo_ant[leg] = {
                        "saldo":              row["saldo"],
                        "tomados_al_corte":   row["tomados_al_corte"]   or 0,
                        "gen_extra_al_corte": row["gen_extra_al_corte"] or 0,
                        "fecha_corte":        row["fecha_corte"]         or "2026-05-21",
                    }
                else:
                    saldo_ant[leg] = {"saldo": 0, "tomados_al_corte": 0,
                                      "gen_extra_al_corte": 0, "fecha_corte": "2026-05-21"}

            conn.execute(
                "UPDATE periodos SET saldo_anterior=? WHERE id=?",
                (json.dumps(saldo_ant, ensure_ascii=False), pid)
            )

            # Leer totales actuales de tomados y gen_extra para cada legajo del cierre
            tomados_totales = {r["legajo"]: (r["total"] or 0) for r in conn.execute(
                "SELECT legajo, SUM(dias) as total FROM francos_tomados WHERE COALESCE(estado,'') != 'Anulado' GROUP BY legajo"
            )}
            gen_extra_totales = {}
            for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_generados GROUP BY legajo"):
                gen_extra_totales[r["legajo"]] = gen_extra_totales.get(r["legajo"], 0) + (r["total"] or 0)
            for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_semana_manual GROUP BY legajo"):
                gen_extra_totales[r["legajo"]] = gen_extra_totales.get(r["legajo"], 0) + (r["total"] or 0)

            # Delta propio de este cierre: solo lo que quedó grabado bajo este
            # pid en periodo_empleados.francos y francos_cierre_detalle.
            deltas = _delta_francos_cierre(conn, pid, legajos_cierre_set)

            # Actualizar saldo_inicial con el saldo nuevo calculado
            for leg in legajos_cierre_set:
                base = saldo_ant[leg]
                d = deltas.get(leg, {"generados_periodo": 0, "tomados_periodo": 0})
                gen_extra_nuevo = max(0, gen_extra_totales.get(leg, 0) - base["gen_extra_al_corte"])
                saldo_nuevo = base["saldo"] + d["generados_periodo"] + gen_extra_nuevo - d["tomados_periodo"]

                emp_resumen = next((e for e in resumen if str(e.get("legajo","")) == leg), None)
                nombre = (emp_resumen.get("nombre") if emp_resumen else None) or nombres_conocidos.get(leg, leg)

                conn.execute("""
                    INSERT INTO francos_saldo_inicial
                        (legajo, nombre, saldo, nota, cargado_en, tomados_al_corte, gen_extra_al_corte, fecha_corte)
                    VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(legajo) DO UPDATE SET
                        saldo              = excluded.saldo,
                        nota               = excluded.nota,
                        cargado_en         = excluded.cargado_en,
                        tomados_al_corte   = excluded.tomados_al_corte,
                        gen_extra_al_corte = excluded.gen_extra_al_corte,
                        fecha_corte        = excluded.fecha_corte
                """, (
                    leg, nombre,
                    saldo_nuevo,
                    f"Actualizado automáticamente al cerrar período #{pid} ({fecha_corte})",
                    ahora_str,
                    tomados_totales.get(leg, 0),
                    gen_extra_totales.get(leg, 0),
                    fecha_corte,
                ))
            conn.commit()
        actualizados = len(legajos_cierre_set)
        print(f"[INFO] saldo_inicial actualizado para {actualizados} empleados del cierre #{pid} ({departamento})")
        saldo_warning = None
    except Exception as exc_saldo:
        import traceback
        tb = traceback.format_exc()
        saldo_warning = str(exc_saldo)
        print(f"[ERROR] saldo_inicial NO actualizado en cierre #{pid}: {exc_saldo}\n{tb}")

    return jsonify({"ok": True, "periodo_archivado": f"periodo_{ts}.json",
                    **({"saldo_warning": saldo_warning} if saldo_warning else {})})


@app.route("/periodos/historial")
def periodos_historial():
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        rows = conn.execute("""
            SELECT p.id, p.cerrado_en, p.semana_desde, p.semana_hasta,
                   p.fecha_desde, p.fecha_hasta,
                   COALESCE(p.estado, 'ACTIVO') as estado,
                   p.fecha_anulacion, p.motivo_anulacion, p.usuario_anulacion,
                   GROUP_CONCAT(DISTINCT pe.departamento) as departamentos,
                   COUNT(pe.id) as total,
                   SUM(pe.confirmado) as confirmados
            FROM periodos p
            LEFT JOIN periodo_empleados pe ON pe.periodo_id = p.id
            GROUP BY p.id
            ORDER BY p.id DESC
        """).fetchall()
        # Semanas visibles reales por período (ya guardadas como números visibles en periodo_empleados)
        sem_rows = conn.execute("SELECT periodo_id, semanas FROM periodo_empleados").fetchall()
    semanas_por_pid = {}
    for sr in sem_rows:
        pid = sr["periodo_id"]
        try:
            nums = json.loads(sr["semanas"] or "[]")
            semanas_por_pid.setdefault(pid, set()).update(nums)
        except Exception:
            pass
    with _get_db() as conn:
        fcd_rows = conn.execute(
            "SELECT periodo_id, COUNT(*) as cnt FROM francos_cierre_detalle GROUP BY periodo_id"
        ).fetchall()
    francos_count_por_pid = {r["periodo_id"]: r["cnt"] for r in fcd_rows}
    cierres = []
    for r in rows:
        total = r["total"] or 0
        conf  = r["confirmados"] or 0
        sems = sorted(semanas_por_pid.get(r["id"], set())) or list(range(r["semana_desde"], r["semana_hasta"]+1))
        francos_count = francos_count_por_pid.get(r["id"], 0)
        cierres.append({
            "id":          r["id"],
            "cerrado_en":  (r["cerrado_en"] or "")[:16].replace("T", " "),
            "semanas":     sems,
            "fecha_desde": r["fecha_desde"] or "",
            "fecha_hasta": r["fecha_hasta"] or "",
            "estado":      r["estado"] or "ACTIVO",
            "fecha_anulacion": (r["fecha_anulacion"] or "")[:16].replace("T", " "),
            "motivo_anulacion": r["motivo_anulacion"] or "",
            "usuario_anulacion": r["usuario_anulacion"] or "",
            "departamentos": r["departamentos"] or "",
            "total":       total,
            "confirmados": conf,
            "pendientes":  total - conf,
            "francos_count": francos_count,
            "francos_cerrados": francos_count > 0,
            "francos_snapshot_ok": francos_count > 0,
        })
    with _get_db() as conn:
        cf_rows = conn.execute(
            "SELECT * FROM cierres_francos ORDER BY id DESC"
        ).fetchall()
    cierres_francos = [dict(r) for r in cf_rows]
    ultimos_activos_cf = {}
    for c in cierres_francos:
        dep = _normalizar_departamento_web(c["departamento"] or "")
        if (c.get("estado") or "ACTIVO") != "ANULADO" and dep not in ultimos_activos_cf:
            ultimos_activos_cf[dep] = c["id"]
    for c in cierres_francos:
        c["cerrado_en"] = (c["cerrado_en"] or "")[:16].replace("T", " ")
        c["fecha_anulacion"] = (c.get("fecha_anulacion") or "")[:16].replace("T", " ")
        depto_visible = c["departamento"]
        depto_norm = _normalizar_departamento_web(depto_visible)
        c["es_ultimo_activo"] = ultimos_activos_cf.get(depto_norm) == c["id"]
        c["es_reversible"] = bool(c.get("base_anterior") and c.get("base_anterior") != "{}")
        with _get_db() as conn:
            legajos_cf = [str(r["legajo"]) for r in conn.execute(
                "SELECT DISTINCT legajo FROM periodo_empleados WHERE LOWER(departamento)=? "
                "UNION SELECT legajo FROM empleados_extra WHERE activo=1 AND LOWER(departamento)=?",
                (depto_visible.lower(), depto_norm)
            )]
            if legajos_cf:
                ph = ",".join("?" * len(legajos_cf))
                c["total_registros"] = conn.execute(
                    f"SELECT COUNT(*) FROM francos_tomados "
                    f"WHERE legajo IN ({ph}) AND COALESCE(estado,'') != 'Anulado' AND fecha_desde <= ?",
                    (*legajos_cf, c["fecha_hasta"])
                ).fetchone()[0]
            else:
                c["total_registros"] = 0
    deptos_conocidos = _departamentos_francos_disponibles()
    return render_template("periodos_historial.html", cierres=cierres,
                           cierres_francos=cierres_francos,
                           deptos_conocidos=deptos_conocidos)


_MESES_PLANILLA = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _cierres_francos_del_mes(conn, mes):
    """Último cierre activo del mes para cada departamento exportable."""
    encontrados = {}
    rows = conn.execute(
        "SELECT * FROM cierres_francos "
        "WHERE COALESCE(estado, 'ACTIVO') <> 'ANULADO' "
        "AND substr(fecha_hasta, 1, 7)=? ORDER BY id DESC",
        (mes,),
    ).fetchall()
    for row in rows:
        departamento = _normalizar_departamento_web(row["departamento"] or "")
        if departamento in ("guardias", "ingenieros") and departamento not in encontrados:
            encontrados[departamento] = dict(row)
    return encontrados


def _actualizar_planilla_francos(contenido, mes, cierres):
    """Actualiza exclusivamente Franco Orig. (G) y Franco Tom. (H)."""
    from openpyxl import load_workbook

    try:
        anio, numero_mes = (int(p) for p in mes.split("-"))
        nombre_hoja = f"{_MESES_PLANILLA[numero_mes]} {anio}"
    except (ValueError, KeyError):
        raise ValueError("El mes indicado no es válido.")
    try:
        libro = load_workbook(BytesIO(contenido), data_only=False)
    except Exception as exc:
        raise ValueError("No se pudo abrir la planilla. Verificá que sea un archivo .xlsx válido.") from exc
    if nombre_hoja not in libro.sheetnames:
        raise ValueError(f"La planilla no contiene la hoja '{nombre_hoja}'.")

    hoja = libro[nombre_hoja]
    filas_por_legajo = {}
    for fila in range(1, hoja.max_row + 1):
        valor = hoja.cell(fila, 1).value
        if valor is not None:
            legajo = str(valor).strip()
            if legajo.endswith(".0"):
                legajo = legajo[:-2]
            filas_por_legajo[legajo] = fila

    actualizados, faltantes = [], []
    for departamento in ("ingenieros", "guardias"):
        try:
            snapshot = json.loads(cierres[departamento].get("saldo_anterior") or "{}")
        except (TypeError, json.JSONDecodeError):
            snapshot = {}
        if not snapshot:
            raise ValueError(
                f"El cierre de {_nombre_departamento_visible(departamento)} no tiene datos de saldo para exportar."
            )
        for legajo, datos in snapshot.items():
            fila = filas_por_legajo.get(str(legajo))
            if not fila:
                faltantes.append(str(legajo))
                continue
            generados = datos.get("generados", 0) or 0
            tomados = datos.get("tomados", 0) or 0
            hoja.cell(fila, 7).value = generados if generados else None
            hoja.cell(fila, 8).value = tomados if tomados else None
            actualizados.append(str(legajo))

    if faltantes:
        raise ValueError("Faltan estos legajos en la hoja mensual: " + ", ".join(sorted(faltantes)))
    if not actualizados:
        raise ValueError("No se encontraron empleados para actualizar en la hoja mensual.")
    if getattr(libro, "calculation", None):
        libro.calculation.fullCalcOnLoad = True
        libro.calculation.forceFullCalc = True
    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida


@app.route("/francos/planilla/actualizar", methods=["POST"])
def francos_planilla_actualizar():
    if not _autenticado():
        return jsonify({"error": "No autorizado"}), 401
    mes = (request.form.get("mes") or "").strip()
    archivo = request.files.get("planilla")
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", mes):
        return jsonify({"error": "Seleccioná un mes válido."}), 400
    if not archivo or not (archivo.filename or "").lower().endswith(".xlsx"):
        return jsonify({"error": "Seleccioná la planilla Excel en formato .xlsx."}), 400

    with _get_db() as conn:
        cierres = _cierres_francos_del_mes(conn, mes)
    faltan = [_nombre_departamento_visible(dep)
              for dep in ("ingenieros", "guardias") if dep not in cierres]
    if faltan:
        return jsonify({
            "error": "No se puede actualizar: falta el cierre activo del mes para " + " y ".join(faltan) + "."
        }), 409
    try:
        salida = _actualizar_planilla_francos(archivo.read(), mes, cierres)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return send_file(
        salida, as_attachment=True,
        download_name=f"Horas_Extras_actualizada_{mes}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        etag=False, max_age=0,
    )


def _vincular_movimientos_cierre_francos(conn, cierre_id, legajos, fecha_hasta):
    """Asocia de forma irreversible (hasta una futura anulación controlada)
    los movimientos manuales pendientes con el cierre del departamento.

    - tomados y generados extraordinarios: por fecha de CARGA (cargado_en),
      no por la fecha tomada -- es lo que decide a qué cierre pertenece un
      movimiento, mismo criterio que _seleccionar_francos_ventana_cierre
      para el mecanismo de periodos;
    - semanas manuales: hasta el mes de la fecha de corte (no tienen
      cargado_en individual, se agrupan por mes).

    Los movimientos ya asociados a otro cierre nunca se reasignan.
    """
    legajos = [str(leg) for leg in legajos if str(leg)]
    if not legajos:
        return {"tomados": 0, "generados": 0, "semanales": 0}

    ph = ",".join("?" * len(legajos))
    mes_hasta = (fecha_hasta or "")[:7]

    cur_tom = conn.execute(
        f"UPDATE francos_tomados "
        f"SET cierre_francos_id=?, estado_antes_cierre=estado, estado='Cerrado' "
        f"WHERE legajo IN ({ph}) "
        f"AND cierre_francos_id IS NULL "
        f"AND COALESCE(estado,'') NOT IN ('Anulado','Cerrado') "
        f"AND {_SQL_CARGADO_EN_DIA} <= ?",
        (cierre_id, *legajos, fecha_hasta),
    )
    cur_gen = conn.execute(
        f"UPDATE francos_generados SET cierre_francos_id=? "
        f"WHERE legajo IN ({ph}) AND cierre_francos_id IS NULL AND {_SQL_CARGADO_EN_DIA} <= ?",
        (cierre_id, *legajos, fecha_hasta),
    )
    cur_sem = conn.execute(
        f"UPDATE francos_semana_manual SET cierre_francos_id=? "
        f"WHERE legajo IN ({ph}) AND cierre_francos_id IS NULL AND mes <= ?",
        (cierre_id, *legajos, mes_hasta),
    )
    return {
        "tomados": cur_tom.rowcount,
        "generados": cur_gen.rowcount,
        "semanales": cur_sem.rowcount,
    }


def _snapshot_base_saldo_manual(conn, legajos):
    """Foto exacta de francos_saldo_inicial antes del cierre manual."""
    snapshot = {}
    for legajo in (str(x) for x in legajos):
        row = conn.execute(
            "SELECT nombre, saldo, nota, cargado_en, tomados_al_corte, "
            "gen_extra_al_corte, fecha_corte FROM francos_saldo_inicial "
            "WHERE CAST(legajo AS TEXT)=?", (legajo,)
        ).fetchone()
        snapshot[legajo] = dict(row) if row else None
    return snapshot


@app.route("/francos/cierre/nuevo", methods=["POST"])
def francos_cierre_nuevo():
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    departamento = _normalizar_departamento_web(request.form.get("departamento", ""))
    fecha_hasta  = request.form.get("fecha_hasta", "").strip()
    if not departamento or not fecha_hasta:
        return jsonify({"error": "Completá departamento y fecha."}), 400
    if _depto_francos_oculto(departamento):
        return jsonify({"error": "Francos de este departamento está oculto temporalmente."}), 400
    depto_visible = _nombre_departamento_visible(departamento)
    ahora = datetime.now().isoformat()
    with _get_db() as conn:
        legajos = [str(r["legajo"]) for r in conn.execute(
            "SELECT DISTINCT legajo FROM periodo_empleados WHERE LOWER(departamento)=? "
            "UNION SELECT legajo FROM empleados_extra WHERE activo=1 AND LOWER(departamento)=?",
            (depto_visible.lower(), departamento)
        )]
        if not legajos:
            return jsonify({"error": f"Sin empleados para {depto_visible}."}), 400
        ph = ",".join("?" * len(legajos))
        # Calcular total días -- por fecha de CARGA (cargado_en), no por la
        # fecha tomada: es lo que decide a qué cierre pertenece un franco
        # (mismo criterio que _seleccionar_francos_ventana_cierre).
        total = conn.execute(
            f"SELECT COALESCE(SUM(dias),0) FROM francos_tomados "
            f"WHERE legajo IN ({ph}) AND COALESCE(estado,'') != 'Anulado' "
            f"AND {_SQL_CARGADO_EN_DIA} <= ?",
            (*legajos, fecha_hasta)
        ).fetchone()[0]
        # Registrar cierre
        cur = conn.execute(
            "INSERT INTO cierres_francos (cerrado_en, departamento, fecha_hasta, total_dias, estado) VALUES (?,?,?,?,?)",
            (ahora, depto_visible, fecha_hasta, total, "ACTIVO")
        )
        cid = cur.lastrowid
        base_anterior = _snapshot_base_saldo_manual(conn, legajos)
        conn.execute(
            "UPDATE cierres_francos SET base_anterior=? WHERE id=?",
            (json.dumps(base_anterior, ensure_ascii=False), cid),
        )
        _vincular_movimientos_cierre_francos(
            conn, cid, legajos, fecha_hasta
        )
        conn.commit()

    # Actualizar saldo_inicial con el saldo calculado AL CORTE (fecha_hasta) --
    # NO con _calcular_saldos() (saldo "en vivo", sin límite superior de
    # fecha, que mezclaría movimientos cargados después de este corte, ej.
    # francos de julio en un cierre de junio). Mismo criterio que
    # periodo_cerrar / _delta_francos_cierre para el mecanismo de periodos.
    try:
        ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mes_hasta = (fecha_hasta or "")[:7]
        with _get_db() as conn:
            # Por fecha de CARGA (cargado_en), no por la fecha tomada --
            # mismo criterio que _vincular_movimientos_cierre_francos.
            tomados_al_hasta = {r["legajo"]: (r["total"] or 0) for r in conn.execute(
                f"SELECT legajo, SUM(dias) as total FROM francos_tomados "
                f"WHERE legajo IN ({ph}) AND COALESCE(estado,'') != 'Anulado' "
                f"AND {_SQL_CARGADO_EN_DIA} <= ? "
                f"GROUP BY legajo", (*legajos, fecha_hasta)
            )}
            gen_extra_al_hasta = {}
            for r in conn.execute(
                f"SELECT legajo, SUM(dias) as total FROM francos_generados "
                f"WHERE legajo IN ({ph}) AND {_SQL_CARGADO_EN_DIA} <= ? GROUP BY legajo",
                (*legajos, fecha_hasta)
            ):
                gen_extra_al_hasta[r["legajo"]] = gen_extra_al_hasta.get(r["legajo"], 0) + (r["total"] or 0)
            # francos_semana_manual no tiene cargado_en individual -- se
            # agrupa por mes, igual que en _vincular_movimientos_cierre_francos.
            for r in conn.execute(
                f"SELECT legajo, SUM(dias) as total FROM francos_semana_manual "
                f"WHERE legajo IN ({ph}) AND mes <= ? GROUP BY legajo",
                (*legajos, mes_hasta)
            ):
                gen_extra_al_hasta[r["legajo"]] = gen_extra_al_hasta.get(r["legajo"], 0) + (r["total"] or 0)

            snap = {}
            for leg in legajos:
                base = base_anterior.get(leg) or {}
                base_saldo     = (base.get("saldo") or 0) if base else 0
                base_tom_corte = (base.get("tomados_al_corte") or 0) if base else 0
                base_gen_corte = (base.get("gen_extra_al_corte") or 0) if base else 0

                tomados_periodo   = max(0, tomados_al_hasta.get(leg, 0) - base_tom_corte)
                generados_periodo = max(0, gen_extra_al_hasta.get(leg, 0) - base_gen_corte)
                saldo_nuevo = base_saldo + generados_periodo - tomados_periodo

                emp_row = next((e for e in _empleados_conocidos() if str(e["legajo"]) == leg), None)
                nombre = emp_row["nombre"] if emp_row else ((base.get("nombre") if base else "") or leg)

                snap[leg] = {
                    "nombre": nombre, "saldo_anterior": base_saldo,
                    "generados": generados_periodo, "tomados": tomados_periodo,
                    "saldo_final": saldo_nuevo,
                }

                conn.execute("""
                    INSERT INTO francos_saldo_inicial
                        (legajo, nombre, saldo, nota, cargado_en, tomados_al_corte, gen_extra_al_corte, fecha_corte)
                    VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(legajo) DO UPDATE SET
                        saldo=excluded.saldo, nota=excluded.nota, cargado_en=excluded.cargado_en,
                        tomados_al_corte=excluded.tomados_al_corte,
                        gen_extra_al_corte=excluded.gen_extra_al_corte,
                        fecha_corte=excluded.fecha_corte
                """, (leg, nombre, saldo_nuevo,
                      f"Cierre manual francos {depto_visible} al {fecha_hasta}",
                      ahora_str, tomados_al_hasta.get(leg, 0),
                      gen_extra_al_hasta.get(leg, 0), fecha_hasta))

            conn.execute("UPDATE cierres_francos SET saldo_anterior=? WHERE id=?",
                         (json.dumps(snap, ensure_ascii=False), cid))
            conn.commit()
    except Exception as e:
        print(f"[ADVERTENCIA] Error actualizando saldo en cierre manual: {e}")

    # Generar PDF -- exactamente los francos que _vincular_movimientos_cierre_francos
    # ya asoció a este cierre (cierre_francos_id=cid), sin re-derivar el
    # filtro de fecha por separado (evita que ambos puedan divergir).
    with _get_db() as conn:
        rows = conn.execute(
            f"SELECT legajo, nombre, tipo, fecha_desde, fecha_hasta, fechas_sueltas, dias, estado, "
            f"fecha_emision, autorizado_por, observaciones FROM francos_tomados "
            f"WHERE cierre_francos_id = ? "
            f"ORDER BY CAST(legajo AS INTEGER), fecha_desde",
            (cid,)
        ).fetchall()
    francos_list = [{**dict(r), "departamento": depto_visible} for r in rows]
    _generar_pdf_francos_cierre(f"cf{cid}", francos_list, fecha_hasta)
    return jsonify({"ok": True, "id": cid, "total": total})


@app.route("/francos/cierre/anular/<int:cid>", methods=["POST"])
def francos_cierre_anular(cid):
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    motivo = request.form.get("motivo", "").strip()
    if not motivo:
        return jsonify({"error": "Indicá el motivo de la anulación."}), 400
    usuario = session.get("usuario", "") or FIRMA_SUPERVISOR or "sistema"
    ahora = datetime.now().isoformat()

    conn = _get_db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        cierre = conn.execute(
            "SELECT * FROM cierres_francos WHERE id=?", (cid,)
        ).fetchone()
        if not cierre:
            conn.rollback()
            return jsonify({"error": "Cierre no encontrado."}), 404
        if (cierre["estado"] or "ACTIVO") == "ANULADO":
            conn.rollback()
            return jsonify({"error": "El cierre ya estaba anulado."}), 400

        ultimo = conn.execute(
            "SELECT id FROM cierres_francos "
            "WHERE LOWER(departamento)=LOWER(?) "
            "AND COALESCE(estado,'ACTIVO') <> 'ANULADO' "
            "ORDER BY cerrado_en DESC, id DESC LIMIT 1",
            (cierre["departamento"],),
        ).fetchone()
        if not ultimo or ultimo["id"] != cid:
            conn.rollback()
            return jsonify({
                "error": "Solo se puede anular el último cierre activo del departamento."
            }), 409

        try:
            base_anterior = json.loads(cierre["base_anterior"] or "{}")
        except Exception:
            base_anterior = {}
        if not base_anterior:
            conn.rollback()
            return jsonify({
                "error": "Este cierre histórico no tiene una base reversible segura."
            }), 409

        for legajo, vals in base_anterior.items():
            if vals is None:
                conn.execute(
                    "DELETE FROM francos_saldo_inicial WHERE CAST(legajo AS TEXT)=?",
                    (str(legajo),),
                )
                continue
            conn.execute("""
                INSERT INTO francos_saldo_inicial
                    (legajo,nombre,saldo,nota,cargado_en,tomados_al_corte,
                     gen_extra_al_corte,fecha_corte)
                VALUES (?,?,?,?,?,?,?,?)
                ON CONFLICT(legajo) DO UPDATE SET
                    nombre=excluded.nombre, saldo=excluded.saldo,
                    nota=excluded.nota, cargado_en=excluded.cargado_en,
                    tomados_al_corte=excluded.tomados_al_corte,
                    gen_extra_al_corte=excluded.gen_extra_al_corte,
                    fecha_corte=excluded.fecha_corte
            """, (
                str(legajo), vals.get("nombre", ""), vals.get("saldo", 0),
                vals.get("nota", ""), vals.get("cargado_en", ""),
                vals.get("tomados_al_corte", 0), vals.get("gen_extra_al_corte", 0),
                vals.get("fecha_corte", "2026-05-21"),
            ))

        tomados = conn.execute(
            "UPDATE francos_tomados SET "
            "estado=COALESCE(NULLIF(estado_antes_cierre,''),'Aprobado'), "
            "estado_antes_cierre='', cierre_francos_id=NULL "
            "WHERE cierre_francos_id=?", (cid,)
        ).rowcount
        generados = conn.execute(
            "UPDATE francos_generados SET cierre_francos_id=NULL "
            "WHERE cierre_francos_id=?", (cid,)
        ).rowcount
        semanales = conn.execute(
            "UPDATE francos_semana_manual SET cierre_francos_id=NULL "
            "WHERE cierre_francos_id=?", (cid,)
        ).rowcount
        conn.execute("""
            UPDATE cierres_francos
            SET estado='ANULADO', fecha_anulacion=?, motivo_anulacion=?,
                usuario_anulacion=?
            WHERE id=?
        """, (ahora, motivo, usuario, cid))
        conn.commit()
        return jsonify({
            "ok": True,
            "desbloqueados": {
                "tomados": tomados, "generados": generados, "semanales": semanales,
            },
        })
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


@app.route("/francos/cierre/informe/<int:cid>")
def francos_cierre_informe(cid):
    """PDF con saldo al cierre: saldo anterior + generados + tomados + saldo final."""
    if not _autenticado(): return _requiere_auth()
    from fpdf import FPDF
    with _get_db() as conn:
        cf = conn.execute("SELECT * FROM cierres_francos WHERE id=?", (cid,)).fetchone()
    if not cf:
        return "Cierre no encontrado.", 404
    cf = dict(cf)
    snap = {}
    try:
        snap = json.loads(cf.get("saldo_anterior") or "{}")
    except Exception:
        pass
    if not snap:
        return "Este cierre no tiene snapshot de saldo guardado.", 404

    depto_visible = cf["departamento"]
    fecha_hasta   = cf["fecha_hasta"]
    cerrado_en    = (cf.get("cerrado_en") or "")[:16].replace("T", " ")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    try:
        import os
        font_dir = os.path.join(os.path.dirname(__file__), "recursos", "fonts")
        pdf.add_font("DejaVu", "",  os.path.join(font_dir, "DejaVuSans.ttf"),        uni=True)
        pdf.add_font("DejaVu", "B", os.path.join(font_dir, "DejaVuSans-Bold.ttf"),   uni=True)
        fam = "DejaVu"
    except Exception:
        fam = "Helvetica"

    # Título
    pdf.set_font(fam, "B", 13)
    pdf.cell(0, 8, f"SALDO DE FRANCOS AL CIERRE — {depto_visible}", ln=1, align="C")
    pdf.set_font(fam, "", 9)
    pdf.cell(0, 5, f"Período hasta: {fecha_hasta}   |   Cerrado: {cerrado_en}", ln=1, align="C")
    pdf.ln(4)

    # Cabecera tabla
    COLS = ["Leg.", "Nombre", "Saldo anterior", "Generados", "Tomados", "Saldo final"]
    ANCH = [14, 60, 32, 28, 24, 28]
    pdf.set_fill_color(200, 215, 240)
    pdf.set_font(fam, "B", 8)
    for col, ancho in zip(COLS, ANCH):
        pdf.cell(ancho, 7, col, 1, 0, "C", fill=True)
    pdf.ln()

    pdf.set_font(fam, "", 8)
    tot_ant = tot_gen = tot_tom = tot_fin = 0
    for leg, d in sorted(snap.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        sf = d.get("saldo_final", 0)
        if sf > 0:   pdf.set_fill_color(220, 252, 231)
        elif sf < 0: pdf.set_fill_color(254, 202, 202)
        else:        pdf.set_fill_color(241, 245, 249)
        ant = d.get("saldo_anterior", 0)
        gen = d.get("generados", 0)
        tom = d.get("tomados", 0)
        tot_ant += ant; tot_gen += gen; tot_tom += tom; tot_fin += sf
        pdf.cell(ANCH[0], 6, leg,            1, 0, "C", True)
        pdf.cell(ANCH[1], 6, d.get("nombre",""), 1, 0, "L", True)
        pdf.cell(ANCH[2], 6, str(ant),       1, 0, "C", True)
        pdf.cell(ANCH[3], 6, str(gen),       1, 0, "C", True)
        pdf.cell(ANCH[4], 6, str(tom),       1, 0, "C", True)
        pdf.cell(ANCH[5], 6, str(sf),        1, 1, "C", True)

    pdf.set_fill_color(23, 32, 51); pdf.set_text_color(255, 255, 255)
    pdf.set_font(fam, "B", 8)
    pdf.cell(ANCH[0] + ANCH[1], 6, "TOTALES", 1, 0, "R", True)
    pdf.cell(ANCH[2], 6, str(tot_ant), 1, 0, "C", True)
    pdf.cell(ANCH[3], 6, str(tot_gen), 1, 0, "C", True)
    pdf.cell(ANCH[4], 6, str(tot_tom), 1, 0, "C", True)
    pdf.cell(ANCH[5], 6, str(tot_fin), 1, 1, "C", True)
    pdf.set_text_color(0, 0, 0)

    # Notas de ajuste manual al pie
    legajos_snap = list(snap.keys())
    if legajos_snap:
        ph_n = ",".join("?" * len(legajos_snap))
        with _get_db() as conn:
            notas_aj = [(str(r["leg"]), snap.get(str(r["leg"]), {}).get("nombre", str(r["leg"])), r["nota"])
                        for r in conn.execute(
                            f"SELECT CAST(legajo AS TEXT) as leg, nota FROM francos_saldo_inicial "
                            f"WHERE CAST(legajo AS TEXT) IN ({ph_n}) AND nota LIKE 'Ajuste manual%'",
                            legajos_snap)]
        if notas_aj:
            pdf.ln(6)
            pdf.set_font(fam, "B", 7); pdf.set_text_color(90, 90, 90)
            pdf.cell(0, 5, "Ajustes manuales de saldo inicial:", ln=1)
            pdf.set_font(fam, "", 7)
            for leg_n, nom_n, nota_n in notas_aj:
                pdf.cell(0, 4, f"  {leg_n} {nom_n}: {nota_n}", ln=1)
            pdf.set_text_color(0, 0, 0)

    raw = pdf.output(dest="S")
    data = raw.encode("latin-1") if isinstance(raw, str) else bytes(raw)
    from flask import Response
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition": f"inline; filename=saldo_cierre_{depto_visible}_{fecha_hasta}.pdf"})


@app.route("/francos/cierre/pdf/<int:cid>")
def francos_cierre_pdf(cid):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        cf = conn.execute("SELECT * FROM cierres_francos WHERE id=?", (cid,)).fetchone()
    if not cf:
        return "Cierre no encontrado.", 404
    cf = dict(cf)
    depto_visible = cf["departamento"]
    fecha_hasta   = cf["fecha_hasta"]
    import glob as _glob
    matches = sorted(_glob.glob(str(Path(f"reportes/francos_cierre_cf{cid}_*.pdf"))))
    # Un cierre anulado conserva el PDF producido al cerrarlo. Nunca se
    # regenera con los movimientos ya desbloqueados o corregidos.
    if (cf.get("estado") or "ACTIVO") == "ANULADO":
        if not matches:
            return "El PDF histórico de este cierre anulado no está disponible.", 404
        return send_file(matches[-1], mimetype="application/pdf", as_attachment=False,
                         download_name=f"francos_{depto_visible.lower()}_{fecha_hasta}.pdf")
    departamento  = _normalizar_departamento_web(depto_visible)
    with _get_db() as conn:
        legajos = [str(r["legajo"]) for r in conn.execute(
            "SELECT DISTINCT legajo FROM periodo_empleados WHERE LOWER(departamento)=? "
            "UNION SELECT legajo FROM empleados_extra WHERE activo=1 AND LOWER(departamento)=?",
            (depto_visible.lower(), departamento)
        )]
    if not legajos:
        return "Sin empleados.", 404
    ph = ",".join("?" * len(legajos))
    with _get_db() as conn:
        rows = conn.execute(
            f"SELECT legajo, nombre, tipo, fecha_desde, fecha_hasta, fechas_sueltas, dias, estado, "
            f"fecha_emision, autorizado_por, observaciones FROM francos_tomados "
            f"WHERE legajo IN ({ph}) AND COALESCE(estado,'') != 'Anulado' AND fecha_desde <= ? "
            f"ORDER BY CAST(legajo AS INTEGER), fecha_desde",
            (*legajos, fecha_hasta)
        ).fetchall()
    francos_list = [{**dict(r), "departamento": depto_visible} for r in rows]
    _generar_pdf_francos_cierre(f"cf{cid}", francos_list, fecha_hasta)
    matches = sorted(_glob.glob(str(Path(f"reportes/francos_cierre_cf{cid}_*.pdf"))))
    if not matches:
        return "Error generando PDF.", 500
    return send_file(matches[-1], mimetype="application/pdf", as_attachment=False,
                     download_name=f"francos_{depto_visible.lower()}_{fecha_hasta}.pdf")


@app.route("/periodos/anular/<int:pid>", methods=["POST"])
def periodo_anular(pid):
    if not _autenticado(): return jsonify({"error":"No autorizado"}), 401
    motivo = request.form.get("motivo", "").strip()
    usuario = session.get("usuario", "") or "sistema"
    ahora = datetime.now().isoformat()

    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return jsonify({"error": "Cierre no encontrado."}), 404
        if (p["estado"] or "ACTIVO") == "ANULADO":
            return jsonify({"error": "El cierre ya estaba anulado."}), 400

        conn.execute("""
            UPDATE periodos
            SET estado='ANULADO',
                fecha_anulacion=?,
                motivo_anulacion=?,
                usuario_anulacion=?
            WHERE id=?
        """, (ahora, motivo, usuario, pid))
        _revertir_estado_francos_cierre(conn, pid)
        conn.commit()

    archivo = p["archivo"]
    if archivo:
        json_path = PERIODOS_DIR / archivo
        if json_path.exists():
            try:
                data = json.loads(json_path.read_text(encoding="utf-8"))
                data["estado"] = "ANULADO"
                data["fecha_anulacion"] = ahora
                data["motivo_anulacion"] = motivo
                data["usuario_anulacion"] = usuario
                json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                pass

    # ── Restaurar saldo_inicial si el cierre tenía snapshot ─────────────────
    try:
        saldo_ant_raw = p["saldo_anterior"] if "saldo_anterior" in p.keys() else None
        if saldo_ant_raw:
            saldo_ant = json.loads(saldo_ant_raw)
            if saldo_ant:
                ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                with _get_db() as conn:
                    for leg, vals in saldo_ant.items():
                        conn.execute("""
                            UPDATE francos_saldo_inicial
                            SET saldo=?, tomados_al_corte=?, gen_extra_al_corte=?,
                                fecha_corte=?, nota=?, cargado_en=?
                            WHERE legajo=?
                        """, (
                            vals.get("saldo", 0),
                            vals.get("tomados_al_corte", 0),
                            vals.get("gen_extra_al_corte", 0),
                            vals.get("fecha_corte", "2026-05-21"),
                            f"Restaurado al anular período #{pid}",
                            ahora_str,
                            leg,
                        ))
                    conn.commit()
    except Exception as exc_rest:
        print(f"[ADVERTENCIA] Error al restaurar saldo_inicial al anular #{pid}: {exc_rest}")

    return jsonify({"ok": True})


@app.route("/periodos/ver/<int:pid>")
def periodos_ver(pid):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return "Período no encontrado", 404
        rows = conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=? ORDER BY departamento, CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()
    empleados = []
    for e in rows:
        if str(e["legajo"]) in LEGAJOS_EXCLUIR_INFORMES:
            continue
        d = dict(e)
        d["semanas"]   = json.loads(d["semanas"] or "[]")
        d["confirmado"] = bool(d["confirmado"])
        empleados.append(d)
    semanas = sorted({s for e in empleados for s in e.get("semanas", [])})
    deptos_cierre = sorted({e["departamento"] for e in empleados if e.get("departamento")})
    departamento = deptos_cierre[0] if len(deptos_cierre) == 1 else "Todos los departamentos"
    fd = p["fecha_desde"] or ""
    fh = p["fecha_hasta"] or ""
    francos_cierre = []
    with _get_db() as conn2:
        rows_ft = conn2.execute("""
            SELECT legajo, nombre, tipo, fecha_desde, fecha_hasta, fechas_sueltas,
                   dias, estado, fecha_emision, autorizado_por, observaciones
            FROM francos_cierre_detalle
            WHERE periodo_id = ?
            ORDER BY departamento, CAST(legajo AS INTEGER), fecha_desde
        """, (pid,)).fetchall()
    for r in rows_ft:
        d = dict(r)
        if d["tipo"] == "SUELTAS":
            try:
                d["fechas_lista"] = json.loads(d["fechas_sueltas"])
            except Exception:
                d["fechas_lista"] = []
        else:
            d["fechas_lista"] = []
        francos_cierre.append(d)
    with _get_db() as conn3:
        devoluciones = [dict(r) for r in conn3.execute(
            "SELECT * FROM francos_anulaciones_cerrados WHERE periodo_aplicado_id=? "
            "ORDER BY CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()]
    return render_template("periodo_detalle.html",
                           pid=pid,
                           departamento=departamento,
                           cerrado_en=(p["cerrado_en"] or "")[:16].replace("T", " "),
                           fecha_desde=fd,
                           fecha_hasta=fh,
                           semanas=semanas,
                           empleados=empleados,
                           francos_cierre=francos_cierre,
                           devoluciones=devoluciones,
                           firma=FIRMA_SUPERVISOR)


def _generar_pdf_cierre_completo(pid):
    """PDF completo del cierre: horas día a día (si existen) + resumen totales
    + saldo de francos + detalle de francos tomados. Un depto por cierre."""
    from pdf_generator import (PDFGeneral, _td_from_any, formato_horas, _round_to_hour,
                                LEGAJOS_EXCLUIR_INFORMES, _total_departamento)

    # ── Cargar datos ──────────────────────────────────────────────────────
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return None
        p = dict(p)
        emps_db = [dict(r) for r in conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=? ORDER BY CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()]
        ft_db = [dict(r) for r in conn.execute(
            "SELECT * FROM francos_cierre_detalle WHERE periodo_id=? "
            "ORDER BY CAST(legajo AS INTEGER), fecha_desde",
            (pid,)
        ).fetchall()]

    saldo_ant = {}
    try:
        saldo_ant = json.loads(p.get("saldo_anterior") or "{}")
    except Exception:
        pass

    departamento_norm = _normalizar_departamento_web(
        emps_db[0].get("departamento", "") if emps_db else ""
    )
    depto_visible = _nombre_departamento_visible(departamento_norm) or "Sin departamento"
    fd = p.get("fecha_desde", "")
    fh = p.get("fecha_hasta", "")
    legajos_cierre = {str(e["legajo"]) for e in emps_db}

    # Agregar empleados del depto con saldo conocido pero sin fichadas en este cierre
    for _emp in _empleados_conocidos():
        if _normalizar_departamento_web(_emp.get("departamento", "")) == departamento_norm:
            _leg = str(_emp["legajo"])
            if _leg not in legajos_cierre:
                emps_db.append({"legajo": _leg, "nombre": _emp["nombre"],
                                 "departamento": depto_visible, "francos": 0,
                                 "ot50": "0h", "ot100": "0h", "comidas": 0,
                                 "tardanzas": 0, "confirmado": False})
                legajos_cierre.add(_leg)

    # ── Horas: recargar CSVs de semanas ───────────────────────────────────
    periodo_json_path = PERIODOS_DIR / (p.get("archivo") or "NADA")
    semanas_numeros = []
    if periodo_json_path.exists():
        try:
            pdata = json.loads(periodo_json_path.read_text(encoding="utf-8"))
            semanas_numeros = pdata.get("semanas", [])
        except Exception:
            pass
    if not semanas_numeros:
        semanas_numeros = list(range(p.get("semana_desde", 1), p.get("semana_hasta", 1) + 1))

    feriados = list(_cargar_feriados_config())
    todos_horas = {}
    for n_sem in semanas_numeros:
        df_sem = _cargar_semana_csv(n_sem)
        if df_sem is None:
            continue
        try:
            emps_proc, _, _ = _procesar_empleados(_normalizar_columnas(df_sem))
        except Exception:
            continue
        for emp in emps_proc:
            if str(emp["legajo"]) not in legajos_cierre:
                continue
            key = str(emp["legajo"])
            if key not in todos_horas:
                todos_horas[key] = {
                    "legajo":      emp["legajo"],
                    "nombre":      emp["nombre"],
                    "departamento": depto_visible,
                    "registros":   [],
                    "excluido_ot": emp.get("excluido_ot", False),
                }
            todos_horas[key]["registros"].extend(emp.get("registros", []))

    tiene_horas = bool(todos_horas)

    # ── Construir PDF ─────────────────────────────────────────────────────
    pdf = PDFGeneral()
    pdf.titulo   = f"Informe de Cierre #{pid} — {depto_visible}"
    pdf.feriados = set(feriados)
    fam = "DejaVu" if pdf._unicode else "Helvetica"

    def titulo_seccion(texto):
        pdf.set_fill_color(23, 32, 51)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(fam, "B", 10)
        pdf.cell(0, 8, f"  {texto}", ln=1, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

    def cabecera_tabla(cols, anchos):
        pdf.set_fill_color(200, 215, 240)
        pdf.set_font(fam, "B", 7)
        for col, ancho in zip(cols, anchos):
            pdf.cell(ancho, 6, col, 1, 0, "C", fill=True)
        pdf.ln()
        pdf.set_font(fam, "", 7)

    def check_pag(alto=6):
        if pdf.get_y() + alto > pdf.h - pdf.b_margin:
            pdf.add_page()
            return True
        return False

    # Portada
    pdf.portada_abreviaciones(f"{depto_visible} — {fd} al {fh}")

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN 1: Detalle de horas día a día
    # ══════════════════════════════════════════════════════════════════════
    if tiene_horas:
        pdf.add_page()
        titulo_seccion(f"DETALLE DE HORAS POR EMPLEADO — {depto_visible}")
        pdf.set_font(fam, "I", 8)
        pdf.cell(0, 5, f"Período: {fd} al {fh}", ln=1)
        pdf.ln(3)

        emps_sorted = sorted(
            (e for e in todos_horas.values() if str(e["legajo"]) not in LEGAJOS_EXCLUIR_INFORMES),
            key=lambda e: int(str(e["legajo"])) if str(e["legajo"]).isdigit() else 0
        )
        totales_depto = {"norm": timedelta(0), "ot50": timedelta(0), "ot100": timedelta(0),
                          "com": 0, "fr": 0, "tarde": 0}
        for i, emp in enumerate(emps_sorted):
            if i > 0:
                if pdf.h - pdf.get_y() - pdf.b_margin < 50:
                    pdf.add_page()
                else:
                    pdf.ln(5)
                    pdf.set_draw_color(110, 110, 110)
                    pdf.set_line_width(0.5)
                    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
                    pdf.set_draw_color(0, 0, 0)
                    pdf.set_line_width(0.2)
                    pdf.ln(5)
            pdf.encabezado_empleado(emp["legajo"], emp["nombre"], emp.get("departamento", ""))
            totales_emp = pdf.tabla_registros(emp.get("registros", []), excluido_ot=emp.get("excluido_ot", False))
            for k in totales_depto:
                totales_depto[k] += totales_emp[k]

        if emps_sorted:
            _total_departamento(pdf, totales_depto)

        # ══════════════════════════════════════════════════════════════════
        # SECCIÓN 2: Resumen de totales
        # ══════════════════════════════════════════════════════════════════
        pdf.add_page()
        titulo_seccion(f"RESUMEN DE TOTALES DEL PERÍODO — {depto_visible}")
        pdf.set_font(fam, "I", 8)
        pdf.cell(0, 5, f"Período: {fd} al {fh}", ln=1)
        pdf.ln(3)

        COLS_R = ["Leg.", "Nombre", "OT 50%", "OT 100%", "Comidas", "Francos", "Tard.", "Confirmado"]
        ANCH_R = [12,     52,       22,        22,         16,        16,         16,     20]

        cabecera_tabla(COLS_R, ANCH_R)

        emps_db_incl = [e for e in emps_db if str(e["legajo"]) not in LEGAJOS_EXCLUIR_INFORMES]
        tot_ot50 = tot_ot100 = timedelta(0)
        tot_com = tot_fr = tot_tard = tot_conf = 0
        for e in sorted(emps_db_incl, key=lambda x: int(x["legajo"]) if str(x["legajo"]).isdigit() else 0):
            check_pag()
            ot50  = _parse_hm(e.get("ot50")  or "0h")
            ot100 = _parse_hm(e.get("ot100") or "0h")
            com   = e.get("comidas", 0) or 0
            fr    = e.get("francos", 0) or 0
            tard  = e.get("tardanzas", 0) or 0
            conf  = bool(e.get("confirmado"))
            tot_ot50  += ot50; tot_ot100 += ot100
            tot_com   += com;  tot_fr    += fr; tot_tard += tard
            if conf: tot_conf += 1
            pdf.cell(ANCH_R[0], 6, str(e["legajo"]), 1, 0, "C")
            pdf.cell(ANCH_R[1], 6, e.get("nombre", ""), 1, 0, "L")
            pdf.cell(ANCH_R[2], 6, formato_horas(_round_to_hour(ot50)),  1, 0, "C")
            pdf.cell(ANCH_R[3], 6, formato_horas(_round_to_hour(ot100)), 1, 0, "C")
            pdf.cell(ANCH_R[4], 6, str(com),  1, 0, "C")
            pdf.cell(ANCH_R[5], 6, str(fr),   1, 0, "C")
            pdf.cell(ANCH_R[6], 6, str(tard), 1, 0, "C")
            pdf.cell(ANCH_R[7], 6, "✓ Sí" if conf else "Pendiente", 1, 1, "C")

        # Fila de totales
        pdf.set_fill_color(23, 32, 51); pdf.set_text_color(255, 255, 255)
        pdf.set_font(fam, "B", 7)
        pdf.cell(ANCH_R[0] + ANCH_R[1], 6, "TOTALES", 1, 0, "R", fill=True)
        pdf.cell(ANCH_R[2], 6, formato_horas(_round_to_hour(tot_ot50)),  1, 0, "C", fill=True)
        pdf.cell(ANCH_R[3], 6, formato_horas(_round_to_hour(tot_ot100)), 1, 0, "C", fill=True)
        pdf.cell(ANCH_R[4], 6, str(tot_com),  1, 0, "C", fill=True)
        pdf.cell(ANCH_R[5], 6, str(tot_fr),   1, 0, "C", fill=True)
        pdf.cell(ANCH_R[6], 6, str(tot_tard), 1, 0, "C", fill=True)
        pdf.cell(ANCH_R[7], 6, f"{tot_conf}/{len(emps_db_incl)}", 1, 1, "C", fill=True)
        pdf.set_text_color(0, 0, 0)

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN 3: Saldo de francos al cierre
    # ══════════════════════════════════════════════════════════════════════
    pdf.add_page()
    titulo_seccion(f"SALDO DE FRANCOS AL CIERRE — {depto_visible}")
    pdf.set_font(fam, "I", 8)
    pdf.cell(0, 5, f"Período: {fd} al {fh}", ln=1)
    pdf.ln(3)

    COLS_S = ["Leg.", "Nombre", "Saldo anterior", "Generados período", "Tomados", "Saldo final"]
    ANCH_S = [12,     52,       28,                28,                  20,         24]

    cabecera_tabla(COLS_S, ANCH_S)

    tomados_por_leg = {}
    for ft in ft_db:
        leg = str(ft["legajo"])
        tomados_por_leg[leg] = tomados_por_leg.get(leg, 0) + (ft.get("dias") or 0)

    tot_sal_ant = tot_gen = tot_tom = tot_sal_fin = 0
    for e in sorted(emps_db, key=lambda x: int(x["legajo"]) if str(x["legajo"]).isdigit() else 0):
        check_pag()
        leg = str(e["legajo"])
        ant_data = saldo_ant.get(leg, {})
        si_ant  = ant_data.get("saldo", 0) if isinstance(ant_data, dict) else 0
        gen_per = e.get("francos", 0) or 0
        tom     = tomados_por_leg.get(leg, 0)
        sf      = si_ant + gen_per - tom
        tot_sal_ant += si_ant; tot_gen += gen_per
        tot_tom     += tom;    tot_sal_fin += sf

        if sf > 0:   pdf.set_fill_color(220, 252, 231)
        elif sf < 0: pdf.set_fill_color(254, 202, 202)
        else:        pdf.set_fill_color(241, 245, 249)
        fill = True

        pdf.cell(ANCH_S[0], 6, leg,               1, 0, "C", fill)
        pdf.cell(ANCH_S[1], 6, e.get("nombre",""),1, 0, "L", fill)
        pdf.cell(ANCH_S[2], 6, str(si_ant),        1, 0, "C", fill)
        pdf.cell(ANCH_S[3], 6, str(gen_per),       1, 0, "C", fill)
        pdf.cell(ANCH_S[4], 6, str(tom),           1, 0, "C", fill)
        pdf.cell(ANCH_S[5], 6, str(sf),            1, 1, "C", fill)
        pdf.set_fill_color(255, 255, 255)

    pdf.set_fill_color(23, 32, 51); pdf.set_text_color(255, 255, 255)
    pdf.set_font(fam, "B", 7)
    pdf.cell(ANCH_S[0] + ANCH_S[1], 6, "TOTALES", 1, 0, "R", fill=True)
    pdf.cell(ANCH_S[2], 6, str(tot_sal_ant), 1, 0, "C", fill=True)
    pdf.cell(ANCH_S[3], 6, str(tot_gen),     1, 0, "C", fill=True)
    pdf.cell(ANCH_S[4], 6, str(tot_tom),     1, 0, "C", fill=True)
    pdf.cell(ANCH_S[5], 6, str(tot_sal_fin), 1, 1, "C", fill=True)
    pdf.set_text_color(0, 0, 0)

    # Notas de ajuste manual al pie de la sección saldo
    leg_emps_s = [str(e["legajo"]) for e in emps_db]
    if leg_emps_s:
        ph_ns = ",".join("?" * len(leg_emps_s))
        with _get_db() as conn:
            notas_ajs = [(str(r["leg"]),
                          next((e.get("nombre","") for e in emps_db if str(e["legajo"]) == str(r["leg"])), str(r["leg"])),
                          r["nota"])
                         for r in conn.execute(
                             f"SELECT CAST(legajo AS TEXT) as leg, nota FROM francos_saldo_inicial "
                             f"WHERE CAST(legajo AS TEXT) IN ({ph_ns}) AND nota LIKE 'Ajuste manual%'",
                             leg_emps_s)]
        if notas_ajs:
            pdf.ln(5)
            pdf.set_font(fam, "B", 7); pdf.set_text_color(90, 90, 90)
            pdf.cell(0, 5, "Ajustes manuales de saldo inicial:", ln=1)
            pdf.set_font(fam, "", 7)
            for leg_n, nom_n, nota_n in notas_ajs:
                pdf.cell(0, 4, f"  {leg_n} {nom_n}: {nota_n}", ln=1)
            pdf.set_text_color(0, 0, 0)

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN 4: Detalle de francos tomados
    # ══════════════════════════════════════════════════════════════════════
    if ft_db:
        pdf.add_page()
        titulo_seccion(f"DETALLE DE FRANCOS TOMADOS — {depto_visible}")
        pdf.set_font(fam, "I", 8)
        pdf.cell(0, 5, f"Período: {fd} al {fh}", ln=1)
        pdf.ln(3)

        COLS_F = ["Leg.", "Nombre", "Tipo", "Fechas", "Días", "Estado", "Observaciones"]
        ANCH_F = [12,     52,       14,      60,        10,     20,       22]

        cabecera_tabla(COLS_F, ANCH_F)

        for ft in ft_db:
            tipo = ft.get("tipo", "")
            if tipo == "UNICO":
                fechas = ft.get("fecha_desde", "")
            elif tipo == "RANGO":
                fechas = f"{ft.get('fecha_desde','')} → {ft.get('fecha_hasta','')}"
            else:
                try:
                    fl = json.loads(ft.get("fechas_sueltas") or "[]")
                    fechas = ", ".join(fl[:3]) + (f" +{len(fl)-3}" if len(fl) > 3 else "")
                except Exception:
                    fechas = ft.get("fecha_desde", "")

            check_pag()
            pdf.cell(ANCH_F[0], 6, str(ft.get("legajo","")),       1, 0, "C")
            pdf.cell(ANCH_F[1], 6, ft.get("nombre",""),             1, 0, "L")
            pdf.cell(ANCH_F[2], 6, tipo,                            1, 0, "C")
            pdf.cell(ANCH_F[3], 6, fechas,                          1, 0, "L")
            pdf.cell(ANCH_F[4], 6, str(ft.get("dias", 0)),          1, 0, "C")
            pdf.cell(ANCH_F[5], 6, ft.get("estado",""),             1, 0, "C")
            pdf.cell(ANCH_F[6], 6, ft.get("observaciones","") or "", 1, 1, "L")

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN 5: Movimientos compensatorios (devoluciones de francos cerrados
    # anulados antes de este cierre) — no modifica cierres anteriores, solo
    # constancia de que la devolución figura en este cierre.
    # ══════════════════════════════════════════════════════════════════════
    with _get_db() as conn:
        devoluciones = [dict(r) for r in conn.execute(
            "SELECT * FROM francos_anulaciones_cerrados WHERE periodo_aplicado_id=? "
            "ORDER BY CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()]
    if devoluciones:
        pdf.ln(5)
        titulo_seccion("MOVIMIENTOS COMPENSATORIOS")
        pdf.set_font(fam, "", 8)
        total_dev = 0
        for d in devoluciones:
            check_pag()
            dias_dev = d.get("dias", 0) or 0
            total_dev += dias_dev
            tipo = d.get("tipo", "")
            if tipo == "RANGO":
                fechas = f"{d.get('fecha_desde','')} → {d.get('fecha_hasta','')}"
            elif tipo == "SUELTAS":
                try:
                    fechas = ", ".join(json.loads(d.get("fechas_sueltas") or "[]"))
                except Exception:
                    fechas = ""
            else:
                fechas = d.get("fecha_desde", "")
            pdf.set_text_color(0, 120, 0)
            pdf.multi_cell(0, 5,
                f"  {d.get('legajo','')} - {d.get('nombre','')}: Devolución por anulación de "
                f"franco cerrado ({fechas}) +{dias_dev} — Motivo: {d.get('motivo','')}")
            pdf.set_text_color(0, 0, 0)
        pdf.set_font(fam, "B", 8)
        pdf.set_text_color(0, 120, 0)
        pdf.cell(0, 6, f"  Total devolución por anulación de francos cerrados: +{total_dev}", ln=1)
        pdf.set_text_color(0, 0, 0)

    raw = pdf.output(dest="S")
    return raw.encode("latin-1") if isinstance(raw, str) else bytes(raw)


@app.route("/periodos/informe_mensual")
def periodos_informe_mensual():
    """PDF combinado de todos los cierres ACTIVOS de un mes (un depto por página)."""
    if not _autenticado(): return _requiere_auth()
    mes = request.args.get("mes", "").strip()  # "YYYY-MM"
    if not mes or len(mes) != 7 or mes[4] != "-":
        return "Parámetro 'mes' requerido con formato YYYY-MM.", 400

    with _get_db() as conn:
        rows = conn.execute(
            "SELECT id, fecha_desde, fecha_hasta FROM periodos "
            "WHERE COALESCE(estado,'ACTIVO')='ACTIVO' AND fecha_desde LIKE ? ORDER BY id",
            (f"{mes}%",)
        ).fetchall()

    if not rows:
        return f"No hay cierres activos para {mes}.", 404

    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        return "Librería pypdf no instalada en el servidor.", 500

    writer = PdfWriter()
    for r in rows:
        pdf_bytes = _generar_pdf_cierre_completo(r["id"])
        if pdf_bytes:
            for page in PdfReader(BytesIO(pdf_bytes)).pages:
                writer.add_page(page)

    if len(writer.pages) == 0:
        return "No se pudo generar el informe.", 500

    out = BytesIO()
    writer.write(out)
    out.seek(0)
    nombre = f"informe_mensual_{mes.replace('-', '_')}.pdf"
    return send_file(out, mimetype="application/pdf",
                     as_attachment=False, download_name=nombre)


@app.route("/periodos/<int:pid>/informe_completo")
def periodos_informe_completo(pid):
    """PDF completo del cierre: horas + saldo francos + detalle tomados
    + detalle de confirmaciones/descripciones por empleado (anexado al final)."""
    if not _autenticado(): return _requiere_auth()
    try:
        pdf_data = _generar_pdf_cierre_completo(pid)
    except Exception as e:
        return f"Error generando informe: {e}", 500
    if pdf_data is None:
        return "Cierre no encontrado.", 404
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        rows = conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=? ORDER BY departamento, CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()

    # Anexar el detalle de confirmaciones/descripciones por empleado (con o sin
    # confirmar) al final del informe. Si algo falla, se devuelve igualmente
    # el informe base sin esta sección extra.
    if p is not None:
        try:
            periodo = dict(p)
            empleados = []
            for e in rows:
                d = dict(e)
                d["semanas"] = json.loads(d["semanas"] or "[]")
                d["confirmado"] = bool(d["confirmado"])
                empleados.append(d)
            datos_csv = _datos_semanas_periodo(periodo)
            conf_pdf_data = _generar_pdf_confirmaciones_cierre(periodo, empleados, datos_csv)

            from pypdf import PdfWriter, PdfReader
            writer = PdfWriter()
            for page in PdfReader(BytesIO(pdf_data)).pages:
                writer.add_page(page)
            for page in PdfReader(BytesIO(conf_pdf_data)).pages:
                writer.add_page(page)
            out = BytesIO()
            writer.write(out)
            pdf_data = out.getvalue()
        except Exception:
            pass

    fd = (p["fecha_desde"] if p else "").replace("-","")
    fh = (p["fecha_hasta"] if p else "").replace("-","")
    nombre = f"informe_cierre_{pid}_{fd}_{fh}.pdf"
    return send_file(BytesIO(pdf_data), mimetype="application/pdf",
                     as_attachment=False, download_name=nombre)


@app.route("/periodos/francos_pdf/<int:pid>")
def periodos_francos_pdf(pid):
    """Muestra (inline, en el navegador) el detalle de francos tomados
    snapshoteado al momento del cierre #pid (francos_cierre_detalle) — nunca
    recalcula desde francos_tomados vivos. El encabezado muestra la fecha
    real del cierre (periodos.cerrado_en) como "Cargados hasta el", y la
    fecha de hoy solo como "Reimpreso el"."""
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return "Cierre no encontrado.", 404
        francos = conn.execute("""
            SELECT legajo, nombre, departamento, tipo, fecha_desde, fecha_hasta,
                   fechas_sueltas, dias, estado, fecha_emision, autorizado_por, observaciones
            FROM francos_cierre_detalle
            WHERE periodo_id=?
            ORDER BY CAST(legajo AS INTEGER), fecha_desde
        """, (pid,)).fetchall()
        devoluciones = conn.execute(
            "SELECT * FROM francos_anulaciones_cerrados WHERE periodo_aplicado_id=? ORDER BY CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()
    fecha_corte = _normalizar_cargado_en(p["cerrado_en"] or "")
    hoy = datetime.now().strftime("%Y-%m-%d")
    francos_list = [dict(r) for r in francos]
    devoluciones_list = [dict(d) for d in devoluciones]
    pdf_bytes = _generar_pdf_francos_cierre(pid, francos_list, fecha_corte,
                                             reimpreso_el=hoy, escribir_archivo=False,
                                             devoluciones=devoluciones_list)
    if pdf_bytes is None:
        return "Error generando PDF.", 500
    fd = (p["fecha_desde"] or "").replace("-", "")
    fh = (p["fecha_hasta"] or "").replace("-", "")
    nombre = f"francos_cierre_{pid}_{fd}_{fh}.pdf"
    return send_file(BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=False, download_name=nombre)


@app.route("/periodos/confirmaciones_pdf/<int:pid>")
def periodos_confirmaciones_pdf(pid):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return "Periodo no encontrado", 404
        rows = conn.execute(
            "SELECT * FROM periodo_empleados WHERE periodo_id=? ORDER BY departamento, CAST(legajo AS INTEGER)",
            (pid,)
        ).fetchall()

    periodo = dict(p)
    empleados = []
    for e in rows:
        d = dict(e)
        d["semanas"] = json.loads(d["semanas"] or "[]")
        d["confirmado"] = bool(d["confirmado"])
        empleados.append(d)
    # semanas en periodo_empleados ya son números visibles — no remapear
    datos_csv = _datos_semanas_periodo(periodo)

    try:
        pdf_data = _generar_pdf_confirmaciones_cierre(periodo, empleados, datos_csv)
        nombre = f"confirmaciones_cierre_{pid}.pdf"
        return send_file(
            BytesIO(pdf_data),
            mimetype="application/pdf",
            as_attachment=False,
            download_name=nombre,
        )
    except Exception as exc:
        confirmaciones = _leer_confirmaciones_cierre(periodo)
        return render_template(
            "confirmaciones_cierre.html",
            periodo=periodo,
            empleados=empleados,
            confirmaciones=confirmaciones,
            pendientes=_pendientes_cierre(confirmaciones, empleados),
            error_pdf=str(exc),
        )


@app.route("/admin/recalcular-saldos", methods=["POST"])
def admin_recalcular_saldos():
    """Recalcula y actualiza francos_saldo_inicial para un departamento o todos."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    depto_param = request.form.get("departamento", "").strip()
    ahora_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    saldos = _calcular_saldos()
    empleados = _empleados_conocidos()
    emp_map = {str(e["legajo"]): e for e in empleados}
    actualizados = 0
    with _get_db() as conn:
        tomados_totales = {r["legajo"]: (r["total"] or 0) for r in conn.execute(
            "SELECT legajo, SUM(dias) as total FROM francos_tomados WHERE COALESCE(estado,'') != 'Anulado' GROUP BY legajo"
        )}
        gen_extra_totales = {}
        for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_generados GROUP BY legajo"):
            gen_extra_totales[r["legajo"]] = gen_extra_totales.get(r["legajo"], 0) + (r["total"] or 0)
        for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_semana_manual GROUP BY legajo"):
            gen_extra_totales[r["legajo"]] = gen_extra_totales.get(r["legajo"], 0) + (r["total"] or 0)
        # Buscar fecha_corte del cierre manual más reciente por depto
        fecha_corte_por_leg = {}
        for cf in conn.execute("SELECT departamento, fecha_hasta FROM cierres_francos ORDER BY id DESC").fetchall():
            dep_cf = _normalizar_departamento_web(cf["departamento"])
            for emp in empleados:
                if _normalizar_departamento_web(emp["departamento"]) == dep_cf:
                    leg = str(emp["legajo"])
                    if leg not in fecha_corte_por_leg:
                        fecha_corte_por_leg[leg] = cf["fecha_hasta"]
        for s in saldos:
            leg = str(s["legajo"])
            emp = emp_map.get(leg)
            if not emp:
                continue
            if depto_param and _normalizar_departamento_web(emp["departamento"]) != _normalizar_departamento_web(depto_param):
                continue
            nombre = emp["nombre"]
            fecha_corte = fecha_corte_por_leg.get(leg, ahora_str[:10])
            conn.execute("""
                INSERT INTO francos_saldo_inicial
                    (legajo, nombre, saldo, nota, cargado_en, tomados_al_corte, gen_extra_al_corte, fecha_corte)
                VALUES (?,?,?,?,?,?,?,?)
                ON CONFLICT(legajo) DO UPDATE SET
                    saldo=excluded.saldo, nota=excluded.nota, cargado_en=excluded.cargado_en,
                    tomados_al_corte=excluded.tomados_al_corte,
                    gen_extra_al_corte=excluded.gen_extra_al_corte,
                    fecha_corte=excluded.fecha_corte
            """, (leg, nombre, s["saldo_actual"],
                  f"Recalculado manualmente {ahora_str[:10]}",
                  ahora_str, tomados_totales.get(leg, 0),
                  gen_extra_totales.get(leg, 0), fecha_corte))
            actualizados += 1
        conn.commit()
    return jsonify({"ok": True, "actualizados": actualizados})


@app.route("/admin/corregir-francos-cierre/<int:pid>", methods=["GET", "POST"])
def admin_corregir_francos_cierre(pid):
    """Corrige retroactivamente el snapshot de francos de un cierre ya hecho.

    Criterio: un franco entra al cierre #pid si su fecha/hora de carga
    (cargado_en) cae dentro de la ventana

        fecha_corte_anterior < cargado_en <= fecha_corte

    donde fecha_corte es la fecha/hora real de este cierre
    (periodos.cerrado_en, sin truncar) y fecha_corte_anterior es la del
    cierre activo (no anulado) anterior del mismo departamento (sin límite
    inferior si no existe uno) — sin importar la fecha tomada. La ventana se
    calcula siempre respecto del cierre #pid, no del último cierre global,
    por lo que sirve también para corregir cierres históricos. La selección
    usa _seleccionar_francos_ventana_cierre, la misma fuente de verdad que
    _snapshot_francos_cierre.

    Los francos sin cargado_en (vacío/NULL/solo espacios — "legacy") nunca
    se asignan a esta ventana: se reportan aparte en
    "legacy_sin_cargado_en" para revisión/tratamiento manual.

    ?confirmar=si aplica la corrección (backup previo de la DB, agrega a
    francos_cierre_detalle, marca francos_tomados.estado='Cerrado', ajusta
    francos_saldo_inicial solo donde corresponda, y deja constancia en
    periodos.francos_corregido_en). Sin ese parámetro es solo diagnóstico
    de lectura (faltantes + legacy + saldos), sin importar el método. No
    borra ni modifica registros existentes de francos_cierre_detalle,
    periodo_empleados, confirmaciones ni semanas, y no toca los francos
    "legacy".

    ?excluir_legajos=13,14 (opcional): esos legajos quedan afuera del ajuste
    de saldo/tomados_al_corte (útil cuando "correcto" para ese legajo no
    refleja bien su historial acumulado de cierres anteriores — el cálculo
    de la ventana de este cierre puntual puede no coincidir con el
    acumulado real si el legajo ya tenía tomados contados en un cierre
    previo). Los francos que le faltan a ese legajo en francos_cierre_detalle
    igual se agregan normalmente; solo se saltea el ajuste de saldo_inicial."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    excluir_legajos = {
        s.strip() for s in request.args.get("excluir_legajos", "").split(",") if s.strip()
    }

    with _get_db() as conn:
        periodo = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not periodo:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        fecha_corte = _normalizar_cargado_en(periodo["cerrado_en"] or "")
        if not fecha_corte:
            return jsonify({"error": f"Período #{pid} no tiene fecha de cierre registrada"}), 400

        emp_rows = conn.execute(
            "SELECT DISTINCT legajo, nombre, departamento FROM periodo_empleados WHERE periodo_id=?", (pid,)
        ).fetchall()
        legajos = [str(r["legajo"]) for r in emp_rows]
        if not legajos:
            return jsonify({"error": f"Período #{pid} no tiene empleados asociados"}), 400
        departamento = _nombre_departamento_visible(emp_rows[0]["departamento"] or "")
        nombre_por_leg = {str(r["legajo"]): r["nombre"] for r in emp_rows}

        actuales = conn.execute(
            "SELECT * FROM francos_cierre_detalle WHERE periodo_id=?", (pid,)
        ).fetchall()

        correctos, fecha_corte_anterior, legacy = _seleccionar_francos_ventana_cierre(
            conn, pid, legajos, departamento, fecha_corte)

        def _clave(r):
            return (str(r["legajo"]), r["tipo"] or "", r["fecha_desde"] or "",
                    r["fecha_hasta"] or "", r["fechas_sueltas"] or "[]", r["dias"] or 0)

        claves_actuales = {_clave(r) for r in actuales}
        faltantes = [r for r in correctos if _clave(r) not in claves_actuales]

        # tomados_al_corte es un puntero ACUMULADO (todo lo tomado, no anulado,
        # hasta fecha_corte) -- no la ventana incremental de este cierre
        # puntual. Un legajo con historial de un cierre activo anterior ya
        # tiene ese acumulado correctamente contado; usar solo la ventana de
        # ESTE cierre (como antes) lo resetearía por debajo del valor real y
        # generaría un doble descuento futuro. Ver CLAUDE.md / plan de
        # rediseño de cierre de francos.
        ph_legajos = ",".join("?" * len(legajos))
        tomados_acumulado_correcto = {r["legajo"]: (r["total"] or 0) for r in conn.execute(f"""
            SELECT legajo, SUM(dias) as total FROM francos_tomados
            WHERE COALESCE(estado,'') != 'Anulado'
              AND {_SQL_CARGADO_EN_NORM} != ''
              AND {_SQL_CARGADO_EN_NORM} <= ?
              AND legajo IN ({ph_legajos})
            GROUP BY legajo
        """, (fecha_corte, *legajos))}

        saldos_iniciales = {r["legajo"]: dict(r) for r in conn.execute("SELECT * FROM francos_saldo_inicial")}

        saldos_diff = []
        for leg in legajos:
            si = saldos_iniciales.get(leg, {})
            tomados_actual   = si.get("tomados_al_corte") or 0
            tomados_correcto = tomados_acumulado_correcto.get(leg, 0)
            diferencia = tomados_correcto - tomados_actual
            saldo_actual = si.get("saldo") or 0
            saldos_diff.append({
                "legajo": leg,
                "nombre": si.get("nombre") or nombre_por_leg.get(leg, ""),
                "tomados_al_corte_actual": tomados_actual,
                "tomados_al_corte_correcto": tomados_correcto,
                "diferencia_dias": diferencia,
                "saldo_actual": saldo_actual,
                "saldo_corregido": saldo_actual - diferencia,
                "excluido_del_ajuste_saldo": leg in excluir_legajos,
            })

        diagnostico = {
            "periodo_id": pid,
            "departamento": departamento,
            "fecha_corte": fecha_corte,
            "fecha_corte_anterior": fecha_corte_anterior,
            "ya_corregido_en": periodo["francos_corregido_en"] or "",
            "registros_actuales": len(actuales),
            "registros_correctos": len(correctos),
            "registros_faltantes": len(faltantes),
            "dias_agregados_total": sum((r["dias"] or 0) for r in faltantes),
            "faltantes": [{
                "legajo": r["legajo"], "nombre": r["nombre"], "tipo": r["tipo"],
                "fecha_desde": r["fecha_desde"], "fecha_hasta": r["fecha_hasta"],
                "fechas_sueltas": r["fechas_sueltas"], "dias": r["dias"],
                "estado": r["estado"], "cargado_en": r["cargado_en"],
            } for r in faltantes],
            "legacy_sin_cargado_en": [{
                "legajo": r["legajo"], "nombre": r["nombre"], "tipo": r["tipo"],
                "fecha_desde": r["fecha_desde"], "fecha_hasta": r["fecha_hasta"],
                "fechas_sueltas": r["fechas_sueltas"], "dias": r["dias"],
                "estado": r["estado"], "cargado_en": r["cargado_en"],
            } for r in legacy],
            "saldos": [s for s in saldos_diff if s["diferencia_dias"] != 0],
        }

    if request.args.get("confirmar") != "si":
        return jsonify(diagnostico)

    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not faltantes:
        with _get_db() as conn:
            conn.execute("UPDATE periodos SET francos_corregido_en=? WHERE id=?", (ahora_str, pid))
            conn.commit()
        return jsonify({"ok": True, "mensaje": "Nada para corregir — el snapshot ya estaba completo.",
                        "diagnostico": diagnostico})

    DATOS_DIR.mkdir(exist_ok=True)
    backup_path = DATOS_DIR / f"cierres_backup_pre_correccion_francos_{pid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    shutil.copy2(DB_FILE, backup_path)

    with _get_db() as conn:
        for r in faltantes:
            conn.execute("""
                INSERT INTO francos_cierre_detalle
                  (periodo_id, legajo, nombre, departamento, tipo, fecha_desde, fecha_hasta,
                   fechas_sueltas, dias, estado, fecha_emision, autorizado_por, observaciones,
                   francos_tomados_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (pid, r["legajo"], r["nombre"], departamento, r["tipo"],
                  r["fecha_desde"] or "", r["fecha_hasta"] or "", r["fechas_sueltas"] or "[]",
                  r["dias"], r["estado"] or "", r["fecha_emision"] or "",
                  r["autorizado_por"] or "", r["observaciones"] or "", r["id"]))
            conn.execute("""
                UPDATE francos_tomados SET estado='Cerrado'
                WHERE id=? AND COALESCE(estado,'') NOT IN ('Anulado','Cerrado')
            """, (r["id"],))

        for s in saldos_diff:
            leg = s["legajo"]
            if leg in excluir_legajos:
                continue
            si = saldos_iniciales.get(leg)
            if s["diferencia_dias"] != 0:
                conn.execute("""
                    INSERT INTO francos_saldo_inicial
                        (legajo, nombre, saldo, nota, cargado_en, tomados_al_corte, gen_extra_al_corte, fecha_corte)
                    VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(legajo) DO UPDATE SET
                        saldo=excluded.saldo, nota=excluded.nota, cargado_en=excluded.cargado_en,
                        tomados_al_corte=excluded.tomados_al_corte, fecha_corte=excluded.fecha_corte
                """, (leg, s["nombre"], s["saldo_corregido"],
                      f"Corrección snapshot francos cierre #{pid} ({fecha_corte})",
                      ahora_str, s["tomados_al_corte_correcto"],
                      (si.get("gen_extra_al_corte", 0) if si else 0) or 0, fecha_corte))
            elif si:
                conn.execute("UPDATE francos_saldo_inicial SET fecha_corte=? WHERE legajo=?",
                              (fecha_corte, leg))

        conn.execute("UPDATE periodos SET francos_corregido_en=? WHERE id=?", (ahora_str, pid))

        francos_actualizados = conn.execute(
            "SELECT * FROM francos_cierre_detalle WHERE periodo_id=? ORDER BY CAST(legajo AS INTEGER), fecha_desde",
            (pid,)
        ).fetchall()
        francos_list = [dict(r) for r in francos_actualizados]
        conn.commit()

    _generar_pdf_francos_cierre(pid, francos_list, fecha_corte)

    return jsonify({
        "ok": True,
        "backup": str(backup_path),
        "registros_agregados": len(faltantes),
        "dias_agregados_total": diagnostico["dias_agregados_total"],
        "saldos_corregidos": [s for s in saldos_diff if s["diferencia_dias"] != 0],
        "fecha_corte": fecha_corte,
    })


@app.route("/admin/restaurar-saldo-desde-periodo/<int:pid>", methods=["GET", "POST"])
def admin_restaurar_saldo_desde_periodo(pid):
    """Restaura filas puntuales de francos_saldo_inicial usando
    periodos.saldo_anterior del cierre #pid (el snapshot guardado antes de
    que ESE cierre corriera su propia actualización automática de saldo),
    para los `legajos` indicados (coma-separado).

    Sirve para corregir un legajo cuyo saldo quedó mal después de que la
    actualización automática de un cierre le sumara algo que no
    correspondía a ese período (por ejemplo un "parcial" de una semana en
    curso de OTRO mes que se coló en el cálculo). No requiere anular el
    cierre ni afecta a otros legajos.

    GET = diagnóstico de solo lectura (valores actuales vs. los del
    saldo_anterior de este cierre). ?confirmar=si aplica."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    legajos = [s.strip() for s in request.args.get("legajos", "").split(",") if s.strip()]
    if not legajos:
        return jsonify({"error": "Falta el parámetro 'legajos' (coma-separado)."}), 400

    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        try:
            saldo_ant = json.loads(p["saldo_anterior"] or "{}")
        except Exception:
            saldo_ant = {}
        saldo_ant = {k: v for k, v in saldo_ant.items() if k in legajos}

        actuales = {
            r["legajo"]: dict(r)
            for r in conn.execute(
                f"SELECT * FROM francos_saldo_inicial WHERE legajo IN ({','.join('?'*len(legajos))})",
                legajos
            ).fetchall()
        }

        comparacion = [{"legajo": leg, "en_saldo_anterior": saldo_ant.get(leg), "actual": actuales.get(leg)}
                       for leg in legajos]

        if request.args.get("confirmar") != "si":
            return jsonify({"periodo_id": pid, "comparacion": comparacion})

        restaurados = []
        ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for leg in legajos:
            vals = saldo_ant.get(leg)
            if not vals:
                continue
            actual = actuales.get(leg, {})
            conn.execute("""
                UPDATE francos_saldo_inicial
                SET saldo=?, tomados_al_corte=?, gen_extra_al_corte=?, fecha_corte=?,
                    nota=?, cargado_en=?
                WHERE legajo=?
            """, (vals.get("saldo", 0), vals.get("tomados_al_corte", 0),
                  vals.get("gen_extra_al_corte", 0), vals.get("fecha_corte", ""),
                  f"Restaurado desde saldo_anterior del cierre #{pid}",
                  ahora_str, leg))
            restaurados.append(leg)
        conn.commit()

    return jsonify({"ok": True, "restaurados": restaurados, "comparacion": comparacion})


@app.route("/admin/restaurar-saldo-desde-backup", methods=["GET", "POST"])
def admin_restaurar_saldo_desde_backup():
    """Restaura filas puntuales de francos_saldo_inicial desde un backup
    sqlite hecho por /admin/corregir-francos-cierre (parámetro `backup`,
    solo nombre de archivo dentro de DATOS_DIR, sin rutas) para los
    `legajos` indicados (coma-separado).

    Sirve para deshacer un ajuste de saldo que resultó incorrecto para
    legajos puntuales, sin tocar francos_cierre_detalle ni el resto de la
    corrección (que puede seguir siendo válida para otros legajos).

    GET = diagnóstico de solo lectura (valores actuales vs. los del
    backup). ?confirmar=si aplica: UPDATE de francos_saldo_inicial con los
    valores (saldo, tomados_al_corte, gen_extra_al_corte, fecha_corte) tal
    como estaban en el backup, solo para esos legajos."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    backup_nombre = request.args.get("backup", "").strip()
    legajos = [s.strip() for s in request.args.get("legajos", "").split(",") if s.strip()]
    if not backup_nombre or "/" in backup_nombre or "\\" in backup_nombre:
        return jsonify({"error": "Parámetro 'backup' inválido o faltante (solo nombre de archivo)."}), 400
    if not legajos:
        return jsonify({"error": "Falta el parámetro 'legajos' (coma-separado)."}), 400

    backup_path = DATOS_DIR / backup_nombre
    if not backup_path.exists():
        return jsonify({"error": f"No existe el backup {backup_nombre} en {DATOS_DIR}."}), 404

    backup_conn = sqlite3.connect(str(backup_path))
    backup_conn.row_factory = sqlite3.Row
    ph = ",".join("?" * len(legajos))
    filas_backup = {
        r["legajo"]: dict(r)
        for r in backup_conn.execute(
            f"SELECT * FROM francos_saldo_inicial WHERE legajo IN ({ph})", legajos
        ).fetchall()
    }
    backup_conn.close()

    with _get_db() as conn:
        actuales = {
            r["legajo"]: dict(r)
            for r in conn.execute(
                f"SELECT * FROM francos_saldo_inicial WHERE legajo IN ({ph})", legajos
            ).fetchall()
        }

        comparacion = []
        for leg in legajos:
            b = filas_backup.get(leg)
            a = actuales.get(leg)
            comparacion.append({
                "legajo": leg,
                "en_backup": b,
                "actual": a,
            })

        if request.args.get("confirmar") != "si":
            return jsonify({"backup": backup_nombre, "comparacion": comparacion})

        restaurados = []
        for leg in legajos:
            b = filas_backup.get(leg)
            if not b:
                continue
            conn.execute("""
                UPDATE francos_saldo_inicial
                SET saldo=?, tomados_al_corte=?, gen_extra_al_corte=?, fecha_corte=?,
                    nota=?, cargado_en=?
                WHERE legajo=?
            """, (b["saldo"], b["tomados_al_corte"], b["gen_extra_al_corte"], b["fecha_corte"],
                  f"Restaurado desde backup {backup_nombre}",
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leg))
            restaurados.append(leg)
        conn.commit()

    return jsonify({"ok": True, "restaurados": restaurados, "comparacion": comparacion})


@app.route("/admin/revertir-francos-cierre-anulado/<int:pid>", methods=["GET", "POST"])
def admin_revertir_francos_cierre_anulado(pid):
    """Aplica retroactivamente lo que ahora hace periodo_anular automáticamente:
    revierte a su estado original los francos_tomados que el cierre #pid dejó
    en estado='Cerrado' antes de que existiera el fix. Solo funciona sobre
    períodos con estado ANULADO (si no, es más seguro anularlo primero y que
    la reversión ocurra sola). Ver _revertir_estado_francos_cierre.

    GET = diagnóstico de solo lectura: lista los francos de
    francos_cierre_detalle del cierre #pid junto al estado actual del
    francos_tomados correspondiente. POST con ?confirmar=si = aplica la
    reversión (solo toca filas todavía en 'Cerrado')."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        if (p["estado"] or "ACTIVO") != "ANULADO":
            return jsonify({"error": f"Período #{pid} no está anulado — esta acción es solo para períodos anulados."}), 400

        detalle = conn.execute(
            "SELECT * FROM francos_cierre_detalle WHERE periodo_id=?", (pid,)
        ).fetchall()
        candidatos = []
        for d in detalle:
            row = conn.execute("""
                SELECT id, estado FROM francos_tomados
                WHERE legajo=? AND tipo=? AND fecha_desde=? AND fecha_hasta=?
                  AND fechas_sueltas=? AND dias=?
            """, (d["legajo"], d["tipo"] or "", d["fecha_desde"] or "",
                  d["fecha_hasta"] or "", d["fechas_sueltas"] or "[]", d["dias"] or 0)).fetchone()
            candidatos.append({
                "legajo": d["legajo"], "nombre": d["nombre"],
                "fecha_desde": d["fecha_desde"], "fecha_hasta": d["fecha_hasta"],
                "dias": d["dias"],
                "estado_en_snapshot": d["estado"],
                "estado_actual_francos_tomados": row["estado"] if row else "(no encontrado)",
            })

        if request.args.get("confirmar") != "si":
            return jsonify({"periodo_id": pid, "candidatos": candidatos})

        revertidos = _revertir_estado_francos_cierre(conn, pid)
        conn.commit()

    return jsonify({"ok": True, "revertidos": revertidos, "candidatos": candidatos})


@app.route("/admin/corregir-fecha-corte/<int:pid>", methods=["GET", "POST"])
def admin_corregir_fecha_corte(pid):
    """Corrige el dato histórico de francos_saldo_inicial.fecha_corte para
    los legajos del cierre #pid.

    Regla definitiva: para francos, fecha_corte debe ser la fecha/hora REAL
    del cierre (periodos.cerrado_en), nunca fecha_hasta del período de horas
    extras que se cerró. La actualización automática de saldo_inicial al
    cerrar un período ya usa cerrado_en desde la commit c34f12e
    (2026-06-11); cierres hechos con código anterior (p.ej. pid=2) quedaron
    con fecha_corte = fecha_hasta del período de horas.

    GET = diagnóstico de solo lectura: para cada legajo del cierre #pid,
    compara el fecha_corte actual en francos_saldo_inicial contra
    cerrado_en y reporta los cambios que correspondería aplicar. No escribe
    nada.

    POST con ?confirmar=si aplica la corrección: backup previo de la DB y
    UPDATE de una sola columna (fecha_corte) en francos_saldo_inicial,
    acotado a los legajos de este cierre. No toca saldo, tomados_al_corte,
    gen_extra_al_corte, nota ni cargado_en, y no afecta legajos de otros
    cierres."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    with _get_db() as conn:
        periodo = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not periodo:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        cerrado_en = _normalizar_cargado_en(periodo["cerrado_en"] or "")
        if not cerrado_en:
            return jsonify({"error": f"Período #{pid} no tiene fecha de cierre registrada"}), 400

        legajos = [str(r["legajo"]) for r in conn.execute(
            "SELECT DISTINCT legajo FROM periodo_empleados WHERE periodo_id=?", (pid,)
        )]
        if not legajos:
            return jsonify({"error": f"Período #{pid} no tiene empleados asociados"}), 400

        placeholders = ",".join("?" * len(legajos))
        actuales = {r["legajo"]: (r["fecha_corte"] or "") for r in conn.execute(
            f"SELECT legajo, fecha_corte FROM francos_saldo_inicial WHERE legajo IN ({placeholders})",
            legajos
        )}

    cambios = [
        {"legajo": leg, "fecha_corte_actual": actuales[leg], "fecha_corte_nuevo": cerrado_en}
        for leg in legajos
        if leg in actuales and actuales[leg] != cerrado_en
    ]
    sin_registro = [leg for leg in legajos if leg not in actuales]

    diagnostico = {
        "periodo_id": pid,
        "fecha_hasta_periodo": periodo["fecha_hasta"] or "",
        "cerrado_en": cerrado_en,
        "legajos": legajos,
        "cambios": cambios,
        "legajos_sin_registro_saldo_inicial": sin_registro,
    }

    if request.method == "GET":
        return jsonify(diagnostico)

    # ---- POST: aplicar corrección (requiere confirmación explícita) ----
    if request.args.get("confirmar") != "si":
        return jsonify({"error": "Falta confirmar=si", "diagnostico": diagnostico}), 400

    if not cambios:
        return jsonify({"ok": True, "mensaje": "Nada para corregir.", "diagnostico": diagnostico})

    DATOS_DIR.mkdir(exist_ok=True)
    backup_path = DATOS_DIR / f"cierres_backup_pre_fecha_corte_{pid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    shutil.copy2(DB_FILE, backup_path)

    with _get_db() as conn:
        for c in cambios:
            conn.execute("UPDATE francos_saldo_inicial SET fecha_corte=? WHERE legajo=?",
                          (cerrado_en, c["legajo"]))
        conn.commit()

    return jsonify({"ok": True, "backup": str(backup_path), "cambios": cambios})


@app.route("/admin/recalcular-saldo-cierre-francos/<int:cid>")
def admin_recalcular_saldo_cierre_francos(cid):
    """Solo lectura por default: recalcula el snapshot de saldo
    (cierres_francos.saldo_anterior, lo que muestra "Ver saldo al cierre")
    de un cierre manual de francos YA HECHO, usando la misma fórmula
    acotada a fecha_hasta que /francos/cierre/nuevo usa hoy -- para
    cierres viejos, creados antes de ese arreglo, cuyo snapshot de saldo
    puede no coincidir con el detalle de francos capturado (ej. un legajo
    aparece en el detalle pero con tomados=0 en el saldo).

    No toca el PDF ya guardado de "Ver francos" (archivo estático) ni
    ningún otro cierre. Con ?confirmar=si, sobreescribe
    cierres_francos.saldo_anterior de este cid con el snapshot recalculado."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    with _get_db() as conn:
        cf = conn.execute("SELECT * FROM cierres_francos WHERE id=?", (cid,)).fetchone()
        if not cf:
            return jsonify({"error": f"Cierre de francos #{cid} no encontrado"}), 404
        cf = dict(cf)
        fecha_hasta = cf["fecha_hasta"]
        departamento = _normalizar_departamento_web(cf["departamento"])

        try:
            base_anterior = json.loads(cf.get("base_anterior") or "{}")
        except Exception:
            base_anterior = {}
        if not base_anterior:
            return jsonify({"error": "Este cierre no tiene base_anterior guardada -- no se puede recalcular con seguridad."}), 409

        legajos = list(base_anterior.keys())
        ph = ",".join("?" * len(legajos))
        mes_hasta = (fecha_hasta or "")[:7]

        tomados_al_hasta = {r["legajo"]: (r["total"] or 0) for r in conn.execute(
            f"SELECT legajo, SUM(dias) as total FROM francos_tomados "
            f"WHERE legajo IN ({ph}) AND COALESCE(estado,'') != 'Anulado' "
            f"AND {_SQL_CARGADO_EN_DIA} <= ? GROUP BY legajo",
            (*legajos, fecha_hasta)
        )}
        gen_extra_al_hasta = {}
        for r in conn.execute(
            f"SELECT legajo, SUM(dias) as total FROM francos_generados "
            f"WHERE legajo IN ({ph}) AND {_SQL_CARGADO_EN_DIA} <= ? GROUP BY legajo",
            (*legajos, fecha_hasta)
        ):
            gen_extra_al_hasta[r["legajo"]] = gen_extra_al_hasta.get(r["legajo"], 0) + (r["total"] or 0)
        for r in conn.execute(
            f"SELECT legajo, SUM(dias) as total FROM francos_semana_manual "
            f"WHERE legajo IN ({ph}) AND mes <= ? GROUP BY legajo",
            (*legajos, mes_hasta)
        ):
            gen_extra_al_hasta[r["legajo"]] = gen_extra_al_hasta.get(r["legajo"], 0) + (r["total"] or 0)

        nombres_conocidos = {str(e["legajo"]): e["nombre"] for e in _empleados_conocidos()}

        snap_nuevo = {}
        diffs = []
        for leg in legajos:
            base = base_anterior.get(leg) or {}
            base_saldo     = (base.get("saldo") or 0) if base else 0
            base_tom_corte = (base.get("tomados_al_corte") or 0) if base else 0
            base_gen_corte = (base.get("gen_extra_al_corte") or 0) if base else 0

            tomados_periodo   = max(0, tomados_al_hasta.get(leg, 0) - base_tom_corte)
            generados_periodo = max(0, gen_extra_al_hasta.get(leg, 0) - base_gen_corte)
            saldo_final = base_saldo + generados_periodo - tomados_periodo
            nombre = nombres_conocidos.get(leg) or (base.get("nombre") if base else "") or leg

            snap_nuevo[leg] = {
                "nombre": nombre, "saldo_anterior": base_saldo,
                "generados": generados_periodo, "tomados": tomados_periodo,
                "saldo_final": saldo_final,
            }

        try:
            snap_actual = json.loads(cf.get("saldo_anterior") or "{}")
        except Exception:
            snap_actual = {}
        for leg in legajos:
            if snap_actual.get(leg) != snap_nuevo.get(leg):
                diffs.append({"legajo": leg, "actual": snap_actual.get(leg), "correcto": snap_nuevo.get(leg)})

        if request.args.get("confirmar") == "si":
            conn.execute(
                "UPDATE cierres_francos SET saldo_anterior=? WHERE id=?",
                (json.dumps(snap_nuevo, ensure_ascii=False), cid)
            )
            conn.commit()

    return jsonify({
        "cierre_francos_id": cid, "departamento": departamento, "fecha_hasta": fecha_hasta,
        "diferencias": diffs, "aplicado": request.args.get("confirmar") == "si",
    })


@app.route("/admin/diagnostico-francos-huerfanos")
def admin_diagnostico_francos_huerfanos():
    """Solo lectura: busca francos_tomados marcados 'Cerrado' que no
    pertenecen a NINGUNO de los dos mecanismos de cierre.

    Un franco 'Cerrado' es normal y esperado en dos casos:
    - mecanismo cierres_francos (Guardias/Ingenieros/etc.): cierre_francos_id
      apunta al cierre que lo tomó.
    - mecanismo periodos (Redes/Administración, vía _snapshot_francos_cierre):
      NUNCA usa cierre_francos_id (ese campo es exclusivo del otro
      mecanismo) -- en cambio queda una copia en francos_cierre_detalle,
      matcheada por francos_tomados_id (o por tupla de campos para filas
      históricas de antes de esa columna).

    Un huérfano real es 'Cerrado' sin cierre_francos_id Y sin ninguna fila
    correspondiente en francos_cierre_detalle -- no lo reclama ninguno de
    los dos mecanismos.

    Sin ?confirmar=si es solo diagnóstico (no escribe nada). Con
    ?confirmar=si, revierte el estado de los huérfanos encontrados a
    'Aprobado' (no toca cierre_francos_id -- ya es NULL -- ni el saldo,
    que ya los contaba igual estando en 'Cerrado'), para que un futuro
    cierre de su departamento los pueda capturar formalmente."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    with _get_db() as conn:
        huerfanos = [dict(r) for r in conn.execute("""
            SELECT ft.id, ft.legajo, ft.nombre, ft.tipo, ft.fecha_desde, ft.fecha_hasta,
                   ft.dias, ft.estado, ft.estado_antes_cierre, ft.cargado_en,
                   ft.cierre_francos_id
            FROM francos_tomados ft
            WHERE ft.estado='Cerrado' AND ft.cierre_francos_id IS NULL
              AND NOT EXISTS (
                  SELECT 1 FROM francos_cierre_detalle fcd
                  WHERE fcd.francos_tomados_id = ft.id
                     OR (
                         fcd.francos_tomados_id IS NULL
                         AND fcd.legajo = ft.legajo AND fcd.tipo = ft.tipo
                         AND fcd.fecha_desde = ft.fecha_desde AND fcd.fecha_hasta = ft.fecha_hasta
                         AND COALESCE(fcd.fechas_sueltas,'[]') = COALESCE(ft.fechas_sueltas,'[]')
                         AND fcd.dias = ft.dias
                     )
              )
            ORDER BY ft.legajo, ft.fecha_desde
        """).fetchall()]
        vinculado_sin_cerrado = [dict(r) for r in conn.execute("""
            SELECT id, legajo, nombre, tipo, fecha_desde, fecha_hasta, dias,
                   estado, estado_antes_cierre, cargado_en, cierre_francos_id
            FROM francos_tomados
            WHERE cierre_francos_id IS NOT NULL AND COALESCE(estado,'') != 'Cerrado'
            ORDER BY legajo, fecha_desde
        """).fetchall()]

        if request.args.get("confirmar") == "si" and huerfanos:
            ids = [h["id"] for h in huerfanos]
            ph_ids = ",".join("?" * len(ids))
            conn.execute(
                f"UPDATE francos_tomados SET estado='Aprobado' WHERE id IN ({ph_ids})",
                ids,
            )
            conn.commit()

    return jsonify({
        "huerfanos_reales": huerfanos,
        "vinculado_pero_no_cerrado": vinculado_sin_cerrado,
        "total_inconsistencias": len(huerfanos) + len(vinculado_sin_cerrado),
        "corregidos": len(huerfanos) if request.args.get("confirmar") == "si" else 0,
    })


@app.route("/admin/historial-legajo/<legajo>")
def admin_historial_legajo(legajo):
    """Solo lectura: historial completo de un legajo en los dos mecanismos
    de cierre de francos (periodos, para deptos con fichadas; cierres_francos,
    para deptos sin fichadas como Ingenieros), su empleados_extra actual y su
    francos_saldo_inicial actual. Pensada para trazar una transición de
    departamento (ej. Redes -> Ingenieros) y confirmar que no se perdió ni
    duplicó saldo en el cambio."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    leg = str(legajo)

    with _get_db() as conn:
        periodos_leg = [dict(r) for r in conn.execute("""
            SELECT pe.periodo_id, pe.departamento, pe.francos, pe.ot50, pe.ot100,
                   p.cerrado_en, p.fecha_desde, p.fecha_hasta,
                   COALESCE(p.estado,'ACTIVO') as estado
            FROM periodo_empleados pe
            JOIN periodos p ON p.id = pe.periodo_id
            WHERE pe.legajo = ?
            ORDER BY p.cerrado_en
        """, (leg,)).fetchall()]

        detalle_cierres = [dict(r) for r in conn.execute("""
            SELECT periodo_id, dias, estado, fecha_desde, fecha_hasta
            FROM francos_cierre_detalle
            WHERE legajo = ?
            ORDER BY periodo_id, fecha_desde
        """, (leg,)).fetchall()]

        try:
            cierres_francos_leg = [dict(r) for r in conn.execute("""
                SELECT cf.id as cierre_francos_id, cf.departamento, cf.cerrado_en,
                       cf.fecha_hasta, COALESCE(cf.estado,'ACTIVO') as estado,
                       ft.dias, ft.tipo, ft.fecha_desde as franco_fecha_desde,
                       ft.fecha_hasta as franco_fecha_hasta
                FROM francos_tomados ft
                JOIN cierres_francos cf ON cf.id = ft.cierre_francos_id
                WHERE ft.legajo = ?
                ORDER BY cf.cerrado_en
            """, (leg,)).fetchall()]
        except Exception:
            cierres_francos_leg = []

        saldo_inicial = conn.execute(
            "SELECT * FROM francos_saldo_inicial WHERE legajo=?", (leg,)
        ).fetchone()
        empleado_extra = conn.execute(
            "SELECT * FROM empleados_extra WHERE legajo=?", (leg,)
        ).fetchone()

        # Fichas puntuales en vivo (todas, sin filtrar por fecha ni cierre) --
        # para poder ver a simple vista cuáles son de qué mes por su
        # cargado_en, antes de cerrar un período manual.
        francos_tomados_leg = [dict(r) for r in conn.execute("""
            SELECT id, tipo, fecha_desde, fecha_hasta, dias, estado, cargado_en, cierre_francos_id
            FROM francos_tomados WHERE legajo=? ORDER BY cargado_en
        """, (leg,)).fetchall()]
        francos_generados_leg = [dict(r) for r in conn.execute("""
            SELECT id, descripcion, dias, cargado_en, cierre_francos_id
            FROM francos_generados WHERE legajo=? ORDER BY cargado_en
        """, (leg,)).fetchall()]
        francos_semana_manual_leg = [dict(r) for r in conn.execute("""
            SELECT id, semana_num, mes, dias, guardado_en, cierre_francos_id
            FROM francos_semana_manual WHERE legajo=? ORDER BY mes
        """, (leg,)).fetchall()]

    ultimo_cierre_periodos = periodos_leg[-1] if periodos_leg else None
    ultimo_cierre_francos = cierres_francos_leg[-1] if cierres_francos_leg else None

    return jsonify({
        "legajo": leg,
        "empleados_extra_actual": dict(empleado_extra) if empleado_extra else None,
        "francos_saldo_inicial_actual": dict(saldo_inicial) if saldo_inicial else None,
        "historial_periodos_horas_extra": periodos_leg,
        "ultimo_cierre_periodos": ultimo_cierre_periodos,
        "historial_francos_cierre_detalle": detalle_cierres,
        "historial_cierres_francos_manual": cierres_francos_leg,
        "ultimo_cierre_francos_manual": ultimo_cierre_francos,
        "francos_tomados_en_vivo": francos_tomados_leg,
        "francos_generados_en_vivo": francos_generados_leg,
        "francos_semana_manual_en_vivo": francos_semana_manual_leg,
    })


@app.route("/admin/inspeccionar-saldo-anterior/<int:pid>")
def admin_inspeccionar_saldo_anterior(pid):
    """Solo lectura: muestra periodos.saldo_anterior (el snapshot de
    francos_saldo_inicial guardado justo ANTES de que ese cierre corriera
    su actualización automática de saldo — lo que periodo_anular usa para
    restaurar) para el/los legajo(s) indicados, junto al valor actual en
    vivo de francos_saldo_inicial. Sirve para rastrear de dónde salió un
    saldo actual que no coincide con lo esperado."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    legajos = [s.strip() for s in request.args.get("legajos", "").split(",") if s.strip()]

    with _get_db() as conn:
        p = conn.execute("SELECT * FROM periodos WHERE id=?", (pid,)).fetchone()
        if not p:
            return jsonify({"error": f"Período #{pid} no encontrado"}), 404
        saldo_ant = {}
        try:
            saldo_ant = json.loads(p["saldo_anterior"] or "{}")
        except Exception:
            pass
        actuales = {
            r["legajo"]: dict(r)
            for r in conn.execute("SELECT * FROM francos_saldo_inicial").fetchall()
        }

    if legajos:
        saldo_ant = {k: v for k, v in saldo_ant.items() if k in legajos}
        actuales = {k: v for k, v in actuales.items() if k in legajos}

    return jsonify({
        "periodo_id": pid,
        "estado": p["estado"] or "ACTIVO",
        "cerrado_en": p["cerrado_en"],
        "saldo_anterior_guardado_en_este_cierre": saldo_ant,
        "saldo_actual_en_vivo": actuales,
    })


@app.route("/admin/diagnostico-francos")
def admin_diagnostico_francos():
    """Diagnóstico de SOLO LECTURA del estado de los snapshots de francos en
    todos los cierres (no escribe nada en la base de datos).

    - Mecanismo A (periodos / francos_cierre_detalle): para cada cierre
      ACTIVO calcula, vía _seleccionar_francos_ventana_cierre (la misma
      fuente de verdad que /admin/corregir-francos-cierre), cuántos
      registros tiene el snapshot actual vs. cuántos "correctos" le
      corresponderían — sin aplicar ninguna corrección.
    - Mecanismo B (cierres_francos): nunca tiene snapshot inmutable; se
      reportan conteos por departamento.
    - Además señala, por cada cierre que requeriría corrección, los cierres
      posteriores del mismo departamento que podrían verse afectados, y
      legajos cuyo francos_saldo_inicial.fecha_corte no coincide con el
      último cierre ACTIVO de su departamento."""
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401

    periodos_info = []
    with _get_db() as conn:
        periodos_rows = conn.execute(
            "SELECT * FROM periodos WHERE COALESCE(estado,'ACTIVO')='ACTIVO' ORDER BY cerrado_en"
        ).fetchall()
        for p in periodos_rows:
            pid = p["id"]
            emp_rows = conn.execute(
                "SELECT DISTINCT legajo, departamento FROM periodo_empleados WHERE periodo_id=?", (pid,)
            ).fetchall()
            if not emp_rows:
                continue
            legajos = [str(r["legajo"]) for r in emp_rows]
            departamento = _nombre_departamento_visible(emp_rows[0]["departamento"] or "")
            fecha_corte = _normalizar_cargado_en(p["cerrado_en"] or "")
            actuales = conn.execute(
                "SELECT * FROM francos_cierre_detalle WHERE periodo_id=?", (pid,)
            ).fetchall()
            if fecha_corte:
                correctos, _fecha_corte_anterior, legacy = _seleccionar_francos_ventana_cierre(
                    conn, pid, legajos, departamento, fecha_corte)
            else:
                correctos, legacy = [], []

            def _clave(r):
                return (str(r["legajo"]), r["tipo"] or "", r["fecha_desde"] or "",
                        r["fecha_hasta"] or "", r["fechas_sueltas"] or "[]", r["dias"] or 0)
            claves_actuales = {_clave(r) for r in actuales}
            faltantes = [r for r in correctos if _clave(r) not in claves_actuales]

            periodos_info.append({
                "periodo_id": pid,
                "departamento": departamento,
                "cerrado_en": fecha_corte or (p["cerrado_en"] or ""),
                "fecha_desde": p["fecha_desde"] or "",
                "fecha_hasta": p["fecha_hasta"] or "",
                "legajos": legajos,
                "registros_snapshot": len(actuales),
                "snapshot_vacio": len(actuales) == 0,
                "registros_correctos_estimados": len(correctos),
                "registros_faltantes_estimados": len(faltantes),
                "requiere_correccion": len(faltantes) > 0,
                "legacy_sin_cargado_en": len(legacy),
                "ya_corregido_en": p["francos_corregido_en"] or "",
            })

        cf_rows = conn.execute("SELECT * FROM cierres_francos ORDER BY id").fetchall()

        emp_depto = {}
        for r in conn.execute("SELECT DISTINCT legajo, departamento FROM periodo_empleados"):
            emp_depto[str(r["legajo"])] = _nombre_departamento_visible(r["departamento"] or "")
        for r in conn.execute("SELECT legajo, departamento FROM empleados_extra WHERE activo=1"):
            emp_depto.setdefault(str(r["legajo"]), _nombre_departamento_visible(r["departamento"] or ""))
        saldo_inicial_rows = conn.execute("SELECT * FROM francos_saldo_inicial").fetchall()

    # ---- Mecanismo A: agregados por departamento ----
    periodos_por_depto = {}
    snapshot_ok_por_depto = {}
    snapshot_vacio_por_depto = {}
    legajos_por_depto = {}
    for p in periodos_info:
        dep = p["departamento"]
        periodos_por_depto[dep] = periodos_por_depto.get(dep, 0) + 1
        if p["registros_snapshot"] > 0:
            snapshot_ok_por_depto[dep] = snapshot_ok_por_depto.get(dep, 0) + 1
        else:
            snapshot_vacio_por_depto[dep] = snapshot_vacio_por_depto.get(dep, 0) + 1
        legajos_por_depto.setdefault(dep, set()).update(p["legajos"])

    requieren_correccion = [p for p in periodos_info if p["requiere_correccion"]]
    cierres_posteriores_afectados = []
    for p in requieren_correccion:
        posteriores = [
            q["periodo_id"] for q in periodos_info
            if q["departamento"] == p["departamento"] and q["cerrado_en"] > p["cerrado_en"]
        ]
        cierres_posteriores_afectados.append({
            "periodo_id": p["periodo_id"],
            "departamento": p["departamento"],
            "cerrado_en": p["cerrado_en"],
            "registros_faltantes_estimados": p["registros_faltantes_estimados"],
            "periodos_posteriores_mismo_depto": posteriores,
        })

    # ---- Mecanismo B: agregados por departamento (nunca tiene snapshot inmutable) ----
    cierres_francos_por_depto = {}
    legajos_cf_por_depto = {}
    for cf in cf_rows:
        depto_visible = cf["departamento"]
        depto_norm = _normalizar_departamento_web(depto_visible)
        cierres_francos_por_depto[depto_visible] = cierres_francos_por_depto.get(depto_visible, 0) + 1
        with _get_db() as conn:
            legs = [str(r["legajo"]) for r in conn.execute(
                "SELECT DISTINCT legajo FROM periodo_empleados WHERE LOWER(departamento)=? "
                "UNION SELECT legajo FROM empleados_extra WHERE activo=1 AND LOWER(departamento)=?",
                (depto_visible.lower(), depto_norm)
            )]
        legajos_cf_por_depto.setdefault(depto_visible, set()).update(legs)

    # ---- Cadena de saldos: fecha_corte de francos_saldo_inicial vs último cierre ACTIVO del depto ----
    ultimo_cierre_por_depto = {}
    for p in periodos_info:
        dep = p["departamento"]
        actual = ultimo_cierre_por_depto.get(dep)
        if actual is None or p["cerrado_en"] > actual["cerrado_en"]:
            ultimo_cierre_por_depto[dep] = p

    inconsistencias_saldo = []
    for r in saldo_inicial_rows:
        leg = str(r["legajo"])
        dep = emp_depto.get(leg, "")
        ultimo = ultimo_cierre_por_depto.get(dep)
        if not ultimo:
            continue
        fecha_corte_saldo  = (r["fecha_corte"] or "")[:10]
        fecha_corte_cierre = (ultimo["cerrado_en"] or "")[:10]
        if fecha_corte_saldo and fecha_corte_cierre and fecha_corte_saldo != fecha_corte_cierre:
            inconsistencias_saldo.append({
                "legajo": leg,
                "nombre": r["nombre"] or "",
                "departamento": dep,
                "fecha_corte_saldo_inicial": fecha_corte_saldo,
                "ultimo_cierre_id": ultimo["periodo_id"],
                "ultimo_cierre_cerrado_en": fecha_corte_cierre,
            })

    return jsonify({
        "periodos": {
            "total": len(periodos_info),
            "por_departamento": periodos_por_depto,
            "con_snapshot": snapshot_ok_por_depto,
            "snapshot_vacio": snapshot_vacio_por_depto,
            "legajos_participantes_por_departamento": {d: len(s) for d, s in legajos_por_depto.items()},
            "requieren_correccion": requieren_correccion,
            "detalle": periodos_info,
        },
        "cierres_francos_manuales": {
            "nota": "Mecanismo B (cierres_francos) no tiene tabla de snapshot inmutable: "
                    "el PDF se recalcula desde francos_tomados al momento de verlo.",
            "total": len(cf_rows),
            "por_departamento": cierres_francos_por_depto,
            "legajos_participantes_por_departamento": {d: len(s) for d, s in legajos_cf_por_depto.items()},
        },
        "cierres_posteriores_afectados_por_correccion": cierres_posteriores_afectados,
        "inconsistencias_cadena_saldos": inconsistencias_saldo,
    })


@app.route("/admin/ajuste-saldo")
def admin_ajuste_saldo():
    if not _autenticado(): return _requiere_auth()
    saldos = _calcular_saldos()
    with _get_db() as conn:
        fsi_rows = {str(r["legajo"]): dict(r) for r in conn.execute(
            "SELECT legajo, nota, fecha_corte, saldo FROM francos_saldo_inicial"
        )}
    empleados = []
    for s in sorted(saldos, key=_legajo_key):
        leg = str(s.get("legajo", ""))
        f = fsi_rows.get(leg, {})
        empleados.append({
            "legajo":       leg,
            "nombre":       s.get("nombre", ""),
            "departamento": s.get("departamento", ""),
            "saldo_inicial": s.get("saldo_inicial", 0),
            "saldo_actual": s.get("saldo_actual", 0),
            "nota":         f.get("nota", "") or "",
            "fecha_corte":  (f.get("fecha_corte") or "")[:10],
        })
    deptos = {}
    for e in empleados:
        deptos.setdefault(e["departamento"], []).append(e)
    return render_template("admin_ajuste_saldo.html", deptos=deptos)


@app.route("/admin/ajuste-saldo/guardar", methods=["POST"])
def admin_ajuste_saldo_guardar():
    if not _autenticado(): return jsonify({"ok": False, "error": "No autorizado"}), 401
    legajo        = request.form.get("legajo", "").strip()
    nuevo_saldo_s = request.form.get("saldo", "").strip()
    motivo        = request.form.get("motivo", "").strip()
    if not legajo or not nuevo_saldo_s.lstrip("-").isdigit():
        return jsonify({"ok": False, "error": "Datos inválidos"}), 400
    if not motivo:
        return jsonify({"ok": False, "error": "El motivo es obligatorio"}), 400
    nuevo_saldo = int(nuevo_saldo_s)
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    nota_nueva = f"Ajuste manual {ahora[:10]}: {motivo}"
    with _get_db() as conn:
        row = conn.execute(
            "SELECT legajo FROM francos_saldo_inicial WHERE CAST(legajo AS TEXT)=?", (legajo,)
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Empleado no encontrado en saldo_inicial"}), 404
        conn.execute(
            "UPDATE francos_saldo_inicial SET saldo=?, nota=?, cargado_en=? WHERE CAST(legajo AS TEXT)=?",
            (nuevo_saldo, nota_nueva, ahora, legajo)
        )
        conn.commit()
    return jsonify({"ok": True, "nota": nota_nueva, "saldo": nuevo_saldo})


@app.route("/admin/reset", methods=["POST"])
def admin_reset():
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    import shutil
    global _sesion

    # Limpiar sesión en memoria y en disco
    _sesion.clear()
    _guardar_sesion(_sesion)

    # Borrar confirmaciones
    if CONFIRM_DIR.exists():
        shutil.rmtree(CONFIRM_DIR)

    # Borrar semanas (CSV + metadata)
    if SEMANAS_DIR.exists():
        shutil.rmtree(SEMANAS_DIR)

    # Borrar periodos JSON
    if PERIODOS_DIR.exists():
        shutil.rmtree(PERIODOS_DIR)

    # Resetear base de datos
    with _get_db() as conn:
        conn.execute("DELETE FROM periodo_empleados")
        conn.execute("DELETE FROM periodos")
        conn.commit()

    return jsonify({"ok": True})


# ═══════════════════════════════════════════════
# FRANCOS TOMADOS
# ═══════════════════════════════════════════════
def _calcular_saldos():
    """Saldo por empleado = saldo_inicial + generados_nuevos - tomados_nuevos.
    Las columnas tomados_al_corte, gen_extra_al_corte y fecha_corte permiten que
    el cierre mensual actualice saldo_inicial sin perder ni duplicar datos."""
    with _get_db() as conn:
        iniciales = {}
        for r in conn.execute(
            "SELECT legajo, saldo, tomados_al_corte, gen_extra_al_corte, fecha_corte "
            "FROM francos_saldo_inicial"
        ):
            iniciales[r["legajo"]] = {
                "saldo":              r["saldo"],
                "tomados_al_corte":   r["tomados_al_corte"]   or 0,
                "gen_extra_al_corte": r["gen_extra_al_corte"] or 0,
                "fecha_corte":        r["fecha_corte"]         or "2026-05-21",
            }

        # Todos los generados por períodos cerrados (se filtra por fecha_corte per-empleado abajo)
        all_pe = conn.execute(
            "SELECT pe.legajo, pe.francos, p.fecha_hasta "
            "FROM periodo_empleados pe "
            "JOIN periodos p ON pe.periodo_id = p.id "
            "WHERE COALESCE(p.estado,'ACTIVO') <> 'ANULADO'"
        ).fetchall()

        gen_manual_raw   = {r["legajo"]: (r["total"] or 0) for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_generados GROUP BY legajo")}
        gen_parcial      = {r["legajo"]: (r["total"] or 0) for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_semana_parcial GROUP BY legajo")}
        gen_manual_sem   = {r["legajo"]: (r["total"] or 0) for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_semana_manual GROUP BY legajo")}
        tomados_raw      = {r["legajo"]: (r["total"] or 0) for r in conn.execute("SELECT legajo, SUM(dias) as total FROM francos_tomados WHERE COALESCE(estado,'') != 'Anulado' GROUP BY legajo")}

    # gen_periodos filtrado por fecha_corte de cada empleado
    gen_periodos_por_emp = {}
    for row in all_pe:
        leg = row["legajo"]
        corte = iniciales.get(leg, {}).get("fecha_corte", "2026-05-21")
        if (row["fecha_hasta"] or "") > corte:
            gen_periodos_por_emp[leg] = gen_periodos_por_emp.get(leg, 0) + (row["francos"] or 0)

    resultado = []
    for emp in _empleados_conocidos():
        leg  = emp["legajo"]
        ini  = iniciales.get(leg, {"saldo": 0, "tomados_al_corte": 0, "gen_extra_al_corte": 0, "fecha_corte": "2026-05-21"})
        si   = ini["saldo"]
        t_corte   = ini["tomados_al_corte"]
        ge_corte  = ini["gen_extra_al_corte"]

        gen_extra_total = gen_manual_raw.get(leg, 0) + gen_manual_sem.get(leg, 0)
        gen = (gen_periodos_por_emp.get(leg, 0)
               + max(0, gen_extra_total - ge_corte)
               + gen_parcial.get(leg, 0))
        tom = max(0, tomados_raw.get(leg, 0) - t_corte)

        resultado.append({
            "legajo":       leg,
            "nombre":       emp["nombre"],
            "departamento": emp["departamento"],
            "saldo_inicial": si,
            "generados":    gen,
            "tomados":      tom,
            "saldo_actual": si + gen - tom,
        })
    return resultado

def _empleados_conocidos():
    """Lista deduplicada de {legajo, nombre, departamento} ordenada por depto y legajo.
    Fuentes (menor a mayor prioridad): empleados_extra → periodo_empleados → sesion.json
    El departamento de cada fuente se normaliza al nombre visible canónico
    (Redes, Administración, Guardias, Internet, Telefonía) para que un mismo
    empleado/depto no quede dividido en dos grupos por diferencias de
    mayúsculas, tildes o espacios entre fuentes.
    """
    vistos = {}
    with _get_db() as conn:
        # empleados_extra primero (base maestra de depts sin fichadas)
        for row in conn.execute(
            "SELECT legajo, nombre, departamento FROM empleados_extra WHERE activo=1"
        ):
            vistos[str(row["legajo"])] = {"nombre": row["nombre"], "departamento": _nombre_departamento_visible(row["departamento"] or "")}
        # periodos cerrados (sobreescribe si coincide legajo)
        for row in conn.execute("SELECT DISTINCT legajo, nombre, departamento FROM periodo_empleados"):
            vistos[str(row["legajo"])] = {"nombre": row["nombre"], "departamento": _nombre_departamento_visible(row["departamento"] or "")}
    # sesion activa (máxima prioridad, es la más reciente)
    for d in _sesion.values():
        leg = str(d.get("legajo", ""))
        if leg:
            vistos[leg] = {"nombre": d.get("nombre", ""), "departamento": _nombre_departamento_visible(d.get("departamento", ""))}
    # En el módulo Francos, 100 y 101 pertenecen exclusivamente a
    # Ingenieros. Una sesión biométrica anterior puede conservarlos como
    # Redes, pero no debe sobreescribir la fuente manual empleados_extra.
    with _get_db() as conn:
        for row in conn.execute(
            "SELECT legajo, nombre, departamento FROM empleados_extra "
            "WHERE activo=1 AND legajo IN ('100','101')"
        ):
            vistos[str(row["legajo"])] = {
                "nombre": row["nombre"],
                "departamento": _nombre_departamento_visible(row["departamento"] or ""),
            }
    return sorted(
        [{"legajo": k, "nombre": v["nombre"], "departamento": v["departamento"]} for k, v in vistos.items()],
        key=lambda e: (e["departamento"].lower(), int(e["legajo"]) if e["legajo"].isdigit() else 0)
    )

def _cargar_feriados_config():
    try:
        cfg = json.loads(Path("config.json").read_text(encoding="utf-8"))
        from datetime import date as _date
        return {datetime.strptime(f, "%Y-%m-%d").date() for f in cfg.get("feriados", [])}
    except Exception:
        return set()

def _guardar_feriados_config(feriados):
    """Persiste la lista de feriados (strings 'YYYY-MM-DD') en config.json sin tocar otras claves."""
    config_path = Path("config.json")
    config = {}
    if config_path.exists():
        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
        except Exception:
            config = {}
    config["feriados"] = sorted(set(feriados))
    config_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")

def _dias_habiles(fecha_desde_str, fecha_hasta_str):
    """Cuenta días hábiles (lun-vie, excluyendo feriados de config.json)."""
    desde = _parse_fecha(fecha_desde_str)
    hasta = _parse_fecha(fecha_hasta_str)
    if not desde or not hasta:
        return 0
    # Normalizar a date (parse_fecha puede devolver datetime o date)
    if isinstance(desde, datetime):
        desde = desde.date()
    if isinstance(hasta, datetime):
        hasta = hasta.date()
    feriados = _cargar_feriados_config()
    total = 0
    cur = desde
    while cur <= hasta:
        if cur.weekday() < 5 and cur not in feriados:
            total += 1
        cur += timedelta(days=1)
    return total

def _fechas_habiles_set(desde, hasta, feriados):
    """Devuelve el conjunto de dates hábiles (lun-vie, no feriado) en [desde, hasta]."""
    if isinstance(desde, datetime): desde = desde.date()
    if isinstance(hasta, datetime): hasta = hasta.date()
    result = set()
    cur = desde
    while cur <= hasta:
        if cur.weekday() < 5 and cur not in feriados:
            result.add(cur)
        cur += timedelta(days=1)
    return result


def _fechas_del_registro(tipo, fecha_desde_str, fecha_hasta_str, fechas_sueltas_json, feriados):
    """Conjunto de dates efectivas de un registro francos_tomados."""
    if tipo == "SUELTAS":
        try:
            lista = json.loads(fechas_sueltas_json or "[]")
        except Exception:
            lista = []
        result = set()
        for s in lista:
            f = _parse_fecha(s)
            if f:
                result.add(f.date() if isinstance(f, datetime) else f)
        return result
    desde = _parse_fecha(fecha_desde_str)
    hasta = _parse_fecha(fecha_hasta_str)
    if not desde or not hasta:
        return set()
    return _fechas_habiles_set(desde, hasta, feriados)


def _es_guardias(conn, legajo):
    """Devuelve True si el empleado pertenece al departamento GUARDIAS."""
    row = conn.execute(
        "SELECT departamento FROM empleados_extra WHERE legajo=? AND activo=1", (str(legajo),)
    ).fetchone()
    if not row:
        return False
    dep = (row["departamento"] or "").upper().strip()
    return dep in ("GUARDIAS", "GUARDIA")

def _fechas_liberadas_por_anulacion(conn, legajo, feriados):
    """Fechas de francos_cierre_detalle ya liberadas vía
    /francos/anular-cerrado (francos_anulaciones_cerrados). No modifica
    francos_cierre_detalle -- el cierre histórico queda intacto -- solo
    cambia qué fechas cuentan como 'ocupadas' en la validación en vivo."""
    liberadas = set()
    for row in conn.execute(
        "SELECT tipo, fecha_desde, fecha_hasta, fechas_sueltas FROM francos_anulaciones_cerrados WHERE legajo=?",
        (str(legajo),)
    ).fetchall():
        liberadas |= _fechas_del_registro(
            row["tipo"], row["fecha_desde"], row["fecha_hasta"], row["fechas_sueltas"], feriados
        )
    return liberadas


def _validar_franco_nuevo(conn, legajo, tipo, fecha_desde_str, fecha_hasta_str,
                          fechas_sueltas_lista, exclude_id=None):
    """
    Valida que las fechas sean hábiles y no se superpongan con registros existentes.
    Devuelve string con el error, o None si todo está bien.
    Los empleados de GUARDIAS pueden tomar franco cualquier día (feriados y fines de semana incluidos).
    """
    feriados = _cargar_feriados_config()
    guardias = _es_guardias(conn, legajo)

    # Construir el set de fechas nuevas
    if tipo == "SUELTAS":
        nuevas = set()
        for s in fechas_sueltas_lista:
            f = _parse_fecha(s)
            if f:
                if isinstance(f, datetime): f = f.date()
                if not guardias and f.weekday() >= 5:
                    return f"La fecha {f.strftime('%d/%m/%Y')} es fin de semana"
                if not guardias and f in feriados:
                    return f"La fecha {f.strftime('%d/%m/%Y')} es feriado"
                nuevas.add(f)
    elif tipo == "UNICO":
        f = _parse_fecha(fecha_desde_str)
        if not f:
            return "Fecha inválida"
        if isinstance(f, datetime): f = f.date()
        if not guardias and f.weekday() >= 5:
            return f"La fecha {f.strftime('%d/%m/%Y')} es fin de semana"
        if not guardias and f in feriados:
            return f"La fecha {f.strftime('%d/%m/%Y')} es feriado"
        nuevas = {f}
    else:  # RANGO
        desde = _parse_fecha(fecha_desde_str)
        hasta = _parse_fecha(fecha_hasta_str)
        if not desde or not hasta:
            return "Fechas inválidas"
        if guardias:
            # Para guardias contar todos los días corridos
            if isinstance(desde, datetime): desde = desde.date()
            if isinstance(hasta, datetime): hasta = hasta.date()
            nuevas = set()
            cur = desde
            while cur <= hasta:
                nuevas.add(cur)
                cur += timedelta(days=1)
        else:
            nuevas = _fechas_habiles_set(desde, hasta, feriados)

    if not nuevas:
        return "El período no contiene días válidos"

    # Verificar superposición con registros vigentes (francos_tomados, no anulados)
    query = "SELECT id, tipo, fecha_desde, fecha_hasta, fechas_sueltas FROM francos_tomados WHERE COALESCE(estado,'') != 'Anulado' AND legajo=?"
    params = [str(legajo)]
    if exclude_id:
        query += " AND id != ?"
        params.append(exclude_id)
    for row in conn.execute(query, params).fetchall():
        existentes = _fechas_del_registro(
            row["tipo"] if isinstance(row, sqlite3.Row) else row[1],
            row["fecha_desde"] if isinstance(row, sqlite3.Row) else row[2],
            row["fecha_hasta"] if isinstance(row, sqlite3.Row) else row[3],
            row["fechas_sueltas"] if isinstance(row, sqlite3.Row) else row[4],
            feriados,
        )
        solapadas = nuevas & existentes
        if solapadas:
            primera = min(solapadas).strftime("%d/%m/%Y")
            return f"El empleado ya tiene franco registrado el {primera}"

    # También verificar contra historial de cierres (cubre francos eliminados antes del cambio a Anular)
    liberadas = _fechas_liberadas_por_anulacion(conn, legajo, feriados)
    for row in conn.execute(
        "SELECT tipo, fecha_desde, fecha_hasta, fechas_sueltas FROM francos_cierre_detalle WHERE legajo=?",
        (str(legajo),)
    ).fetchall():
        existentes = _fechas_del_registro(
            row["tipo"], row["fecha_desde"], row["fecha_hasta"], row["fechas_sueltas"], feriados,
        ) - liberadas
        solapadas = nuevas & existentes
        if solapadas:
            primera = min(solapadas).strftime("%d/%m/%Y")
            return f"El empleado ya tiene franco de cierre registrado el {primera}"

    return None


@app.route("/francos")
def francos():
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        registros = conn.execute(
            "SELECT * FROM francos_tomados ORDER BY cargado_en DESC, id DESC"
        ).fetchall()
    registros = [dict(r) for r in registros]
    for r in registros:
        if r["tipo"] == "SUELTAS":
            try:
                r["fechas_lista"] = json.loads(r["fechas_sueltas"])
            except Exception:
                r["fechas_lista"] = []
        else:
            r["fechas_lista"] = []
    empleados_raw = _empleados_conocidos()
    emp_depto = {e["legajo"]: e["departamento"] for e in empleados_raw}
    for r in registros:
        r["departamento"] = emp_depto.get(r["legajo"], "")
    _reg_dict = {}
    for r in registros:
        dep = r["departamento"] or "Sin departamento"
        _reg_dict.setdefault(dep, []).append(r)
    registros_por_depto = sorted(
        [{"departamento": k, "registros": v} for k, v in _reg_dict.items() if not _depto_francos_oculto(k)],
        key=lambda x: x["departamento"].lower()
    )
    empleados_por_depto = {}
    for e in empleados_raw:
        dep = e["departamento"] or "Sin departamento"
        empleados_por_depto.setdefault(dep, []).append(e)
    empleados_grupos = [{"departamento": k, "empleados": v} for k, v in sorted(empleados_por_depto.items()) if not _depto_francos_oculto(k)]
    saldos_raw = _calcular_saldos()
    sf_grupos = {}
    for s in saldos_raw:
        dep = s["departamento"] or "Sin departamento"
        sf_grupos.setdefault(dep, []).append(s)
    saldos_por_depto_f = []
    for k, v in sorted(sf_grupos.items()):
        if _depto_francos_oculto(k):
            continue
        saldos_por_depto_f.append({
            "departamento":    k,
            "empleados":       v,
            "total_inicial":   sum(e["saldo_inicial"] for e in v),
            "total_generados": sum(e["generados"]     for e in v),
            "total_tomados":   sum(e["tomados"]       for e in v),
            "total_actual":    sum(e["saldo_actual"]  for e in v),
        })
    total_general_f = {
        "inicial":   sum(g["total_inicial"]   for g in saldos_por_depto_f),
        "generados": sum(g["total_generados"] for g in saldos_por_depto_f),
        "tomados":   sum(g["total_tomados"]   for g in saldos_por_depto_f),
        "actual":    sum(g["total_actual"]    for g in saldos_por_depto_f),
    }
    departamentos = _departamentos_francos_disponibles(empleados_raw)
    # Deptos manuales con valores guardados por semana/mes
    mes_actual = datetime.now().strftime("%Y-%m")
    with _get_db() as conn:
        gen_manual = [dict(r) for r in conn.execute(
            "SELECT * FROM francos_generados ORDER BY cargado_en DESC, id DESC"
        ) if not _depto_francos_oculto(r["departamento"])]
        emps_extra = conn.execute(
            "SELECT legajo, nombre, departamento FROM empleados_extra WHERE activo=1 ORDER BY departamento, CAST(legajo AS INTEGER)"
        ).fetchall()
        manual_guardados = {}
        manual_cierres = {}
        for r in conn.execute(
            "SELECT legajo, semana_num, dias, cierre_francos_id "
            "FROM francos_semana_manual WHERE mes=?", (mes_actual,)
        ):
            manual_guardados.setdefault(str(r["legajo"]), {})[r["semana_num"]] = r["dias"]
            if r["cierre_francos_id"] is not None:
                manual_cierres.setdefault(str(r["legajo"]), {})[r["semana_num"]] = r["cierre_francos_id"]
    _DEPTOS_AUTO = {"redes", "administracion"}
    deptos_manuales = {}
    for e in emps_extra:
        dep_norm = _normalizar_departamento_web(e["departamento"] or "")
        if dep_norm in _DEPTOS_AUTO or _depto_francos_oculto(e["departamento"]):
            continue
        dep = e["departamento"] or "Sin departamento"
        deptos_manuales.setdefault(dep, []).append({
            "legajo": e["legajo"], "nombre": e["nombre"],
            "semanas": manual_guardados.get(str(e["legajo"]), {}),
            "cierres": manual_cierres.get(str(e["legajo"]), {}),
        })
    deptos_manuales_list = [{"departamento": k, "empleados": v} for k, v in sorted(deptos_manuales.items())]
    return render_template("francos.html",
                           registros=registros,
                           registros_por_depto=registros_por_depto,
                           saldos_por_depto=saldos_por_depto_f,
                           total_general=total_general_f,
                           empleados_grupos=empleados_grupos,
                           departamentos=departamentos,
                           gen_manual=gen_manual,
                           deptos_manuales=deptos_manuales_list,
                           mes_actual=mes_actual,
                           firma=FIRMA_SUPERVISOR)


@app.route("/francos/saldos")
def francos_saldos():
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        iniciales   = {r["legajo"]: dict(r) for r in conn.execute("SELECT * FROM francos_saldo_inicial")}
        gen_manual  = [dict(r) for r in conn.execute(
            "SELECT * FROM francos_generados ORDER BY cargado_en DESC, id DESC"
        )]
    empleados_raw = _empleados_conocidos()
    por_depto = {}
    for e in empleados_raw:
        dep = e["departamento"] or "Sin departamento"
        por_depto.setdefault(dep, []).append(e)
    empleados_grupos = [{"departamento": dep, "empleados": emps} for dep, emps in sorted(por_depto.items())]
    departamentos = _departamentos_francos_disponibles(empleados_raw, incluir_ocultos=True)
    # Agrupar saldos por departamento
    saldos_raw = _calcular_saldos()
    saldos_grupos = {}
    for s in saldos_raw:
        dep = s["departamento"] or "Sin departamento"
        saldos_grupos.setdefault(dep, []).append(s)
    saldos_por_depto = []
    for k, v in sorted(saldos_grupos.items()):
        saldos_por_depto.append({
            "departamento":    k,
            "empleados":       v,
            "total_inicial":   sum(e["saldo_inicial"] for e in v),
            "total_generados": sum(e["generados"]     for e in v),
            "total_tomados":   sum(e["tomados"]       for e in v),
            "total_actual":    sum(e["saldo_actual"]  for e in v),
        })
    total_general = {
        "inicial":   sum(g["total_inicial"]   for g in saldos_por_depto),
        "generados": sum(g["total_generados"] for g in saldos_por_depto),
        "tomados":   sum(g["total_tomados"]   for g in saldos_por_depto),
        "actual":    sum(g["total_actual"]    for g in saldos_por_depto),
    }
    return render_template("francos_saldos.html",
                           saldos_por_depto=saldos_por_depto,
                           total_general=total_general,
                           iniciales=iniciales,
                           empleados_grupos=empleados_grupos,
                           gen_manual=gen_manual,
                           departamentos=departamentos)


@app.route("/francos/pdf_depto")
def francos_pdf_depto():
    """PDF de francos tomados por departamento (sin cierre — para Guardias, Internet, Telefonía, etc.)."""
    if not _autenticado(): return _requiere_auth()
    depto_param = request.args.get("depto", "").strip()
    hoy = datetime.now().strftime("%Y-%m-%d")
    with _get_db() as conn:
        if depto_param and depto_param != "__todos__":
            depto_norm = _normalizar_departamento_web(depto_param)
            legajos = [str(r["legajo"]) for r in conn.execute(
                "SELECT DISTINCT legajo FROM periodo_empleados WHERE departamento=? "
                "UNION SELECT legajo FROM empleados_extra WHERE activo=1 AND LOWER(departamento)=?",
                (_nombre_departamento_visible(depto_norm), depto_norm)
            )]
            if not legajos:
                return "Sin empleados para ese departamento.", 404
            ph = ",".join("?" * len(legajos))
            rows = conn.execute(f"""
                SELECT ft.legajo, ft.nombre, ft.tipo, ft.fecha_desde, ft.fecha_hasta,
                       ft.fechas_sueltas, ft.dias, ft.estado, ft.fecha_emision,
                       ft.autorizado_por, ft.observaciones
                FROM francos_tomados ft
                WHERE ft.legajo IN ({ph})
                  AND COALESCE(ft.estado,'') NOT IN ('Anulado', 'Cerrado')
                ORDER BY CAST(ft.legajo AS INTEGER), ft.fecha_desde
            """, legajos).fetchall()
            depto_label = _nombre_departamento_visible(depto_norm)
        else:
            rows = conn.execute("""
                SELECT legajo, nombre, tipo, fecha_desde, fecha_hasta,
                       fechas_sueltas, dias, estado, fecha_emision, autorizado_por, observaciones
                FROM francos_tomados
                WHERE COALESCE(estado,'') NOT IN ('Anulado', 'Cerrado')
                ORDER BY CAST(legajo AS INTEGER), fecha_desde
            """).fetchall()
            depto_label = "Todos los departamentos"
    francos_list = [{**dict(r), "departamento": depto_label} for r in rows]
    pid_fake = f"depto_{depto_param or 'todos'}_{hoy.replace('-','')}"
    _generar_pdf_francos_cierre(pid_fake, francos_list, hoy)
    import glob as _glob
    matches = sorted(_glob.glob(str(Path(f"reportes/francos_cierre_{pid_fake}_*.pdf"))))
    if not matches:
        return "Error generando PDF.", 500
    nombre_dl = f"francos_{(depto_param or 'todos').replace(' ','_')}_{hoy}.pdf"
    return send_file(matches[-1], mimetype="application/pdf", as_attachment=False, download_name=nombre_dl)


@app.route("/francos/saldos/exportar")
def francos_saldos_exportar():
    if not _autenticado(): return _requiere_auth()
    depto_filtro = request.args.get("depto", "").strip().lower()
    saldos_raw = _calcular_saldos()
    from io import StringIO
    import csv as csv_mod
    si = StringIO()
    w = csv_mod.writer(si, delimiter=";")
    w.writerow(["Departamento", "Legajo", "Nombre", "Saldo inicial", "Generados", "Tomados", "Saldo actual"])
    for s in saldos_raw:
        dep = s.get("departamento") or ""
        if depto_filtro and dep.lower() != depto_filtro:
            continue
        w.writerow([dep, s["legajo"], s["nombre"],
                    s["saldo_inicial"], s["generados"], s["tomados"], s["saldo_actual"]])
    output = si.getvalue().encode("utf-8-sig")
    from flask import Response
    nombre_archivo = f"saldos_francos_{depto_filtro or 'todos'}.csv"
    return Response(output, mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.route("/francos/saldos/guardar", methods=["POST"])
def francos_saldos_guardar():
    if not _autenticado(): return _requiere_auth()
    legajo = request.form.get("legajo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    try:
        saldo = int(request.form.get("saldo", "0").strip())
    except ValueError:
        saldo = 0
    nota   = request.form.get("nota", "").strip()
    ahora  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not legajo or not nombre:
        return redirect(url_for("francos_saldos") + "?error=datos_requeridos")
    with _get_db() as conn:
        conn.execute("""
            INSERT INTO francos_saldo_inicial (legajo, nombre, saldo, nota, cargado_en)
            VALUES (?,?,?,?,?)
            ON CONFLICT(legajo) DO UPDATE SET nombre=excluded.nombre, saldo=excluded.saldo, nota=excluded.nota, cargado_en=excluded.cargado_en
        """, (legajo, nombre, saldo, nota, ahora))
        conn.commit()
    return redirect(url_for("francos_saldos"))


@app.route("/francos/generados/nuevo", methods=["POST"])
def francos_generados_nuevo():
    if not _autenticado(): return _requiere_auth()
    legajo      = request.form.get("legajo", "").strip()
    nombre      = request.form.get("nombre", "").strip()
    departamento = request.form.get("departamento", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    try:
        dias = int(request.form.get("dias", "1").strip())
    except ValueError:
        dias = 1
    if not legajo or not nombre or dias < 1:
        return redirect(url_for("francos_saldos") + "?error=generados_datos_requeridos")
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _get_db() as conn:
        conn.execute("""
            INSERT INTO francos_generados (legajo, nombre, departamento, descripcion, dias, cargado_en)
            VALUES (?,?,?,?,?,?)
        """, (legajo, nombre, departamento, descripcion, dias, ahora))
        conn.commit()
    return redirect(url_for("francos_saldos"))


@app.route("/francos/generados/eliminar/<int:gen_id>", methods=["POST"])
def francos_generados_eliminar(gen_id):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        generado = conn.execute(
            "SELECT cierre_francos_id FROM francos_generados WHERE id=?", (gen_id,)
        ).fetchone()
        if generado and generado["cierre_francos_id"] is not None:
            return redirect(url_for("francos_saldos") + "?error=movimiento_cerrado")
        conn.execute("DELETE FROM francos_generados WHERE id=?", (gen_id,))
        conn.commit()
    return redirect(url_for("francos_saldos"))


@app.route("/francos/nuevo", methods=["POST"])
def francos_nuevo():
    if not _autenticado(): return _requiere_auth()
    legajo         = request.form.get("legajo", "").strip()
    nombre         = request.form.get("nombre", "").strip()
    tipo           = request.form.get("tipo", "RANGO").strip()
    estado         = request.form.get("estado", "Aprobado").strip()
    obs            = request.form.get("observaciones", "").strip()
    fecha_emision  = request.form.get("fecha_emision", "").strip()
    autorizado_por = request.form.get("autorizado_por", "").strip()
    ahora          = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not legajo or not nombre:
        return redirect(url_for("francos") + "?error=legajo_requerido")

    fechas_sueltas_lista = []
    if tipo == "UNICO":
        fecha_unica = request.form.get("fecha_unica", "").strip()
        if not fecha_unica:
            return redirect(url_for("francos") + "?error=fechas_requeridas")
        fecha_desde = fecha_unica
        fecha_hasta = fecha_unica
        dias = 1
        fechas_sueltas_json = "[]"
    elif tipo == "RANGO":
        fecha_desde = request.form.get("fecha_desde", "").strip()
        fecha_hasta = request.form.get("fecha_hasta", "").strip()
        if not fecha_desde or not fecha_hasta:
            return redirect(url_for("francos") + "?error=fechas_requeridas")
        dias = _dias_habiles(fecha_desde, fecha_hasta)
        if dias <= 0:
            return redirect(url_for("francos") + "?error=dias_cero")
        fechas_sueltas_json = "[]"
    else:
        fechas_raw = request.form.get("fechas_sueltas", "").strip()
        fechas_lista = [f.strip() for f in fechas_raw.replace(",", "\n").split("\n") if f.strip()]
        fechas_lista = [f for f in fechas_lista if _parse_fecha(f)]
        if not fechas_lista:
            return redirect(url_for("francos") + "?error=fechas_requeridas")
        fechas_sueltas_lista = fechas_lista
        dias = len(fechas_lista)
        fechas_sueltas_json = json.dumps(sorted(fechas_lista))
        fecha_desde = min(fechas_lista)
        fecha_hasta = max(fechas_lista)

    with _get_db() as conn:
        error_val = _validar_franco_nuevo(
            conn, legajo, tipo, fecha_desde, fecha_hasta, fechas_sueltas_lista
        )
        if error_val:
            return redirect(url_for("francos") + "?error=" + error_val)
        conn.execute(
            """INSERT INTO francos_tomados
               (legajo, nombre, tipo, fecha_desde, fecha_hasta, fechas_sueltas, dias, estado, observaciones, fecha_emision, autorizado_por, cargado_en)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (legajo, nombre, tipo, fecha_desde, fecha_hasta, fechas_sueltas_json, dias, estado, obs, fecha_emision, autorizado_por, ahora)
        )
        conn.commit()
    return redirect(url_for("francos"))

def _devolver_saldo_franco_anulado(conn, legajo, dias, cargado_en):
    """Si el franco anulado ya estaba contado en tomados_al_corte, devuelve
    los días al saldo_inicial y descuenta de tomados_al_corte para no
    contarlos también en el cálculo en vivo (_calcular_saldos).

    tomados_al_corte se fija en cada cierre como SUM(dias) de TODO
    francos_tomados no anulado en ese momento (sin filtrar por fecha tomada,
    ver periodo_cerrar) — por eso lo que decide si ya estaba contado es
    cuándo se CARGÓ el franco (cargado_en), no la fecha en que se toma
    (fecha_desde/fecha_hasta), que puede ser futura respecto del corte.
    Los francos legacy sin cargado_en son anteriores a que existiera esa
    columna y por lo tanto siempre estaban ya contados en cualquier corte.

    Si el franco se cargó después del corte, no hace falta tocar nada:
    _calcular_saldos ya excluye 'Anulado' de tomados_raw y el saldo sube solo."""
    leg = str(legajo)
    si = conn.execute(
        "SELECT saldo, tomados_al_corte, fecha_corte FROM francos_saldo_inicial WHERE legajo=?", (leg,)
    ).fetchone()
    if not si or dias <= 0:
        return
    cargado_norm = _normalizar_cargado_en(cargado_en or "")
    ya_contado = (not cargado_norm) or (cargado_norm <= (si["fecha_corte"] or ""))
    if ya_contado:
        conn.execute(
            "UPDATE francos_saldo_inicial SET saldo=saldo+?, tomados_al_corte=MAX(0,tomados_al_corte-?) WHERE legajo=?",
            (dias, dias, leg)
        )


@app.route("/francos/eliminar/<int:fid>", methods=["POST"])
def francos_eliminar(fid):
    if not _autenticado(): return _requiere_auth()
    motivo = request.form.get("motivo", "").strip()
    with _get_db() as conn:
        franco = conn.execute(
            "SELECT legajo, dias, cargado_en, observaciones, cierre_francos_id, estado "
            "FROM francos_tomados WHERE id=?", (fid,)
        ).fetchone()
        if not franco:
            return redirect(url_for("francos"))
        if franco["cierre_francos_id"] is not None or (franco["estado"] or "") == "Cerrado":
            # "Cerrado" == capturado por un cierre de periodos (Redes/Administración);
            # para anular ese franco sin romper el snapshot histórico hay que
            # usar /francos/anular-cerrado, que registra la devolución trazable.
            return redirect(url_for("francos") + "?error=movimiento_cerrado")
        obs_actual = franco["observaciones"] or ""
        nueva_obs = f"[ANULADO: {motivo}] {obs_actual}".strip() if motivo else f"[ANULADO] {obs_actual}".strip()
        conn.execute("UPDATE francos_tomados SET estado='Anulado', observaciones=? WHERE id=?", (nueva_obs, fid))
        _devolver_saldo_franco_anulado(conn, franco["legajo"], franco["dias"] or 0, franco["cargado_en"] or "")
        conn.commit()
    return redirect(url_for("francos"))

@app.route("/francos/aprobar/<int:fid>", methods=["POST"])
def francos_aprobar(fid):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        franco = conn.execute(
            "SELECT cierre_francos_id, estado FROM francos_tomados WHERE id=?", (fid,)
        ).fetchone()
        if franco and (franco["cierre_francos_id"] is not None or (franco["estado"] or "") == "Cerrado"):
            return redirect(url_for("francos") + "?error=movimiento_cerrado")
        conn.execute(
            "UPDATE francos_tomados SET estado='Aprobado' "
            "WHERE id=? AND cierre_francos_id IS NULL AND COALESCE(estado,'') != 'Cerrado'", (fid,)
        )
        conn.commit()
    return redirect(url_for("francos"))


@app.route("/francos/anular-cerrado/<int:fid>", methods=["POST"])
def francos_anular_cerrado(fid):
    """Anula un franco ya 'Cerrado' (perteneciente a un cierre histórico) sin
    modificar ese cierre: libera las fechas, devuelve el saldo y registra un
    movimiento compensatorio que se mostrará en el próximo cierre mensual del
    empleado. Ver francos_anulaciones_cerrados y _fechas_liberadas_por_anulacion."""
    if not _autenticado(): return _requiere_auth()
    motivo = request.form.get("motivo", "").strip()
    if not motivo:
        return redirect(url_for("francos") + "?error=motivo_requerido")
    usuario = request.form.get("usuario", "").strip() or FIRMA_SUPERVISOR
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _get_db() as conn:
        franco = conn.execute(
            "SELECT * FROM francos_tomados WHERE id=?", (fid,)
        ).fetchone()
        if not franco:
            return redirect(url_for("francos") + "?error=franco_no_encontrado")
        if franco["cierre_francos_id"] is not None:
            return redirect(url_for("francos") + "?error=movimiento_cerrado")
        if (franco["estado"] or "") != "Cerrado":
            return redirect(url_for("francos") + "?error=franco_no_cerrado")

        conn.execute(
            "UPDATE francos_tomados SET estado='Anulado', fecha_anulacion=?, "
            "motivo_anulacion=?, usuario_anulacion=? WHERE id=?",
            (ahora, motivo, usuario, fid)
        )
        _devolver_saldo_franco_anulado(conn, franco["legajo"], franco["dias"] or 0, franco["cargado_en"] or "")

        # Buscar en qué cierre histórico figuraba (solo informativo, no se modifica)
        origen = conn.execute(
            "SELECT periodo_id, departamento FROM francos_cierre_detalle "
            "WHERE legajo=? AND tipo=? AND fecha_desde=? AND fecha_hasta=? AND fechas_sueltas=? "
            "ORDER BY id DESC LIMIT 1",
            (franco["legajo"], franco["tipo"], franco["fecha_desde"] or "",
             franco["fecha_hasta"] or "", franco["fechas_sueltas"] or "[]")
        ).fetchone()

        conn.execute("""
            INSERT INTO francos_anulaciones_cerrados
              (francos_tomados_id, legajo, nombre, departamento, tipo, fecha_desde, fecha_hasta,
               fechas_sueltas, dias, motivo, usuario, anulado_en, periodo_origen_id, periodo_aplicado_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,NULL)
        """, (fid, franco["legajo"], franco["nombre"],
              (origen["departamento"] if origen else ""), franco["tipo"],
              franco["fecha_desde"] or "", franco["fecha_hasta"] or "", franco["fechas_sueltas"] or "[]",
              franco["dias"] or 0, motivo, usuario, ahora,
              (origen["periodo_id"] if origen else None)))
        conn.commit()
    return redirect(url_for("francos") + "?ok=anulado_cerrado")


@app.route("/francos/guardar-manual-semana", methods=["POST"])
def francos_guardar_manual_semana():
    if not _autenticado(): return jsonify({"error": "No autorizado"}), 401
    semana_num = int(request.form.get("semana_num", 0))
    mes        = request.form.get("mes", "").strip()
    if not semana_num or not mes:
        return jsonify({"error": "Semana y mes son requeridos"}), 400
    ahora = datetime.now().isoformat(timespec="seconds")
    guardados = 0
    bloqueados = []
    with _get_db() as conn:
        emps = conn.execute(
            "SELECT legajo, nombre, departamento FROM empleados_extra WHERE activo=1"
        ).fetchall()
        for emp in emps:
            key = f"dias_{emp['legajo']}"
            val = request.form.get(key, "").strip()
            dias = int(val) if val.isdigit() else 0
            existente = conn.execute(
                "SELECT cierre_francos_id FROM francos_semana_manual "
                "WHERE legajo=? AND semana_num=? AND mes=?",
                (emp["legajo"], semana_num, mes),
            ).fetchone()
            if existente and existente["cierre_francos_id"] is not None:
                bloqueados.append({
                    "legajo": str(emp["legajo"]),
                    "cierre_id": existente["cierre_francos_id"],
                })
                continue
            conn.execute("""
                INSERT INTO francos_semana_manual (legajo, nombre, departamento, semana_num, mes, dias, guardado_en)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(legajo, semana_num, mes) DO UPDATE SET dias=excluded.dias, guardado_en=excluded.guardado_en
            """, (emp["legajo"], emp["nombre"], emp["departamento"] or "", semana_num, mes, dias, ahora))
            guardados += 1
        conn.commit()
    return jsonify({"ok": True, "guardados": guardados, "bloqueados": bloqueados})


# ═══════════════════════════════════════════════
# CONFIGURACION DE EMAIL
# ═══════════════════════════════════════════════
_EMAIL_CFG_FILE = Path("config_email.json")


def _leer_email_cfg():
    try:
        return json.loads(_EMAIL_CFG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _guardar_email_cfg(data):
    _EMAIL_CFG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/configuracion/email", methods=["GET"])
def configuracion_email():
    if not _autenticado(): return _requiere_auth()
    cfg = _leer_email_cfg()
    with _get_db() as conn:
        supervisores = [dict(r) for r in conn.execute(
            "SELECT * FROM supervisores ORDER BY nombre"
        )]
    for s in supervisores:
        try:
            s["deptos_lista"] = json.loads(s["departamentos"])
        except Exception:
            s["deptos_lista"] = []
    departamentos = _departamentos_francos_disponibles(incluir_ocultos=True)
    return render_template("configuracion_email.html",
                           cfg=cfg,
                           supervisores=supervisores,
                           departamentos=departamentos)


@app.route("/configuracion/email/smtp", methods=["POST"])
def configuracion_email_smtp():
    if not _autenticado(): return _requiere_auth()
    cfg = _leer_email_cfg()
    cfg["smtp_user"]  = request.form.get("smtp_user",  "").strip()
    cfg["smtp_pass"]  = request.form.get("smtp_pass",  "").strip()
    cfg["smtp_host"]  = request.form.get("smtp_host",  "smtp.gmail.com").strip()
    cfg["smtp_port"]  = int(request.form.get("smtp_port", "587") or "587")
    cfg["smtp_from_name"] = request.form.get("smtp_from_name", "CM Horas Extras").strip()
    _guardar_email_cfg(cfg)
    return redirect(url_for("configuracion_email") + "?ok=smtp")


@app.route("/configuracion/supervisores/nuevo", methods=["POST"])
def supervisores_nuevo():
    if not _autenticado(): return _requiere_auth()
    nombre    = request.form.get("nombre", "").strip()
    email     = request.form.get("email",  "").strip()
    deptos    = request.form.getlist("departamentos")
    if not nombre or not email or not deptos:
        return redirect(url_for("configuracion_email") + "?error=sup_datos_requeridos")
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _get_db() as conn:
        conn.execute(
            "INSERT INTO supervisores (nombre, email, departamentos, cargado_en) VALUES (?,?,?,?)",
            (nombre, email, json.dumps(deptos), ahora)
        )
        conn.commit()
    return redirect(url_for("configuracion_email") + "?ok=supervisor")


@app.route("/configuracion/supervisores/eliminar/<int:sid>", methods=["POST"])
def supervisores_eliminar(sid):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        conn.execute("DELETE FROM supervisores WHERE id=?", (sid,))
        conn.commit()
    return redirect(url_for("configuracion_email"))


@app.route("/configuracion/feriados", methods=["GET"])
def configuracion_feriados():
    if not _autenticado(): return _requiere_auth()
    feriados = sorted(_cargar_feriados_config())
    return render_template("configuracion_feriados.html", feriados=feriados)


@app.route("/configuracion/feriados/nuevo", methods=["POST"])
def feriados_nuevo():
    if not _autenticado(): return _requiere_auth()
    fecha = request.form.get("fecha", "").strip()
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError:
        return redirect(url_for("configuracion_feriados") + "?error=fecha")
    feriados = {f.strftime("%Y-%m-%d") for f in _cargar_feriados_config()}
    feriados.add(fecha)
    _guardar_feriados_config(feriados)
    return redirect(url_for("configuracion_feriados") + "?ok=agregado")


@app.route("/configuracion/feriados/eliminar/<fecha>", methods=["POST"])
def feriados_eliminar(fecha):
    if not _autenticado(): return _requiere_auth()
    feriados = {f.strftime("%Y-%m-%d") for f in _cargar_feriados_config()}
    feriados.discard(fecha)
    _guardar_feriados_config(feriados)
    return redirect(url_for("configuracion_feriados") + "?ok=eliminado")


@app.route("/configuracion/supervisores/enviar/<int:sid>", methods=["POST"])
def supervisores_enviar(sid):
    if not _autenticado(): return _requiere_auth()
    import reporte_saldos_francos as rrf
    import traceback
    from datetime import datetime as _dt

    with _get_db() as conn:
        row = conn.execute("SELECT * FROM supervisores WHERE id=?", (sid,)).fetchone()
        sup = {k: row[k] for k in row.keys()} if row else None
    if not sup:
        return redirect(url_for("configuracion_email") + "?error=sin_datos")

    deptos = json.loads(sup["departamentos"])
    fecha_str  = _dt.now().strftime("%d/%m/%Y")
    fecha_arch = _dt.now().strftime("%Y%m%d")

    try:
        saldos = rrf._calcular_saldos()
    except Exception:
        app.logger.error("supervisores_enviar: _calcular_saldos\n" + traceback.format_exc())
        return redirect(url_for("configuracion_email") + "?error=sin_datos")

    por_depto = {}
    for s in saldos:
        por_depto.setdefault(s["departamento"], []).append(s)

    # Versión filtrada para envío externo (excluye no_enviar_reportes.json)
    por_depto_envio = rrf._filtrar_no_enviar(por_depto)

    reportes_dir = Path(__file__).parent / "reportes"
    reportes_dir.mkdir(exist_ok=True)

    adjuntos = []
    for dep in deptos:
        emps = next((v for k, v in por_depto_envio.items() if k.lower() == dep.lower()), None)
        if not emps:
            continue
        dep_limpio = dep.replace(" ", "_").replace("/", "-").upper()
        output_path = reportes_dir / f"reporte_francos_{dep_limpio}_{fecha_arch}.pdf"
        try:
            rrf._hacer_pdf(emps, dep.upper(), output_path, fecha_str)
            adjuntos.append(output_path)
        except Exception:
            app.logger.error(f"supervisores_enviar: _hacer_pdf {dep}\n" + traceback.format_exc())

    if not adjuntos:
        return redirect(url_for("configuracion_email") + "?error=sin_datos")

    cfg = rrf._leer_email_cfg()
    try:
        html_body = rrf._hacer_html_email(sup["nombre"], deptos, por_depto_envio, fecha_str)
        rrf._enviar_email(cfg, sup["nombre"], sup["email"], adjuntos, html_body, fecha_str)
    except Exception:
        app.logger.error("supervisores_enviar: _enviar_email\n" + traceback.format_exc())
        return redirect(url_for("configuracion_email") + "?error=email")

    return redirect(url_for("configuracion_email") + f"?ok=reporte&sup_nombre={sup['nombre']}")


# ═══════════════════════════════════════════════
# EMPLEADOS EXTRA (depts sin fichadas automáticas)
# ═══════════════════════════════════════════════
_ODS_PERSONAL = Path(r"I:\Desde Facturacion\PERSONAL POR SECTOR.ods")


def _leer_ods_personal():
    """Parsea el ODS y devuelve lista de {nombre, departamento}."""
    try:
        import pandas as pd
        df = pd.read_excel(str(_ODS_PERSONAL), engine="odf", header=None)
    except Exception as exc:
        return [], str(exc)

    resultado = []
    depto_actual = ""
    for _, row in df.iterrows():
        val = str(row[0]).strip() if len(row) > 0 and not (hasattr(row[0], '__class__') and row[0].__class__.__name__ == 'float') else ""
        import math
        raw = row[0]
        if raw is None or (isinstance(raw, float) and math.isnan(raw)):
            continue
        val = str(raw).strip()
        if not val:
            continue
        # Heurística: nombre en mayúsculas sin coma y sin espacios intermedios de 1 palabra
        # → es encabezado de sección si no tiene dos palabras separadas por espacio (apellido nombre)
        palabras = val.split()
        if len(palabras) <= 2 and val == val.upper() and not any(c.isdigit() for c in val):
            # probablemente es nombre de sector
            depto_actual = val.title()
        else:
            if depto_actual:
                resultado.append({"nombre": val.upper(), "departamento": depto_actual.upper()})
    return resultado, None


@app.route("/empleados/importar")
def empleados_importar():
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        existentes = {r["legajo"]: dict(r) for r in conn.execute(
            "SELECT * FROM empleados_extra WHERE activo=1"
        )}
    candidatos, error = _leer_ods_personal()
    # Solo mostramos sectores que no son los que ya están en sesion/periodos
    sectores_activos = {e["departamento"].upper() for e in _empleados_conocidos()}
    return render_template("empleados_importar.html",
                           candidatos=candidatos,
                           existentes=existentes,
                           error=error,
                           sectores_activos=sectores_activos)


@app.route("/empleados/importar/guardar", methods=["POST"])
def empleados_importar_guardar():
    if not _autenticado(): return _requiere_auth()
    nombres      = request.form.getlist("nombre")
    departamentos = request.form.getlist("departamento")
    legajos      = request.form.getlist("legajo")
    ahora        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    guardados    = 0
    with _get_db() as conn:
        for leg, nom, dep in zip(legajos, nombres, departamentos):
            leg = leg.strip()
            nom = nom.strip()
            dep = dep.strip()
            if not leg or not nom:
                continue
            conn.execute("""
                INSERT INTO empleados_extra (legajo, nombre, departamento, cargado_en)
                VALUES (?,?,?,?)
                ON CONFLICT(legajo) DO UPDATE
                SET nombre=excluded.nombre, departamento=excluded.departamento, activo=1
            """, (leg, nom, dep))
            guardados += 1
        conn.commit()
    return redirect(url_for("empleados_importar") + f"?ok={guardados}")


@app.route("/empleados/extra/eliminar/<legajo>", methods=["POST"])
def empleados_extra_eliminar(legajo):
    if not _autenticado(): return _requiere_auth()
    with _get_db() as conn:
        conn.execute("UPDATE empleados_extra SET activo=0 WHERE legajo=?", (legajo,))
        conn.commit()
    return redirect(url_for("empleados_importar"))


@app.route("/plus-vacacional")
def plus_vacacional():
    if not _autenticado(): return _requiere_auth()
    depto_param = _normalizar_departamento_web(request.args.get("depto", "redes"))
    if depto_param not in {"redes", "administracion"}:
        depto_param = "redes"
    depto_visible = _nombre_departamento_visible(depto_param)

    with _get_db() as conn:
        periodos_rows = conn.execute("""
            SELECT DISTINCT p.id, p.fecha_desde, p.fecha_hasta, p.cerrado_en
            FROM periodos p
            JOIN periodo_empleados pe ON pe.periodo_id = p.id
            WHERE COALESCE(p.estado, 'ACTIVO') = 'ACTIVO'
              AND pe.departamento = ?
            ORDER BY p.fecha_desde DESC
            LIMIT 6
        """, (depto_visible,)).fetchall()

        if not periodos_rows:
            return render_template("plus_vacacional.html",
                departamento=depto_visible, depto_param=depto_param,
                labels=[], empleados=[],
                generado=datetime.now().strftime("%d/%m/%Y %H:%M"))

        periodos_list = list(reversed([dict(p) for p in periodos_rows]))
        pids = [p["id"] for p in periodos_list]
        ph   = ",".join("?" * len(pids))

        rows = conn.execute(f"""
            SELECT legajo, nombre, periodo_id, ot50, ot100
            FROM periodo_empleados
            WHERE periodo_id IN ({ph})
              AND legajo NOT IN ('100','101')
            ORDER BY CAST(legajo AS INTEGER)
        """, pids).fetchall()

    emp_map = {}
    for r in rows:
        leg = str(r["legajo"])
        if leg not in emp_map:
            emp_map[leg] = {"nombre": r["nombre"], "por_pid": {}}
        td = _parse_hm(r["ot50"] or "0h") + _parse_hm(r["ot100"] or "0h")
        emp_map[leg]["por_pid"][r["periodo_id"]] = emp_map[leg]["por_pid"].get(r["periodo_id"], timedelta(0)) + td

    empleados = []
    for leg, emp in sorted(emp_map.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 9999):
        horas_td = [emp["por_pid"].get(p["id"], timedelta(0)) for p in periodos_list]
        total_td = sum(horas_td, timedelta(0))
        empleados.append({
            "legajo": leg,
            "nombre": emp["nombre"],
            "horas":  [_fmt_hhcmm(h) for h in horas_td],
            "total":  _fmt_hhcmm(total_td),
            "tiene_horas": total_td.total_seconds() > 0,
        })

    labels = [_label_mes(p["fecha_desde"]) for p in periodos_list]

    return render_template("plus_vacacional.html",
        departamento=depto_visible, depto_param=depto_param,
        labels=labels, empleados=empleados,
        generado=datetime.now().strftime("%d/%m/%Y %H:%M"))


# ═══════════════════════════════════════════════
# TAREAS PROGRAMADAS VÍA CRON EXTERNO
# ═══════════════════════════════════════════════

@app.route("/tareas/reporte-francos-semanal", methods=["GET", "POST"])
def tarea_reporte_francos_semanal():
    """Dispara reporte_saldos_francos.main() (PDF de saldos por depto +
    envío por email a supervisores) desde un servicio de cron externo
    (ej. cron-job.org), sin necesitar sesión de supervisor.

    Reemplaza a la Scheduled Task de PythonAnywhere, que requiere plan
    pago. Protegida por REPORTE_FRANCOS_TOKEN (variable de entorno) —
    sin configurar, la ruta rechaza cualquier pedido (falla cerrado)."""
    if not REPORTE_FRANCOS_TOKEN or request.args.get("token", "") != REPORTE_FRANCOS_TOKEN:
        return jsonify({"error": "No autorizado"}), 401

    import reporte_saldos_francos as rrf
    try:
        rrf.main()
        return jsonify({"ok": True})
    except SystemExit as se:
        # main() usa sys.exit(0)/(1) en algunos caminos (ej. sin datos) —
        # no es un error del servidor, solo cómo señaliza su propio final.
        return jsonify({"ok": se.code in (0, None), "exit_code": se.code})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ═══════════════════════════════════════════════
# ARRANQUE LOCAL
# ═══════════════════════════════════════════════
if __name__ == "__main__":
    try:
        ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        ip = "127.0.0.1"
    print(f"\n  Supervisor:  http://{ip}:5000")
    print(f"  Contraseña:  {SUPERVISOR_PASS}\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
