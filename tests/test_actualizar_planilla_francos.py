import json
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

import servidor


def _planilla(legajos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Julio 2026"
    for fila, legajo in enumerate(legajos, start=8):
        ws.cell(fila, 1).value = legajo
        ws.cell(fila, 3).value = 99
        ws.cell(fila, 6).value = 5
        ws.cell(fila, 9).value = f"=F{fila}+G{fila}-H{fila}"
        ws.cell(fila, 12).value = 30
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def test_actualiza_solo_francos_orig_y_tomados():
    """Desde julio 2026 cada depto tiene su propia planilla separada (antes
    era una sola combinada) -- _actualizar_planilla_francos toma UN
    departamento y UN cierre por vez."""
    cierre_ingenieros = {"saldo_anterior": json.dumps({
        "100": {"generados": 1, "tomados": 2},
        "101": {"generados": 0, "tomados": 1},
    })}
    resultado = servidor._actualizar_planilla_francos(
        _planilla([100, 101]), "2026-07", "ingenieros", cierre_ingenieros
    )
    ws = load_workbook(resultado, data_only=False)["Julio 2026"]

    assert (ws["G8"].value, ws["H8"].value) == (1, 2)
    assert (ws["G9"].value, ws["H9"].value) == (None, 1)
    assert ws["C8"].value == 99
    assert ws["F8"].value == 5
    assert ws["I8"].value == "=F8+G8-H8"
    assert ws["L8"].value == 30


def test_actualiza_guardias_por_separado_de_ingenieros():
    cierre_guardias = {"saldo_anterior": json.dumps({
        "113": {"generados": 3, "tomados": 0},
        "118": {"generados": 1, "tomados": 1},
    })}
    resultado = servidor._actualizar_planilla_francos(
        _planilla([113, 118]), "2026-07", "guardias", cierre_guardias
    )
    ws = load_workbook(resultado, data_only=False)["Julio 2026"]

    assert (ws["G8"].value, ws["H8"].value) == (3, None)
    assert (ws["G9"].value, ws["H9"].value) == (1, 1)


def test_rechaza_planilla_si_falta_un_legajo():
    cierre = {"saldo_anterior": json.dumps({"999": {"generados": 1, "tomados": 0}})}
    with pytest.raises(ValueError, match="999"):
        servidor._actualizar_planilla_francos(_planilla([100]), "2026-07", "ingenieros", cierre)


def test_ruta_requiere_solo_el_cierre_del_depto_elegido(monkeypatch, tmp_path):
    """Antes exigía cierre activo de Guardias E Ingenieros para actualizar
    cualquiera de las dos planillas (ambas venían combinadas en una sola).
    Ahora cada planilla es independiente: alcanza con el cierre del depto
    que se está actualizando."""
    db_file = tmp_path / "cierres_test.db"
    monkeypatch.setattr(servidor, "DATOS_DIR", tmp_path)
    monkeypatch.setattr(servidor, "DB_FILE", db_file)
    servidor._init_db()
    monkeypatch.setattr(servidor, "_autenticado", lambda: True)

    import sqlite3
    conn = sqlite3.connect(str(db_file))
    conn.execute(
        "INSERT INTO cierres_francos (cerrado_en, departamento, fecha_hasta, total_dias, estado, saldo_anterior) "
        "VALUES (?,?,?,?,?,?)",
        ("2026-07-08T10:00:00", "Ingenieros", "2026-07-08", 1, "ACTIVO",
         json.dumps({"100": {"generados": 1, "tomados": 0}})),
    )
    conn.commit()
    conn.close()

    servidor.app.config["TESTING"] = True
    with servidor.app.test_client() as c:
        data = {
            "mes": "2026-07",
            "departamento": "ingenieros",
            "planilla": (BytesIO(_planilla([100])), "planilla.xlsx"),
        }
        resp = c.post("/francos/planilla/actualizar", data=data, content_type="multipart/form-data")
        assert resp.status_code == 200, resp.get_data(as_text=True)

        # Guardias no tiene cierre este mes -- debe rechazar SU planilla,
        # sin que Ingenieros (ya probado arriba) le haga falta a Guardias.
        data_guardias = {
            "mes": "2026-07",
            "departamento": "guardias",
            "planilla": (BytesIO(_planilla([113])), "planilla.xlsx"),
        }
        resp_guardias = c.post("/francos/planilla/actualizar", data=data_guardias, content_type="multipart/form-data")
        assert resp_guardias.status_code == 409
        assert "Guardias" in resp_guardias.get_json()["error"]
